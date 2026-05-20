"""
Индекс страха и жадности для LINK / Chainlink — 2015 -> апрель 2026
===================================================================

Источник:
    Alternative.me Fear & Greed Index
    Бесплатно, без ключа

Важно:
    Alternative.me даёт общий Crypto Fear & Greed Index,
    а не отдельный индекс только для LINK / Chainlink.

    То есть это общий рыночный крипто-сентимент.
    Мы просто используем его как детерминанту для датасета LINK.

Данные доступны примерно с 1 февраля 2018 года.

Установка:
    pip install requests pandas openpyxl

Запуск:
    python fear_greed_link.py

Результат:
    fear_greed_link.xlsx
"""

import requests
import pandas as pd
import os


OUT = "fear_greed_link.xlsx"

UA = {
    "User-Agent": "Mozilla/5.0 (research)"
}


START_DATE = "2015-01-01"
END_DATE = "2026-04-30"


def get_fear_greed():
    print("Скачиваю Fear & Greed Index для LINK-датасета (Alternative.me)...")

    url = "https://api.alternative.me/fng/?limit=0&format=json&date_format=cn"

    r = requests.get(
        url,
        headers=UA,
        timeout=60
    )

    r.raise_for_status()

    data = r.json()["data"]

    rows = []

    for d in data:
        rows.append(
            {
                "Дата": d["timestamp"],
                "Индекс страха и жадности LINK": int(d["value"]),
                "Классификация": d["value_classification"],
            }
        )

    df = pd.DataFrame(rows)

    df["Дата"] = pd.to_datetime(
        df["Дата"],
        format="%Y-%m-%d"
    )

    df = df.sort_values("Дата")

    print(f"   Последние даты в API: {df['Дата'].tail(5).dt.date.tolist()}")

    before = len(df)

    df = df.drop_duplicates(
        subset="Дата",
        keep="last"
    )

    dupes = before - len(df)

    if dupes:
        print(f"   Удалено дубликатов: {dupes}")

    df = df.set_index("Дата").sort_index()

    full_range = pd.date_range(
        start=df.index.min(),
        end=df.index.max(),
        freq="D"
    )

    missing = full_range.difference(df.index)

    if len(missing) > 0:
        print(f"   Пропущенных дней в API: {len(missing)} -> заполняю forward fill")
        print(f"   Пропуски: {[d.date() for d in missing]}")

        df = df.reindex(full_range).ffill()

    df.index.name = "Дата"

    df = df.reset_index()

    # Изменение индекса день к дню
    df["fear_greed_link_change"] = (
        df["Индекс страха и жадности LINK"].diff()
    )

    # Экстремальный страх
    df["link_fear_dummy"] = (
        df["Индекс страха и жадности LINK"] <= 25
    ).astype(int)

    # Экстремальная жадность
    df["link_greed_dummy"] = (
        df["Индекс страха и жадности LINK"] >= 75
    ).astype(int)

    df = df[
        (df["Дата"] >= START_DATE)
        & (df["Дата"] <= END_DATE)
    ].copy()

    df = df.reset_index(drop=True)

    print(
        f"   OK {len(df)} строк | "
        f"{df['Дата'].min().date()} -> {df['Дата'].max().date()}"
    )

    print(
        f"   Экстремальный страх LINK-сентимент <=25: "
        f"{df['link_fear_dummy'].sum()} дней"
    )

    print(
        f"   Экстремальная жадность LINK-сентимент >=75: "
        f"{df['link_greed_dummy'].sum()} дней"
    )

    return df


def validate_dataset(df: pd.DataFrame):
    print("\nПроверка данных:")

    if df.empty:
        print("   Датасет пустой.")
        return

    print("   Строк:", len(df))
    print(
        "   Диапазон:",
        df["Дата"].min().date(),
        "->",
        df["Дата"].max().date()
    )

    full_range = pd.date_range(
        start=df["Дата"].min(),
        end=df["Дата"].max(),
        freq="D"
    )

    missing = full_range.difference(pd.DatetimeIndex(df["Дата"]))

    print("   Пропущенных дат внутри диапазона:", len(missing))

    if len(missing) > 0:
        print("   Первые пропуски:")
        print([d.date() for d in missing[:20]])

    print("\nПервые строки:")
    print(df.head(10))

    print("\nПоследние строки:")
    print(df.tail(10))

    print("\nОписание индекса:")
    print(df["Индекс страха и жадности LINK"].describe())

    print("\nПустые значения:")
    na_report = df.isna().sum()
    na_report = na_report[na_report > 0]

    if len(na_report) == 0:
        print("   Пустых значений нет.")
    else:
        print(na_report)


def save_excel(df, path):
    print(f"\nСохраняю Excel: {os.path.basename(path)}")

    df.to_excel(
        path,
        sheet_name="Fear & Greed LINK",
        index=False
    )

    from openpyxl import load_workbook
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from openpyxl.formatting.rule import ColorScaleRule

    wb = load_workbook(path)

    ws = wb["Fear & Greed LINK"]

    hdr_fill = PatternFill("solid", fgColor="1F3864")

    hdr_font = Font(
        color="FFFFFF",
        bold=True,
        size=10,
        name="Calibri"
    )

    body_font = Font(
        size=10,
        name="Calibri"
    )

    thin = Side(
        style="thin",
        color="DDDDDD"
    )

    brd = Border(
        left=thin,
        right=thin,
        top=thin,
        bottom=thin
    )

    for cell in ws[1]:
        cell.fill = hdr_fill
        cell.font = hdr_font
        cell.alignment = Alignment(
            horizontal="center",
            vertical="center"
        )
        cell.border = brd

    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.font = body_font
            cell.border = brd

    for col_cells in ws.columns:
        max_len = max(
            (
                len(str(c.value or ""))
                for c in col_cells
            ),
            default=8
        )

        ws.column_dimensions[
            get_column_letter(col_cells[0].column)
        ].width = min(max_len + 3, 40)

    # Цветовая шкала для самого индекса:
    # низкие значения красные, средние жёлтые, высокие зелёные.
    ws.conditional_formatting.add(
        f"B2:B{ws.max_row}",
        ColorScaleRule(
            start_type="min",
            start_color="F8696B",
            mid_type="num",
            mid_value=50,
            mid_color="FFEB84",
            end_type="max",
            end_color="63BE7B"
        )
    )

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    wb.save(path)

    print(f"   OK Строк: {len(df)}")


if __name__ == "__main__":
    print("=" * 60)
    print("  Fear & Greed Index LINK / Chainlink  2015 -> 2026")
    print("=" * 60)

    df = get_fear_greed()

    validate_dataset(df)

    out_path = os.path.join(
        os.path.dirname(os.path.abspath(__file__)),
        OUT
    )

    save_excel(df, out_path)

    print("\n" + "=" * 60)
    print(f"  Готово: {out_path}")
    print("=" * 60)