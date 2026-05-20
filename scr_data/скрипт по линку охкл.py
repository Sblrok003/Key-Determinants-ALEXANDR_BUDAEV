"""
Chainlink LINK OHLCV — CoinMarketCap + fallback — 2015 -> март 2026
===================================================================

Основной источник:
    CoinMarketCap public data-api

Fallback, если CMC пропустил конкретную дату:
    1. Yahoo Finance LINK-USD:
        Open / High / Low / Close
    2. CoinGecko:
        Объём торгов ($)
        Рыночная капитализация ($)

Зачем fallback:
    У CoinMarketCap по LINK есть пропущенная дата 2022-07-31.
    CMC не отдаёт её даже отдельным запросом.
    Поэтому код сам находит пропуски и восстанавливает их из альтернативных источников.

Что получаем:
    date
    LINK Open
    LINK High
    LINK Low
    LINK Close
    Объём торгов ($)
    Рыночная капитализация ($)

Также создаёт диагностические файлы:
    link_fallback_rows.csv
    link_missing_dates.csv, если что-то всё равно не восстановилось

Установка:
    pip install requests pandas openpyxl

Запуск:
    python link_ohlcv_cmc_fixed.py

Результат:
    link_ohlcv_2015_2026_march.xlsx
"""

from pathlib import Path
from datetime import datetime
import time
import random

import pandas as pd
import requests


START_DATE = "2015-01-01"
END_DATE = "2026-03-31"

OUTPUT_FILE = "link_ohlcv_2015_2026_march.xlsx"

CMC_ID_LINK = 1975
CMC_ID_USD = 2781

CMC_URL = "https://api.coinmarketcap.com/data-api/v3/cryptocurrency/historical"

YAHOO_CHART_URL = "https://query1.finance.yahoo.com/v8/finance/chart/LINK-USD"

COINGECKO_HISTORY_URL = "https://api.coingecko.com/api/v3/coins/chainlink/history"

BINANCE_KLINES_URL = "https://api.binance.com/api/v3/klines"

CHUNK_DAYS = 180
OVERLAP_DAYS = 3
MAX_RETRIES = 6

SLEEP_BETWEEN_CHUNKS = (1.2, 2.8)

# Если CMC пропускает дату, включаем восстановление через Yahoo/CoinGecko/Binance.
ENABLE_FALLBACK = True

# Если True, в итоговый Excel добавится колонка data_source.
# Для твоей общей таблицы обычно лучше оставить False, чтобы структура была как раньше.
INCLUDE_SOURCE_COLUMN_IN_EXCEL = False


def get_cmc_headers() -> dict:
    return {
        "User-Agent": (
            "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
            "AppleWebKit/537.36 (KHTML, like Gecko) "
            "Chrome/120.0 Safari/537.36"
        ),
        "Accept": "application/json, text/plain, */*",
        "Referer": "https://coinmarketcap.com/currencies/chainlink/historical-data/",
        "Origin": "https://coinmarketcap.com",
    }


def get_common_headers() -> dict:
    return {
        "User-Agent": (
            "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
            "AppleWebKit/537.36 (KHTML, like Gecko) "
            "Chrome/120.0 Safari/537.36"
        ),
        "Accept": "application/json, text/plain, */*",
    }


def to_unix_start(date_str: str) -> int:
    return int(pd.Timestamp(date_str, tz="UTC").timestamp())


def to_unix_end(date_str: str) -> int:
    ts = (
        pd.Timestamp(date_str, tz="UTC")
        + pd.Timedelta(days=1)
        - pd.Timedelta(milliseconds=1)
    )
    return int(ts.timestamp())


def make_chunks_with_overlap(start_date: str, end_date: str):
    start = pd.Timestamp(start_date)
    end = pd.Timestamp(end_date)

    current = start

    while current <= end:
        chunk_end = min(current + pd.Timedelta(days=CHUNK_DAYS - 1), end)

        real_start = max(current - pd.Timedelta(days=OVERLAP_DAYS), start)
        real_end = min(chunk_end + pd.Timedelta(days=OVERLAP_DAYS), end)

        yield real_start.strftime("%Y-%m-%d"), real_end.strftime("%Y-%m-%d")

        current = chunk_end + pd.Timedelta(days=1)


def extract_cmc_quote(q: dict) -> dict:
    quote = q.get("quote", {})

    if not isinstance(quote, dict):
        return {}

    if "USD" in quote and isinstance(quote["USD"], dict):
        return quote["USD"]

    if str(CMC_ID_USD) in quote and isinstance(quote[str(CMC_ID_USD)], dict):
        return quote[str(CMC_ID_USD)]

    return quote


def parse_cmc_quotes(data: dict) -> pd.DataFrame:
    root = data.get("data", {})
    quotes = root.get("quotes", [])

    rows = []

    for q in quotes:
        quote = extract_cmc_quote(q)

        date_raw = (
            q.get("timeOpen")
            or q.get("timeClose")
            or q.get("timestamp")
            or quote.get("timestamp")
        )

        if not date_raw:
            continue

        date = pd.to_datetime(date_raw, utc=True, errors="coerce")

        if pd.isna(date):
            continue

        date = date.tz_localize(None).normalize()

        rows.append(
            {
                "date": date,
                "LINK Open": quote.get("open"),
                "LINK High": quote.get("high"),
                "LINK Low": quote.get("low"),
                "LINK Close": quote.get("close"),
                "Объём торгов ($)": quote.get("volume"),
                "Рыночная капитализация ($)": quote.get("marketCap"),
                "data_source": "CoinMarketCap",
            }
        )

    return pd.DataFrame(rows)


def request_cmc(start_date: str, end_date: str) -> dict:
    params = {
        "id": CMC_ID_LINK,
        "convertId": CMC_ID_USD,
        "timeStart": to_unix_start(start_date),
        "timeEnd": to_unix_end(end_date),
        "interval": "1d",
    }

    response = requests.get(
        CMC_URL,
        params=params,
        headers=get_cmc_headers(),
        timeout=60,
    )

    response.raise_for_status()

    data = response.json()

    status = data.get("status", {})
    error_code = status.get("error_code", 0)

    if error_code not in (0, "0", None):
        raise RuntimeError(f"CMC status error: {status}")

    return data


def download_cmc_period(start_date: str, end_date: str) -> pd.DataFrame:
    last_error = None

    for attempt in range(1, MAX_RETRIES + 1):
        try:
            print(f"  {start_date} -> {end_date} | попытка {attempt}/{MAX_RETRIES}")

            data = request_cmc(start_date, end_date)
            df = parse_cmc_quotes(data)

            if df.empty:
                print("    Пустой кусок")
                return df

            print(
                f"    строк: {len(df)} | "
                f"{df['date'].min().date()} -> {df['date'].max().date()}"
            )

            return df

        except Exception as e:
            last_error = e
            wait = attempt * 4 + random.uniform(1, 5)
            print(f"    Ошибка: {e}")
            print(f"    Жду {wait:.1f} сек...")
            time.sleep(wait)

    raise RuntimeError(
        f"Не удалось скачать период {start_date} -> {end_date}: {last_error}"
    )


def clean_dataframe(df: pd.DataFrame) -> pd.DataFrame:
    if df.empty:
        return df

    numeric_cols = [
        "LINK Open",
        "LINK High",
        "LINK Low",
        "LINK Close",
        "Объём торгов ($)",
        "Рыночная капитализация ($)",
    ]

    for col in numeric_cols:
        if col in df.columns:
            df[col] = pd.to_numeric(df[col], errors="coerce")

    df["date"] = pd.to_datetime(df["date"]).dt.normalize()

    df = df.dropna(subset=["date", "LINK Close"])

    df = df[df["LINK Close"] > 0].copy()

    df = df.drop_duplicates(subset=["date"], keep="last")

    df = df.sort_values("date")

    df = df[
        (df["date"] >= pd.to_datetime(START_DATE))
        & (df["date"] <= pd.to_datetime(END_DATE))
    ].copy()

    df = df.reset_index(drop=True)

    return df


def get_missing_dates(df: pd.DataFrame) -> list[pd.Timestamp]:
    if df.empty:
        return []

    first_date = df["date"].min()
    last_required = pd.to_datetime(END_DATE)

    full_range = pd.date_range(first_date, last_required, freq="D")
    existing = pd.DatetimeIndex(df["date"])

    missing = full_range.difference(existing)

    return list(missing)


def download_yahoo_ohlc(target_date: pd.Timestamp) -> dict | None:
    """
    Берём OHLC с Yahoo Finance LINK-USD.
    Volume отсюда НЕ используем как долларовый объём.
    """

    start_ts = int(pd.Timestamp(target_date, tz="UTC").timestamp())
    end_ts = int((pd.Timestamp(target_date, tz="UTC") + pd.Timedelta(days=1)).timestamp())

    params = {
        "period1": start_ts,
        "period2": end_ts,
        "interval": "1d",
        "includePrePost": "false",
        "events": "history",
    }

    response = requests.get(
        YAHOO_CHART_URL,
        params=params,
        headers=get_common_headers(),
        timeout=60,
    )

    response.raise_for_status()

    data = response.json()

    chart = data.get("chart", {})
    error = chart.get("error")

    if error:
        raise RuntimeError(f"Yahoo error: {error}")

    results = chart.get("result", [])

    if not results:
        return None

    result = results[0]

    timestamps = result.get("timestamp", [])
    indicators = result.get("indicators", {})
    quotes = indicators.get("quote", [])

    if not timestamps or not quotes:
        return None

    q = quotes[0]

    opens = q.get("open", [])
    highs = q.get("high", [])
    lows = q.get("low", [])
    closes = q.get("close", [])

    if not opens or not highs or not lows or not closes:
        return None

    return {
        "LINK Open": opens[0],
        "LINK High": highs[0],
        "LINK Low": lows[0],
        "LINK Close": closes[0],
    }


def download_binance_ohlc(target_date: pd.Timestamp) -> dict | None:
    """
    Резервный fallback, если Yahoo не сработал.
    Берём LINKUSDT с Binance.

    Важно:
        Это уже данные одной биржи, а не агрегированный рынок.
        Используется только если Yahoo недоступен.
    """

    start_ms = int(pd.Timestamp(target_date, tz="UTC").timestamp() * 1000)
    end_ms = int(
        (pd.Timestamp(target_date, tz="UTC") + pd.Timedelta(days=1)).timestamp() * 1000
    )

    params = {
        "symbol": "LINKUSDT",
        "interval": "1d",
        "startTime": start_ms,
        "endTime": end_ms,
        "limit": 1,
    }

    response = requests.get(
        BINANCE_KLINES_URL,
        params=params,
        headers=get_common_headers(),
        timeout=60,
    )

    response.raise_for_status()

    data = response.json()

    if not data:
        return None

    k = data[0]

    return {
        "LINK Open": float(k[1]),
        "LINK High": float(k[2]),
        "LINK Low": float(k[3]),
        "LINK Close": float(k[4]),
        "binance_quote_volume_usdt": float(k[7]),
    }


def download_coingecko_market_data(target_date: pd.Timestamp) -> dict:
    """
    Берём market cap и volume с CoinGecko на конкретную дату.

    CoinGecko date format:
        dd-mm-yyyy
    """

    date_str = target_date.strftime("%d-%m-%Y")

    params = {
        "date": date_str,
        "localization": "false",
    }

    response = requests.get(
        COINGECKO_HISTORY_URL,
        params=params,
        headers=get_common_headers(),
        timeout=60,
    )

    response.raise_for_status()

    data = response.json()

    market_data = data.get("market_data", {})

    market_cap = (
        market_data.get("market_cap", {}).get("usd")
        if isinstance(market_data.get("market_cap", {}), dict)
        else None
    )

    total_volume = (
        market_data.get("total_volume", {}).get("usd")
        if isinstance(market_data.get("total_volume", {}), dict)
        else None
    )

    current_price = (
        market_data.get("current_price", {}).get("usd")
        if isinstance(market_data.get("current_price", {}), dict)
        else None
    )

    return {
        "Рыночная капитализация ($)": market_cap,
        "Объём торгов ($)": total_volume,
        "coingecko_close_snapshot": current_price,
    }


def download_fallback_row(target_date: pd.Timestamp) -> pd.DataFrame:
    """
    Восстанавливает одну пропущенную дату.

    Приоритет:
        OHLC:
            1. Yahoo LINK-USD
            2. Binance LINKUSDT

        Market Cap / Volume:
            1. CoinGecko
            2. Если CoinGecko volume не дал, можно использовать Binance quote volume как грубый резерв.
    """

    d = target_date.strftime("%Y-%m-%d")

    print(f"\nВосстанавливаю дату через fallback: {d}")

    ohlc = None
    ohlc_source = None
    binance_quote_volume = None

    try:
        ohlc = download_yahoo_ohlc(target_date)
        if ohlc is not None:
            ohlc_source = "Yahoo Finance LINK-USD"
            print("  OHLC взят из Yahoo Finance")
    except Exception as e:
        print(f"  Yahoo не сработал: {e}")

    if ohlc is None:
        try:
            binance_data = download_binance_ohlc(target_date)
            if binance_data is not None:
                binance_quote_volume = binance_data.pop("binance_quote_volume_usdt", None)
                ohlc = binance_data
                ohlc_source = "Binance LINKUSDT"
                print("  OHLC взят из Binance LINKUSDT")
        except Exception as e:
            print(f"  Binance не сработал: {e}")

    if ohlc is None:
        raise RuntimeError(f"Не удалось восстановить OHLC для {d}")

    market = {}

    try:
        market = download_coingecko_market_data(target_date)
        print("  Market cap / volume взяты из CoinGecko")
    except Exception as e:
        print(f"  CoinGecko не сработал: {e}")
        market = {
            "Рыночная капитализация ($)": None,
            "Объём торгов ($)": None,
            "coingecko_close_snapshot": None,
        }

    volume = market.get("Объём торгов ($)")

    if volume is None and binance_quote_volume is not None:
        volume = binance_quote_volume
        volume_source = "Binance quote volume USDT"
    else:
        volume_source = "CoinGecko total_volume USD"

    row = {
        "date": target_date.normalize(),
        "LINK Open": ohlc.get("LINK Open"),
        "LINK High": ohlc.get("LINK High"),
        "LINK Low": ohlc.get("LINK Low"),
        "LINK Close": ohlc.get("LINK Close"),
        "Объём торгов ($)": volume,
        "Рыночная капитализация ($)": market.get("Рыночная капитализация ($)"),
        "data_source": f"fallback: OHLC={ohlc_source}; volume={volume_source}; mcap=CoinGecko",
    }

    df = pd.DataFrame([row])
    df = clean_dataframe(df)

    if df.empty:
        raise RuntimeError(f"Fallback вернул пустую строку для {d}")

    return df


def repair_missing_dates(df: pd.DataFrame) -> pd.DataFrame:
    missing = get_missing_dates(df)

    if not missing:
        print("\nПропущенных дат нет.")
        return df

    print("\nНайдены пропущенные даты:")
    print([d.date() for d in missing])

    if not ENABLE_FALLBACK:
        missing_df = pd.DataFrame(
            {"missing_date": [d.strftime("%Y-%m-%d") for d in missing]}
        )
        missing_df.to_csv("link_missing_dates.csv", index=False, encoding="utf-8-sig")
        raise RuntimeError("Есть пропуски, а ENABLE_FALLBACK = False")

    recovered_parts = []

    for d in missing:
        try:
            recovered = download_fallback_row(d)

            if not recovered.empty:
                recovered_parts.append(recovered)

            time.sleep(random.uniform(1.0, 2.5))

        except Exception as e:
            print(f"  Не удалось восстановить {d.date()}: {e}")

    if recovered_parts:
        fallback_df = pd.concat(recovered_parts, ignore_index=True)
        fallback_df.to_csv("link_fallback_rows.csv", index=False, encoding="utf-8-sig")

        print("\nВосстановленные fallback-строки сохранены:")
        print(Path("link_fallback_rows.csv").resolve())

        df = pd.concat([df] + recovered_parts, ignore_index=True)
        df = clean_dataframe(df)

    final_missing = get_missing_dates(df)

    if final_missing:
        missing_df = pd.DataFrame(
            {"missing_date": [d.strftime("%Y-%m-%d") for d in final_missing]}
        )
        missing_df.to_csv("link_missing_dates.csv", index=False, encoding="utf-8-sig")

        print("\nПосле fallback всё равно остались пропуски:")
        print([d.date() for d in final_missing])
        print("Файл с пропусками сохранён: link_missing_dates.csv")

        raise RuntimeError(
            "Данные не являются непрерывными. "
            "Смотри link_missing_dates.csv"
        )

    print("\nВсе пропущенные даты восстановлены.")

    return df


def download_link_cmc() -> pd.DataFrame:
    print("Скачиваю LINK OHLCV с CoinMarketCap...")
    print("Период запроса:", START_DATE, "->", END_DATE)
    print("Куски скачиваются с перекрытием:", OVERLAP_DAYS, "дня")
    print("Fallback включён:", ENABLE_FALLBACK)
    print()

    parts = []

    for start_date, end_date in make_chunks_with_overlap(START_DATE, END_DATE):
        part = download_cmc_period(start_date, end_date)

        if not part.empty:
            parts.append(part)

        time.sleep(random.uniform(*SLEEP_BETWEEN_CHUNKS))

    if not parts:
        raise RuntimeError("CoinMarketCap не вернул данных.")

    df = pd.concat(parts, ignore_index=True)
    df = clean_dataframe(df)

    df = repair_missing_dates(df)

    return df


def validate_dataset(df: pd.DataFrame):
    print("\nПроверка данных:")

    print("Строк:", len(df))

    if df.empty:
        print("Датасет пустой.")
        return

    first_date = df["date"].min()
    last_date = df["date"].max()

    print("Диапазон:", first_date.date(), "->", last_date.date())

    print("\nПервые строки:")
    print(df.head(10))

    print("\nПоследние строки:")
    print(df.tail(10))

    full_range = pd.date_range(
        start=first_date,
        end=last_date,
        freq="D",
    )

    missing = full_range.difference(pd.DatetimeIndex(df["date"]))

    print("\nПропущенных дат внутри фактического диапазона:", len(missing))

    if len(missing) > 0:
        print("Первые пропуски:")
        print([d.date() for d in missing[:30]])

    bad_ohlc = df[
        (df["LINK High"] < df["LINK Low"])
        | (df["LINK High"] < df["LINK Open"])
        | (df["LINK High"] < df["LINK Close"])
        | (df["LINK Low"] > df["LINK Open"])
        | (df["LINK Low"] > df["LINK Close"])
    ]

    print("Ошибочных OHLC-строк:", len(bad_ohlc))

    if len(bad_ohlc) > 0:
        print("\nПлохие OHLC строки:")
        print(bad_ohlc.head(30))

    nan_report = df.isna().sum()
    nan_report = nan_report[nan_report > 0]

    print("\nПустые значения по колонкам:")

    if len(nan_report) == 0:
        print("Пустых значений нет.")
    else:
        print(nan_report)

    duplicates = df[df.duplicated(subset=["date"], keep=False)]

    print("\nДублей по date:", len(duplicates))

    if len(duplicates) > 0:
        print(duplicates.head(30))

    if "data_source" in df.columns:
        print("\nИсточники данных:")
        print(df["data_source"].value_counts())

        fallback_rows = df[df["data_source"].astype(str).str.contains("fallback", case=False, na=False)]

        if not fallback_rows.empty:
            print("\nFallback-строки:")
            print(fallback_rows)


def save_excel(df: pd.DataFrame, output_file: str):
    final_file = output_file

    df_to_save = df.copy()

    if not INCLUDE_SOURCE_COLUMN_IN_EXCEL and "data_source" in df_to_save.columns:
        df_to_save = df_to_save.drop(columns=["data_source"])

    try:
        df_to_save.to_excel(final_file, index=False)

    except PermissionError:
        print("\nФайл уже открыт или заблокирован:", output_file)
        stamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        final_file = output_file.replace(".xlsx", f"_{stamp}.xlsx")
        print("Сохраняю под новым именем:", final_file)
        df_to_save.to_excel(final_file, index=False)

    from openpyxl import load_workbook
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    wb = load_workbook(final_file)
    ws = wb.active
    ws.title = "LINK OHLCV"

    hdr_fill = PatternFill("solid", fgColor="1F3864")
    hdr_font = Font(color="FFFFFF", bold=True, size=10, name="Calibri")
    body_font = Font(size=10, name="Calibri")

    thin = Side(style="thin", color="DDDDDD")
    brd = Border(left=thin, right=thin, top=thin, bottom=thin)

    for cell in ws[1]:
        cell.fill = hdr_fill
        cell.font = hdr_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = brd

    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.font = body_font
            cell.border = brd

            if cell.column == 1:
                cell.number_format = "yyyy-mm-dd"
            elif isinstance(cell.value, float):
                cell.number_format = "#,##0.000000"

    for col_cells in ws.columns:
        max_len = max(
            len(str(c.value or ""))
            for c in col_cells
        )
        ws.column_dimensions[
            get_column_letter(col_cells[0].column)
        ].width = min(max_len + 3, 35)

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    wb.save(final_file)

    print("\nФайл сохранён:")
    print(Path(final_file).resolve())


def main():
    link = download_link_cmc()

    validate_dataset(link)

    save_excel(link, OUTPUT_FILE)


if __name__ == "__main__":
    main()