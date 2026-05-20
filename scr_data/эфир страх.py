"""
Индекс страха и жадности Ethereum — 2015 → апрель 2026
======================================================
Источник: Alternative.me (бесплатно, без ключа)

Важно:
Alternative.me даёт общий Crypto Fear & Greed Index,
а не отдельный индекс только для Ethereum.
Данные доступны с 1 февраля 2018 года.

Установка:
    pip install requests pandas openpyxl

Запуск:
    python fear_greed_eth.py

Результат: fear_greed_eth.xlsx
"""

import requests
import pandas as pd
import os

OUT = "fear_greed_eth.xlsx"
UA  = {"User-Agent": "Mozilla/5.0 (research)"}


def get_fear_greed():
    print("▶ Скачиваю Fear & Greed Index для Ethereum-датасета (Alternative.me)...")

    url = "https://api.alternative.me/fng/?limit=0&format=json&date_format=cn"

    r = requests.get(url, headers=UA, timeout=60)
    r.raise_for_status()

    data = r.json()["data"]

    rows = []

    for d in data:
        rows.append({
            "Дата": d["timestamp"],
            "Индекс страха и жадности Ethereum": int(d["value"]),
            "Классификация": d["value_classification"],
        })

    df = pd.DataFrame(rows)

    df["Дата"] = pd.to_datetime(df["Дата"], format="%Y-%m-%d")

    df = df.sort_values("Дата")

    print(f"   Последние даты в API: {df['Дата'].tail(5).dt.date.tolist()}")

    before = len(df)

    df = df.drop_duplicates(subset="Дата", keep="last")

    dupes = before - len(df)

    if dupes:
        print(f"   Удалено дубликатов: {dupes}")

    df = df.set_index("Дата").sort_index()

    full_range = pd.date_range(
        start=df.index.min(),
        end=df.index.max(),
        freq="D"
    )

    missing = full_range.difference(df.index)

    if len(missing) > 0:
        print(f"   Пропущенных дней в API: {len(missing)} → заполняю forward fill")
        print(f"   Пропуски: {[d.date() for d in missing]}")

        df = df.reindex(full_range).ffill()

    df.index.name = "Дата"

    df = df.reset_index()

    df["fear_greed_eth_change"] = df["Индекс страха и жадности Ethereum"].diff()

    df["eth_fear_dummy"] = (
        df["Индекс страха и жадности Ethereum"] <= 25
    ).astype(int)

    df["eth_greed_dummy"] = (
        df["Индекс страха и жадности Ethereum"] >= 75
    ).astype(int)

    df = df[
        (df["Дата"] >= "2015-01-01")
        & (df["Дата"] <= "2026-04-30")
    ]

    print(
        f"   ✓ {len(df)} строк | "
        f"{df['Дата'].min().date()} → {df['Дата'].max().date()}"
    )

    print(f"   Экстрем. страх Ethereum-сентимент (≤25): {df['eth_fear_dummy'].sum()} дней")
    print(f"   Экстрем. жадность Ethereum-сентимент (≥75): {df['eth_greed_dummy'].sum()} дней")

    return df


def save_excel(df, path):
    print(f"▶ Сохраняю Excel: {os.path.basename(path)}")

    df.to_excel(
        path,
        sheet_name="Fear & Greed ETH",
        index=False
    )

    from openpyxl import load_workbook
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from openpyxl.formatting.rule import ColorScaleRule

    wb = load_workbook(path)

    ws = wb["Fear & Greed ETH"]

    hdr_fill = PatternFill("solid", fgColor="1F3864")

    hdr_font = Font(
        color="FFFFFF",
        bold=True,
        size=10,
        name="Calibri"
    )

    body_font = Font(
        size=10,
        name="Calibri"
    )

    thin = Side(
        style="thin",
        color="DDDDDD"
    )

    brd = Border(
        left=thin,
        right=thin,
        top=thin,
        bottom=thin
    )

    for cell in ws[1]:
        cell.fill = hdr_fill
        cell.font = hdr_font
        cell.alignment = Alignment(
            horizontal="center",
            vertical="center"
        )
        cell.border = brd

    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.font = body_font
            cell.border = brd

    for col_cells in ws.columns:
        max_len = max(
            (
                len(str(c.value or ""))
                for c in col_cells
            ),
            default=8
        )

        ws.column_dimensions[
            get_column_letter(col_cells[0].column)
        ].width = min(max_len + 3, 40)

    ws.conditional_formatting.add(
        f"B2:B{ws.max_row}",
        ColorScaleRule(
            start_type="min",
            start_color="F8696B",
            mid_type="num",
            mid_value=50,
            mid_color="FFEB84",
            end_type="max",
            end_color="63BE7B"
        )
    )

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    wb.save(path)

    print(f"   ✓ Строк: {len(df)}")


if __name__ == "__main__":
    print("=" * 55)
    print("  Fear & Greed Index ETH  2015 → 2026")
    print("=" * 55)

    df = get_fear_greed()

    out_path = os.path.join(
        os.path.dirname(os.path.abspath(__file__)),
        OUT
    )

    save_excel(df, out_path)

    print("\n" + "=" * 55)
    print(f"  ✅ Готово: {out_path}")
    print("=" * 55)