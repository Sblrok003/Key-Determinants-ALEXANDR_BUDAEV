"""
BTC Market Cap, Total Market Cap, BTC Dominance — 2015 → 2026
==============================================================
Источник: CoinGecko Demo API (бесплатный ключ)
Получить ключ: https://www.coingecko.com/en/api → "Get Free API Key"

Установка:
    pip install requests pandas openpyxl

Запуск:
    python btc_mcap_dominance.py
"""

import requests
import pandas as pd
import numpy as np
import time
import os

# ── ВСТАВЬ СВОЙ КЛЮЧ ────────────────────────────────
CG_API_KEY = "CG-RBjGobW8Atkef2V1VFNf4mot"
# ────────────────────────────────────────────────────

START = "2015-01-01"
END   = "2026-04-30"
OUT   = "btc_mcap_dominance.xlsx"

BASE    = "https://api.coingecko.com/api/v3"
HEADERS = {"User-Agent": "Mozilla/5.0 (research)"}


def cg_get(path, params=None, retries=3):
    """GET к CoinGecko с повторами при 429."""
    url = BASE + path
    # Ключ передаётся как query-параметр (Demo API требует именно так)
    p = dict(params or {})
    p["x_cg_demo_api_key"] = CG_API_KEY
    for attempt in range(retries):
        try:
            r = requests.get(url, headers=HEADERS, params=p, timeout=60)
            if r.status_code == 429:
                wait = 15 * (attempt + 1)
                print(f"   429 Rate limit — жду {wait}с...")
                time.sleep(wait)
                continue
            r.raise_for_status()
            return r.json()
        except requests.HTTPError as e:
            print(f"   HTTP ошибка: {e}")
            if attempt == retries - 1:
                raise
        time.sleep(2)
    return None


# ─────────────────────────────────────────────────────
# 1. BTC Market Cap история
# ─────────────────────────────────────────────────────
def get_btc_mcap():
    print("▶ [1/3] BTC Market Cap (CoinGecko)...")
    data = cg_get("/coins/bitcoin/market_chart", {"vs_currency": "usd", "days": "max"})
    if not data:
        return pd.DataFrame()

    df = pd.DataFrame(data["market_caps"], columns=["ts", "BTC Market Cap (млрд $)"])
    df["date"] = pd.to_datetime(df["ts"], unit="ms").dt.normalize()
    df["BTC Market Cap (млрд $)"] = (df["BTC Market Cap (млрд $)"] / 1e9).round(3)
    df = df.set_index("date")[["BTC Market Cap (млрд $)"]].sort_index()
    print(f"   ✓ {len(df)} строк | {df.index.min().date()} → {df.index.max().date()}")
    return df


# ─────────────────────────────────────────────────────
# 2. Total Market Cap история
# ─────────────────────────────────────────────────────
def get_total_mcap():
    print("▶ [2/3] Total Market Cap (CoinGecko)...")
    time.sleep(2)
    data = cg_get("/global/market_cap_chart", {"vs_currency": "usd", "days": "max"})
    if not data:
        return pd.DataFrame()

    raw = data.get("market_cap_chart", {}).get("market_cap", [])
    if not raw:
        print("   ⚠ Нет данных в ответе")
        return pd.DataFrame()

    df = pd.DataFrame(raw, columns=["ts", "Total Market Cap (млрд $)"])
    df["date"] = pd.to_datetime(df["ts"], unit="ms").dt.normalize()
    df["Total Market Cap (млрд $)"] = (df["Total Market Cap (млрд $)"] / 1e9).round(3)
    df = df.set_index("date")[["Total Market Cap (млрд $)"]].sort_index()
    print(f"   ✓ {len(df)} строк | {df.index.min().date()} → {df.index.max().date()}")
    return df


# ─────────────────────────────────────────────────────
# 3. Объединяем и считаем Dominance
# ─────────────────────────────────────────────────────
def build_master(btc_df, total_df):
    master = btc_df.join(total_df, how="outer")
    if "Total Market Cap (млрд $)" in master.columns:
        master["BTC Dominance (%)"] = (
            master["BTC Market Cap (млрд $)"] /
            master["Total Market Cap (млрд $)"] * 100
        ).round(2)
    master = master.loc[START:END].sort_index()
    return master


# ─────────────────────────────────────────────────────
# 4. Сохранение в Excel
# ─────────────────────────────────────────────────────
def save_excel(df, path):
    print(f"▶ [3/3] Сохраняю Excel...")
    df.index.name = "Дата"
    df.to_excel(path, sheet_name="Market Cap")

    from openpyxl import load_workbook
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from openpyxl.formatting.rule import ColorScaleRule

    wb = load_workbook(path)
    ws = wb["Market Cap"]

    hdr_fill = PatternFill("solid", fgColor="1F3864")
    hdr_font = Font(color="FFFFFF", bold=True, size=10, name="Calibri")
    body_font = Font(size=10, name="Calibri")
    thin = Side(style="thin", color="DDDDDD")
    brd  = Border(left=thin, right=thin, top=thin, bottom=thin)

    for cell in ws[1]:
        cell.fill      = hdr_fill
        cell.font      = hdr_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border    = brd

    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.font   = body_font
            cell.border = brd
            if isinstance(cell.value, float):
                cell.number_format = "#,##0.00"

    for col_cells in ws.columns:
        max_len = max((len(str(c.value or "")) for c in col_cells), default=8)
        ws.column_dimensions[get_column_letter(col_cells[0].column)].width = min(max_len + 3, 30)

    # Градиент для Dominance
    for idx, cell in enumerate(ws[1], 1):
        if "Dominance" in str(cell.value):
            col_ltr = get_column_letter(idx)
            ws.conditional_formatting.add(
                f"{col_ltr}2:{col_ltr}{ws.max_row}",
                ColorScaleRule(
                    start_type="min", start_color="F8696B",
                    mid_type="num",   mid_value=50, mid_color="FFEB84",
                    end_type="max",   end_color="63BE7B"
                )
            )

    ws.freeze_panes = "B2"
    ws.auto_filter.ref = ws.dimensions
    wb.save(path)
    print(f"   ✓ Сохранено: {path}")
    print(f"   Строк: {len(df)} | Столбцов: {len(df.columns)}")


# ─────────────────────────────────────────────────────
# MAIN
# ─────────────────────────────────────────────────────
if __name__ == "__main__":
    print("=" * 55)
    print("  BTC Market Cap & Dominance  2015 → 2026")
    print("=" * 55)

    btc_df   = get_btc_mcap()
    total_df = get_total_mcap()
    master   = build_master(btc_df, total_df)

    print(f"\n  Период: {master.index.min().date()} → {master.index.max().date()}")
    for col in master.columns:
        n = master[col].notna().sum()
        print(f"  {col}: {n} значений")

    out_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), OUT)
    save_excel(master, out_path)

    print("\n" + "=" * 55)
    print(f"  ✅  Готово: {out_path}")
    print("=" * 55)