"""
Google Trends — поисковая активность "Ethereum" — 2015 → апрель 2026
====================================================================
Установка:
    pip install pytrends pandas openpyxl

Запуск:
    python google_trends_eth.py

Прогресс сохраняется в google_trends_eth_checkpoint.pkl после каждого куска.
При повторном запуске уже скачанные куски пропускаются.
Результат: google_trends_eth.xlsx
"""

import pandas as pd
import numpy as np
import time
import random
import os
import pickle
from datetime import datetime, timedelta
from pytrends.request import TrendReq

SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))

OUT        = os.path.join(SCRIPT_DIR, "google_trends_eth.xlsx")
CHECKPOINT = os.path.join(SCRIPT_DIR, "google_trends_eth_checkpoint.pkl")

KEYWORD  = "Ethereum"

START    = datetime(2015, 1, 1)
END      = datetime(2026, 4, 29)

CHUNK    = 180
OVERLAP  = 30
SLEEP    = 30


def human_sleep(base):
    time.sleep(base + random.uniform(0, base * 0.5))


def load_checkpoint():
    if os.path.exists(CHECKPOINT):
        with open(CHECKPOINT, "rb") as f:
            data = pickle.load(f)

        print(
            f"   ✅ Загружен чекпоинт: {len(data['chunks'])} кусков, "
            f"следующий старт: {data['next_start'].date()}"
        )

        return data["chunks"], data["next_start"]

    return [], START


def save_checkpoint(chunks, next_start):
    with open(CHECKPOINT, "wb") as f:
        pickle.dump(
            {
                "chunks": chunks,
                "next_start": next_start
            },
            f
        )


def download_chunks():
    pt = TrendReq(
        hl="en-US",
        tz=0,
        timeout=(15, 45)
    )

    chunks, current = load_checkpoint()

    while current < END:
        chunk_end = min(current + timedelta(days=CHUNK), END)

        tf = (
            f"{current.strftime('%Y-%m-%d')} "
            f"{chunk_end.strftime('%Y-%m-%d')}"
        )

        success = False

        for attempt in range(5):
            try:
                pt.build_payload(
                    [KEYWORD],
                    timeframe=tf,
                    geo=""
                )

                df = pt.interest_over_time()

                if df.empty:
                    print(f"   {tf}: пустой ответ")
                    break

                df = df[[KEYWORD]].copy()
                df.index = pd.to_datetime(df.index).tz_localize(None)

                chunks.append(df)

                print(
                    f"   ✓ {tf}: {len(df)} строк "
                    f"(всего кусков: {len(chunks)})"
                )

                success = True
                break

            except Exception as e:
                wait = 120 * (2 ** attempt) + random.uniform(10, 30)

                print(
                    f"   ❌ {tf} "
                    f"(попытка {attempt + 1}/5): ошибка — жду {wait:.0f}с"
                )
                print(f"      Детали: {e}")

                time.sleep(wait)

        if not success:
            print(f"   ⚠ Пропущен кусок {tf}")

        if chunk_end >= END:
            save_checkpoint(chunks, END)
            break

        next_start = chunk_end - timedelta(days=OVERLAP)

        save_checkpoint(chunks, next_start)

        current = next_start

        human_sleep(SLEEP)

    return chunks


def normalize_and_stitch(chunks):
    if not chunks:
        return pd.DataFrame()

    result = chunks[0].rename(columns={KEYWORD: "raw"}).copy()

    for chunk in chunks[1:]:
        chunk = chunk.rename(columns={KEYWORD: "raw"}).copy()

        overlap_idx = result.index.intersection(chunk.index)

        if len(overlap_idx) > 0:
            prev_mean = (
                result.loc[overlap_idx, "raw"]
                .replace(0, np.nan)
                .mean()
            )

            curr_mean = (
                chunk.loc[overlap_idx, "raw"]
                .replace(0, np.nan)
                .mean()
            )

            if (
                curr_mean is not None
                and prev_mean is not None
                and curr_mean > 0
                and prev_mean > 0
            ):
                scale = prev_mean / curr_mean
                chunk["raw"] = chunk["raw"] * scale

        new_dates = chunk.index.difference(result.index)

        result = pd.concat(
            [
                result,
                chunk.loc[new_dates]
            ]
        )

    result = result.sort_index()

    result.index.name = "Дата"

    result.columns = ["Поисковая активность Ethereum (Google Trends)"]

    col = "Поисковая активность Ethereum (Google Trends)"

    max_val = result[col].max()

    if max_val > 0:
        result[col] = (
            result[col] / max_val * 100
        ).round(2)

    return result


def save_excel(df, path):
    print(f"▶ Сохраняю Excel: {os.path.basename(path)}")

    df.to_excel(
        path,
        sheet_name="Google Trends"
    )

    from openpyxl import load_workbook
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from openpyxl.formatting.rule import ColorScaleRule

    wb = load_workbook(path)
    ws = wb["Google Trends"]

    hdr_fill = PatternFill("solid", fgColor="1F3864")
    hdr_font = Font(
        color="FFFFFF",
        bold=True,
        size=10,
        name="Calibri"
    )

    body_font = Font(
        size=10,
        name="Calibri"
    )

    thin = Side(
        style="thin",
        color="DDDDDD"
    )

    brd = Border(
        left=thin,
        right=thin,
        top=thin,
        bottom=thin
    )

    for cell in ws[1]:
        cell.fill = hdr_fill
        cell.font = hdr_font
        cell.alignment = Alignment(
            horizontal="center",
            vertical="center",
            wrap_text=True
        )
        cell.border = brd

    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.font = body_font
            cell.border = brd

            if isinstance(cell.value, float):
                cell.number_format = "#,##0.00"

    for col_cells in ws.columns:
        max_len = max(
            (
                len(str(c.value or ""))
                for c in col_cells
            ),
            default=8
        )

        ws.column_dimensions[
            get_column_letter(col_cells[0].column)
        ].width = min(max_len + 3, 45)

    ws.conditional_formatting.add(
        f"B2:B{ws.max_row}",
        ColorScaleRule(
            start_type="min",
            start_color="FFFFFF",
            end_type="max",
            end_color="FF6600"
        )
    )

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    wb.save(path)

    print(f"   ✓ Строк: {len(df)}")


if __name__ == "__main__":
    print("=" * 60)
    print("  Google Trends — Ethereum  2015 → апрель 2026")
    print("=" * 60)

    chunks = download_chunks()

    print(f"\n▶ Склеиваю {len(chunks)} кусков...")

    result = normalize_and_stitch(chunks)

    if result.empty:
        print("❌ Нет данных")
    else:
        result = result.loc[
            START.strftime("%Y-%m-%d"):
            END.strftime("%Y-%m-%d")
        ]

        print(
            f"   Итого: {len(result)} строк | "
            f"{result.index.min().date()} → {result.index.max().date()}"
        )

        save_excel(result, OUT)

        if os.path.exists(CHECKPOINT):
            os.remove(CHECKPOINT)
            print("   Чекпоинт удалён.")

        print("\n" + "=" * 60)
        print(f"  ✅ Готово: {OUT}")
        print("=" * 60)