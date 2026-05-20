"""
LINK exact network metrics без API-ключей
=========================================

Итоговые колонки:
    date
    link_tx_count
    link_fees_native
    link_active_addresses

Что считается:
    link_tx_count:
        уникальные Ethereum tx_hash, где был LINK Transfer event

    link_fees_native:
        сумма gasUsed * gasPrice / 1e18 по уникальным tx_hash
        Единица измерения: ETH

    link_active_addresses:
        уникальные from + to адреса LINK Transfer events за день

Важно:
    LINK — токен в Ethereum.
    Поэтому native fee для LINK-transfer платится в ETH.

Без ключей:
    Используется публичный Ethereum JSON-RPC.

Установка:
    pip install requests pandas openpyxl

Запуск:
    python линк_точные_комиссии.py

Результат:
    link_exact_network_metrics_no_key.xlsx

Кэш:
    link_exact_no_key_cache/

Если код остановился:
    Просто запусти ещё раз.
    Уже готовые дни и receipts он не будет качать заново.
"""

from pathlib import Path
from datetime import datetime
import time
import random
import json

import pandas as pd
import requests


# ============================================================
# 1. SETTINGS
# ============================================================

START_DATE = "2017-01-01"
END_DATE = "2026-04-30"

# До этой даты в таблице будут честные нули
LINK_START_DATE = "2017-09-16"

OUTPUT_FILE = "link_exact_network_metrics_no_key.xlsx"

# LINK contract on Ethereum
LINK_CONTRACT = "0x514910771af9ca656af840dff83e8264ecf986ca"

# ERC-20 Transfer(address,address,uint256)
TRANSFER_TOPIC0 = (
    "0xddf252ad1be2c89b69c2b068fc378daa952ba7f"
    "163c4a11628f55a4df523b3ef"
)

RPC_URLS = [
    "https://eth.llamarpc.com",
    "https://rpc.flashbots.net",
    "https://ethereum.publicnode.com",
]

SCRIPT_DIR = Path(__file__).resolve().parent

CACHE_DIR = SCRIPT_DIR / "link_exact_no_key_cache"
DAY_CACHE_DIR = CACHE_DIR / "days"
RECEIPT_CACHE_DIR = CACHE_DIR / "receipts_by_day"

BLOCK_TS_CACHE_FILE = CACHE_DIR / "block_timestamps_cache.json"
DAY_BLOCK_CACHE_FILE = CACHE_DIR / "day_block_boundaries_cache.json"

DAY_CACHE_DIR.mkdir(parents=True, exist_ok=True)
RECEIPT_CACHE_DIR.mkdir(parents=True, exist_ok=True)

MAX_RETRIES = 5

# Не делай слишком большим. Публичные RPC могут отваливаться.
BATCH_SIZE = 40

START_BLOCK_CHUNK = 3000
MIN_BLOCK_CHUNK = 100
MAX_BLOCK_CHUNK = 10000

SLEEP_RPC = 0.12


# ============================================================
# 2. BASIC HELPERS
# ============================================================

def sleep_random(base=SLEEP_RPC):
    time.sleep(base + random.uniform(0, base * 0.5))


def hex_to_int(x):
    if x is None:
        return None

    if isinstance(x, int):
        return x

    if isinstance(x, str):
        if x.startswith("0x"):
            return int(x, 16)
        return int(x)

    return None


def int_to_hex(n: int) -> str:
    return hex(int(n))


def to_unix(date_str: str) -> int:
    return int(pd.Timestamp(date_str, tz="UTC").timestamp())


def topic_to_address(topic: str) -> str:
    topic = str(topic).lower()

    if topic.startswith("0x"):
        topic = topic[2:]

    return "0x" + topic[-40:]


def day_range(start_date: str, end_date: str):
    for d in pd.date_range(start_date, end_date, freq="D"):
        yield d.strftime("%Y-%m-%d")


def next_day_str(date_str: str) -> str:
    return (
        pd.Timestamp(date_str)
        + pd.Timedelta(days=1)
    ).strftime("%Y-%m-%d")


def day_cache_path(date_str: str) -> Path:
    return DAY_CACHE_DIR / f"link_exact_{date_str}.csv"


def receipt_day_cache_path(date_str: str) -> Path:
    return RECEIPT_CACHE_DIR / f"receipts_{date_str}.json"


def tx_day_cache_path(date_str: str) -> Path:
    return RECEIPT_CACHE_DIR / f"txs_{date_str}.json"


def load_json_cache(path: Path) -> dict:
    if path.exists():
        with open(path, "r", encoding="utf-8") as f:
            return json.load(f)

    return {}


def save_json_cache(path: Path, obj: dict):
    tmp = path.with_suffix(".tmp")

    with open(tmp, "w", encoding="utf-8") as f:
        json.dump(obj, f, ensure_ascii=False)

    tmp.replace(path)


_block_ts_cache = load_json_cache(BLOCK_TS_CACHE_FILE)
_day_block_cache = load_json_cache(DAY_BLOCK_CACHE_FILE)
_latest_block_cache = None


# ============================================================
# 3. RPC
# ============================================================

def rpc_post(payload):
    """
    Пробует все публичные RPC.
    Если один отвалился, быстро идём к следующему.
    """
    last_error = None

    for attempt in range(1, MAX_RETRIES + 1):
        for url in RPC_URLS:
            try:
                r = requests.post(
                    url,
                    json=payload,
                    timeout=90,
                    headers={
                        "Content-Type": "application/json",
                        "User-Agent": "Mozilla/5.0 research",
                    },
                )

                if r.status_code != 200:
                    raise RuntimeError(f"HTTP {r.status_code}: {r.text[:500]}")

                data = r.json()

                if isinstance(data, dict) and "error" in data:
                    raise RuntimeError(str(data["error"]))

                return data

            except Exception as e:
                last_error = e
                print(f"   RPC ошибка {url}: {e}")

        wait = attempt * 2 + random.uniform(0.5, 2.0)
        print(f"   Все RPC не ответили. Жду {wait:.1f} сек... попытка {attempt}/{MAX_RETRIES}")
        time.sleep(wait)

    raise RuntimeError(f"Все RPC endpoints отказали: {last_error}")


def rpc_call(method: str, params: list):
    payload = {
        "jsonrpc": "2.0",
        "id": random.randint(1, 10_000_000),
        "method": method,
        "params": params,
    }

    data = rpc_post(payload)

    if "result" not in data:
        raise RuntimeError(f"Нет result в RPC ответе: {data}")

    return data["result"]


def rpc_batch(calls: list[tuple[str, list]]) -> list:
    if not calls:
        return []

    payload = []

    for i, (method, params) in enumerate(calls):
        payload.append(
            {
                "jsonrpc": "2.0",
                "id": i,
                "method": method,
                "params": params,
            }
        )

    data = rpc_post(payload)

    if not isinstance(data, list):
        raise RuntimeError(f"Batch ответ не list: {data}")

    by_id = {}

    for item in data:
        idx = item.get("id")

        if "error" in item:
            by_id[idx] = None
        else:
            by_id[idx] = item.get("result")

    return [by_id.get(i) for i in range(len(calls))]


# ============================================================
# 4. BLOCK BOUNDARIES BY DAY
# ============================================================

def get_latest_block_number() -> int:
    global _latest_block_cache

    if _latest_block_cache is None:
        _latest_block_cache = hex_to_int(
            rpc_call("eth_blockNumber", [])
        )

    return _latest_block_cache


def get_block_timestamp(block_number: int) -> int:
    key = str(int(block_number))

    if key in _block_ts_cache:
        return int(_block_ts_cache[key])

    block = rpc_call(
        "eth_getBlockByNumber",
        [int_to_hex(block_number), False]
    )

    if block is None:
        raise RuntimeError(f"Блок не найден: {block_number}")

    ts = hex_to_int(block["timestamp"])

    _block_ts_cache[key] = ts

    if len(_block_ts_cache) % 50 == 0:
        save_json_cache(BLOCK_TS_CACHE_FILE, _block_ts_cache)

    return ts


def first_block_at_or_after(timestamp: int) -> int:
    """
    Бинарный поиск первого блока с timestamp >= timestamp.
    """
    lo = 0
    hi = get_latest_block_number()

    while lo < hi:
        mid = (lo + hi) // 2
        mid_ts = get_block_timestamp(mid)

        if mid_ts < timestamp:
            lo = mid + 1
        else:
            hi = mid

        sleep_random(0.01)

    return lo


def get_start_block_for_day(date_str: str) -> int:
    """
    Кэшируем start-block каждого дня.
    """
    if date_str in _day_block_cache:
        return int(_day_block_cache[date_str]["start_block"])

    ts = to_unix(date_str)
    start_block = first_block_at_or_after(ts)

    _day_block_cache[date_str] = {
        "start_block": start_block,
    }

    save_json_cache(DAY_BLOCK_CACHE_FILE, _day_block_cache)
    save_json_cache(BLOCK_TS_CACHE_FILE, _block_ts_cache)

    return start_block


def get_day_block_range(date_str: str) -> tuple[int, int]:
    """
    Для дня D:
        start_block = первый блок >= D 00:00 UTC
        next_start_block = первый блок >= D+1 00:00 UTC
        end_block = next_start_block - 1

    Тогда все logs внутри [start_block; end_block] относятся к этому дню.
    Поэтому НЕ нужно скачивать timestamp каждого блока.
    """

    next_date = next_day_str(date_str)

    start_block = get_start_block_for_day(date_str)
    next_start_block = get_start_block_for_day(next_date)

    end_block = next_start_block - 1

    # Дописываем end_block в кэш текущего дня
    _day_block_cache[date_str]["end_block"] = end_block
    _day_block_cache[date_str]["next_start_block"] = next_start_block

    save_json_cache(DAY_BLOCK_CACHE_FILE, _day_block_cache)

    return start_block, end_block


# ============================================================
# 5. LINK TRANSFER LOGS
# ============================================================

def get_logs_block_range(from_block: int, to_block: int) -> list[dict]:
    params = [
        {
            "fromBlock": int_to_hex(from_block),
            "toBlock": int_to_hex(to_block),
            "address": LINK_CONTRACT,
            "topics": [TRANSFER_TOPIC0],
        }
    ]

    logs = rpc_call("eth_getLogs", params)

    if logs is None:
        return []

    return logs


def download_logs_adaptive(start_block: int, end_block: int) -> list[dict]:
    """
    eth_getLogs с адаптивным chunk.
    Если RPC ругается, уменьшаем диапазон.
    """
    all_logs = []

    current = start_block
    chunk = START_BLOCK_CHUNK

    while current <= end_block:
        to_block = min(current + chunk - 1, end_block)

        try:
            logs = get_logs_block_range(current, to_block)

            all_logs.extend(logs)

            print(
                f"   blocks {current} -> {to_block} | "
                f"logs={len(logs)} | total={len(all_logs)} | chunk={chunk}"
            )

            current = to_block + 1

            if len(logs) < 300 and chunk < MAX_BLOCK_CHUNK:
                chunk = min(chunk * 2, MAX_BLOCK_CHUNK)

            sleep_random()

        except Exception as e:
            print(f"   eth_getLogs ошибка {current}->{to_block}: {e}")

            if chunk > MIN_BLOCK_CHUNK:
                chunk = max(chunk // 2, MIN_BLOCK_CHUNK)
                print(f"   уменьшаю chunk до {chunk}")
                time.sleep(2.0)
            else:
                raise RuntimeError(
                    f"Не могу скачать logs даже при chunk={chunk}: "
                    f"{current}->{to_block}"
                )

    return all_logs


def logs_to_events(logs: list[dict]) -> pd.DataFrame:
    rows = []

    for lg in logs:
        if lg.get("removed") is True:
            continue

        topics = lg.get("topics", [])

        if len(topics) < 3:
            continue

        tx_hash = str(lg.get("transactionHash", "")).lower()
        log_index = hex_to_int(lg.get("logIndex"))

        if not tx_hash:
            continue

        rows.append(
            {
                "tx_hash": tx_hash,
                "log_index": log_index,
                "from": topic_to_address(topics[1]),
                "to": topic_to_address(topics[2]),
            }
        )

    if not rows:
        return pd.DataFrame(
            columns=[
                "tx_hash",
                "log_index",
                "from",
                "to",
            ]
        )

    df = pd.DataFrame(rows)

    df = df.drop_duplicates(
        subset=["tx_hash", "log_index"],
        keep="last"
    )

    return df


# ============================================================
# 6. RECEIPTS AND FEES
# ============================================================

def load_receipts_for_day(date_str: str) -> dict:
    path = receipt_day_cache_path(date_str)

    if path.exists():
        return load_json_cache(path)

    return {}


def save_receipts_for_day(date_str: str, receipts: dict):
    save_json_cache(receipt_day_cache_path(date_str), receipts)


def load_txs_for_day(date_str: str) -> dict:
    path = tx_day_cache_path(date_str)

    if path.exists():
        return load_json_cache(path)

    return {}


def save_txs_for_day(date_str: str, txs: dict):
    save_json_cache(tx_day_cache_path(date_str), txs)


def load_receipts_batch_for_day(date_str: str, tx_hashes: list[str]) -> dict[str, dict]:
    tx_hashes = list(dict.fromkeys([str(h).lower() for h in tx_hashes]))

    cached = load_receipts_for_day(date_str)

    out = {}
    need = []

    for h in tx_hashes:
        if h in cached:
            out[h] = cached[h]
        else:
            need.append(h)

    for i in range(0, len(need), BATCH_SIZE):
        batch = need[i:i + BATCH_SIZE]

        calls = [
            ("eth_getTransactionReceipt", [h])
            for h in batch
        ]

        results = rpc_batch(calls)

        for h, receipt in zip(batch, results):
            if receipt is not None:
                out[h] = receipt
                cached[h] = receipt

        save_receipts_for_day(date_str, cached)

        print(f"   receipts: {min(i + BATCH_SIZE, len(need))}/{len(need)}")

        sleep_random()

    return out


def load_transactions_batch_for_day(date_str: str, tx_hashes: list[str]) -> dict[str, dict]:
    tx_hashes = list(dict.fromkeys([str(h).lower() for h in tx_hashes]))

    cached = load_txs_for_day(date_str)

    out = {}
    need = []

    for h in tx_hashes:
        if h in cached:
            out[h] = cached[h]
        else:
            need.append(h)

    for i in range(0, len(need), BATCH_SIZE):
        batch = need[i:i + BATCH_SIZE]

        calls = [
            ("eth_getTransactionByHash", [h])
            for h in batch
        ]

        results = rpc_batch(calls)

        for h, tx in zip(batch, results):
            if tx is not None:
                out[h] = tx
                cached[h] = tx

        save_txs_for_day(date_str, cached)

        print(
            f"   tx gasPrice fallback: "
            f"{min(i + BATCH_SIZE, len(need))}/{len(need)}"
        )

        sleep_random()

    return out


def compute_fee_native_for_day(date_str: str, tx_hashes: list[str]) -> float:
    """
    Считает exact gas fee по уникальным транзакциям.

    fee_native = gasUsed * effectiveGasPrice / 1e18

    До EIP-1559 effectiveGasPrice может отсутствовать.
    Тогда берём gasPrice из eth_getTransactionByHash.
    """
    tx_hashes = list(dict.fromkeys([str(h).lower() for h in tx_hashes]))

    receipts = load_receipts_batch_for_day(date_str, tx_hashes)

    rows = []
    need_tx_fallback = []

    for h in tx_hashes:
        receipt = receipts.get(h)

        if receipt is None:
            continue

        gas_used = hex_to_int(receipt.get("gasUsed"))
        gas_price = hex_to_int(receipt.get("effectiveGasPrice"))

        if gas_used is None:
            continue

        if gas_price is None:
            need_tx_fallback.append(h)

        rows.append(
            {
                "tx_hash": h,
                "gas_used": gas_used,
                "gas_price": gas_price,
            }
        )

    if not rows:
        return 0.0

    fee_df = pd.DataFrame(rows)

    if need_tx_fallback:
        txs = load_transactions_batch_for_day(date_str, need_tx_fallback)

        gas_price_map = {}

        for h, tx in txs.items():
            gp = hex_to_int(tx.get("gasPrice"))

            if gp is not None:
                gas_price_map[h] = gp

        mask = fee_df["gas_price"].isna()

        fee_df.loc[mask, "gas_price"] = (
            fee_df.loc[mask, "tx_hash"].map(gas_price_map)
        )

    fee_df["gas_price"] = pd.to_numeric(
        fee_df["gas_price"],
        errors="coerce"
    )

    fee_df["gas_used"] = pd.to_numeric(
        fee_df["gas_used"],
        errors="coerce"
    )

    fee_df = fee_df.dropna(subset=["gas_used", "gas_price"])

    if fee_df.empty:
        return 0.0

    fee_df["fee_native"] = (
        fee_df["gas_used"].astype(float)
        * fee_df["gas_price"].astype(float)
        / 1e18
    )

    return float(fee_df["fee_native"].sum())


# ============================================================
# 7. DAILY PROCESSING
# ============================================================

def zero_day(date_str: str) -> pd.DataFrame:
    return pd.DataFrame(
        [
            {
                "date": pd.to_datetime(date_str),
                "link_tx_count": 0,
                "link_fees_native": 0.0,
                "link_active_addresses": 0,
            }
        ]
    )


def process_day(date_str: str) -> pd.DataFrame:
    cache_path = day_cache_path(date_str)

    if cache_path.exists():
        print(f"\nКэш дня найден: {date_str}")
        df = pd.read_csv(cache_path)
        df["date"] = pd.to_datetime(df["date"])
        return df

    print("\n" + "=" * 70)
    print(f"День {date_str}")
    print("=" * 70)

    if pd.Timestamp(date_str) < pd.Timestamp(LINK_START_DATE):
        print("До появления LINK. Нули.")
        df_day = zero_day(date_str)
        df_day.to_csv(cache_path, index=False, encoding="utf-8-sig")
        return df_day

    start_block, end_block = get_day_block_range(date_str)

    print(f"Блоки дня: {start_block} -> {end_block}")

    logs = download_logs_adaptive(start_block, end_block)

    print(f"LINK Transfer logs за день: {len(logs)}")

    if not logs:
        df_day = zero_day(date_str)
        df_day.to_csv(cache_path, index=False, encoding="utf-8-sig")
        return df_day

    events = logs_to_events(logs)

    if events.empty:
        df_day = zero_day(date_str)
        df_day.to_csv(cache_path, index=False, encoding="utf-8-sig")
        return df_day

    unique_tx = (
        events["tx_hash"]
        .dropna()
        .astype(str)
        .str.lower()
        .unique()
        .tolist()
    )

    link_tx_count = len(unique_tx)

    addresses = pd.concat(
        [
            events["from"],
            events["to"],
        ],
        ignore_index=True
    )

    link_active_addresses = (
        addresses
        .dropna()
        .astype(str)
        .str.lower()
        .nunique()
    )

    print(f"Уникальных tx: {link_tx_count}")
    print(f"Активных адресов: {link_active_addresses}")
    print("Считаю exact gas fees...")

    link_fees_native = compute_fee_native_for_day(date_str, unique_tx)

    df_day = pd.DataFrame(
        [
            {
                "date": pd.to_datetime(date_str),
                "link_tx_count": int(link_tx_count),
                "link_fees_native": float(link_fees_native),
                "link_active_addresses": int(link_active_addresses),
            }
        ]
    )

    df_day.to_csv(cache_path, index=False, encoding="utf-8-sig")

    print(
        f"OK {date_str}: "
        f"tx={link_tx_count}, "
        f"fees_native_ETH={link_fees_native:.10f}, "
        f"active={link_active_addresses}"
    )

    return df_day


# ============================================================
# 8. BUILD DATASET
# ============================================================

def build_dataset() -> pd.DataFrame:
    parts = []

    for date_str in day_range(START_DATE, END_DATE):
        part = process_day(date_str)
        parts.append(part)

        sleep_random(0.10)

    df = pd.concat(parts, ignore_index=True)

    df["date"] = pd.to_datetime(df["date"]).dt.normalize()

    df = (
        df
        .groupby("date", as_index=False)
        .agg(
            {
                "link_tx_count": "sum",
                "link_fees_native": "sum",
                "link_active_addresses": "max",
            }
        )
    )

    calendar = pd.DataFrame(
        {
            "date": pd.date_range(START_DATE, END_DATE, freq="D")
        }
    )

    df = calendar.merge(df, on="date", how="left")

    df["link_tx_count"] = df["link_tx_count"].fillna(0).astype(int)
    df["link_fees_native"] = df["link_fees_native"].fillna(0.0)
    df["link_active_addresses"] = df["link_active_addresses"].fillna(0).astype(int)

    return df


# ============================================================
# 9. VALIDATION
# ============================================================

def validate(df: pd.DataFrame):
    print("\n" + "=" * 80)
    print("Проверка")
    print("=" * 80)

    print("Строк:", len(df))

    print(
        "Диапазон:",
        df["date"].min().date(),
        "->",
        df["date"].max().date()
    )

    full_range = pd.date_range(START_DATE, END_DATE, freq="D")

    missing = full_range.difference(pd.DatetimeIndex(df["date"]))

    print("Пропущенных дат:", len(missing))

    if len(missing) > 0:
        print([d.date() for d in missing[:20]])
        raise RuntimeError("Есть пропущенные даты.")

    duplicates = df[df.duplicated(subset=["date"], keep=False)]

    print("Дублей:", len(duplicates))

    if len(duplicates) > 0:
        raise RuntimeError("Есть дубли.")

    print("\nПустые значения:")
    na = df.isna().sum()
    na = na[na > 0]

    if len(na) == 0:
        print("Пустых значений нет.")
    else:
        print(na)

    print("\nНенулевых значений:")
    print("link_tx_count:", (df["link_tx_count"] != 0).sum())
    print("link_fees_native:", (df["link_fees_native"] != 0).sum())
    print("link_active_addresses:", (df["link_active_addresses"] != 0).sum())

    print("\nПервые строки:")
    print(df.head(10))

    print("\nПоследние строки:")
    print(df.tail(10))

    print("\nОписание:")
    print(
        df[
            [
                "link_tx_count",
                "link_fees_native",
                "link_active_addresses",
            ]
        ].describe()
    )


# ============================================================
# 10. SAVE EXCEL
# ============================================================

def save_excel(df: pd.DataFrame, path: str):
    final_path = Path(path)

    try:
        df.to_excel(final_path, index=False)

    except PermissionError:
        stamp = datetime.now().strftime("%Y%m%d_%H%M%S")

        final_path = final_path.with_name(
            final_path.stem + f"_{stamp}" + final_path.suffix
        )

        print(f"Файл был открыт. Сохраняю как: {final_path}")

        df.to_excel(final_path, index=False)

    from openpyxl import load_workbook
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    wb = load_workbook(final_path)
    ws = wb.active
    ws.title = "LINK Exact No Key"

    hdr_fill = PatternFill("solid", fgColor="1F3864")
    hdr_font = Font(color="FFFFFF", bold=True, size=10, name="Calibri")
    body_font = Font(size=10, name="Calibri")

    thin = Side(style="thin", color="DDDDDD")
    brd = Border(left=thin, right=thin, top=thin, bottom=thin)

    for cell in ws[1]:
        cell.fill = hdr_fill
        cell.font = hdr_font
        cell.alignment = Alignment(
            horizontal="center",
            vertical="center",
            wrap_text=True
        )
        cell.border = brd

    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.font = body_font
            cell.border = brd

            if cell.column == 1:
                cell.number_format = "yyyy-mm-dd"

            elif isinstance(cell.value, float):
                cell.number_format = "#,##0.0000000000"

    for col_cells in ws.columns:
        max_len = max(
            len(str(c.value or ""))
            for c in col_cells
        )

        ws.column_dimensions[
            get_column_letter(col_cells[0].column)
        ].width = min(max_len + 3, 40)

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    wb.save(final_path)

    print("\nФайл сохранён:")
    print(final_path.resolve())


# ============================================================
# 11. MAIN
# ============================================================

def main():
    print("=" * 80)
    print("LINK exact daily metrics без API-ключей")
    print("=" * 80)

    df = build_dataset()

    validate(df)

    save_excel(df, OUTPUT_FILE)

    print("\nГОТОВО")


if __name__ == "__main__":
    main()