"""
Ethereum OHLCV — CoinMarketCap — 2015 -> март 2026
==================================================

Источник:
    CoinMarketCap public data-api

Что получаем:
    date
    ETH Open
    ETH High
    ETH Low
    ETH Close
    Объём торгов ($)
    Рыночная капитализация ($)

Важно:
    ETH не торговался с 2015-01-01.
    Ethereum mainnet запущен 30 июля 2015.
    Поэтому нормальные рыночные данные начинаются примерно с августа 2015.

Установка:
    pip install requests pandas openpyxl

Запуск:
    python eth_ohlcv_cmc.py

Результат:
    eth_ohlcv_2015_2026_march.xlsx
"""

from pathlib import Path
from datetime import datetime
import time
import random

import pandas as pd
import requests


START_DATE = "2015-01-01"
END_DATE = "2026-03-31"

OUTPUT_FILE = "eth_ohlcv_2015_2026_march.xlsx"

# CoinMarketCap IDs:
# Ethereum = 1027
# USD = 2781
CMC_ID_ETH = 1027
CMC_ID_USD = 2781

# Берём кусками, чтобы CMC не резал ответ.
CHUNK_DAYS = 330

URL = "https://api.coinmarketcap.com/data-api/v3/cryptocurrency/historical"


def to_unix_start(date_str: str) -> int:
    """
    Начало даты YYYY-MM-DD в UTC Unix timestamp.
    """
    return int(pd.Timestamp(date_str, tz="UTC").timestamp())


def to_unix_end(date_str: str) -> int:
    """
    Конец даты YYYY-MM-DD в UTC Unix timestamp.
    """
    ts = (
        pd.Timestamp(date_str, tz="UTC")
        + pd.Timedelta(days=1)
        - pd.Timedelta(seconds=1)
    )
    return int(ts.timestamp())


def make_chunks(start_date: str, end_date: str, chunk_days: int):
    """
    Делит период на куски.
    """
    start = pd.Timestamp(start_date)
    end = pd.Timestamp(end_date)

    current = start

    while current <= end:
        chunk_end = min(current + pd.Timedelta(days=chunk_days - 1), end)
        yield current.strftime("%Y-%m-%d"), chunk_end.strftime("%Y-%m-%d")
        current = chunk_end + pd.Timedelta(days=1)


def parse_cmc_quotes(data: dict) -> pd.DataFrame:
    """
    Разбирает JSON CoinMarketCap public data-api.
    """

    root = data.get("data", {})
    quotes = root.get("quotes", [])

    rows = []

    for q in quotes:
        quote = q.get("quote", {})

        # Обычно quote лежит сразу как dict:
        # quote = {open, high, low, close, volume, marketCap, timestamp}
        #
        # Но на случай другой структуры оставим проверку.
        if isinstance(quote, dict) and "USD" in quote:
            quote = quote["USD"]

        date_raw = (
            q.get("timeOpen")
            or q.get("timeClose")
            or q.get("timestamp")
            or quote.get("timestamp")
        )

        if not date_raw:
            continue

        rows.append(
            {
                "date": pd.to_datetime(date_raw, utc=True)
                .tz_localize(None)
                .normalize(),

                "ETH Open": quote.get("open"),
                "ETH High": quote.get("high"),
                "ETH Low": quote.get("low"),
                "ETH Close": quote.get("close"),
                "Объём торгов ($)": quote.get("volume"),
                "Рыночная капитализация ($)": quote.get("marketCap"),
            }
        )

    return pd.DataFrame(rows)


def download_chunk(start_date: str, end_date: str) -> pd.DataFrame:
    """
    Скачивает один кусок ETH OHLCV с CoinMarketCap.
    """

    params = {
        "id": CMC_ID_ETH,
        "convertId": CMC_ID_USD,
        "timeStart": to_unix_start(start_date),
        "timeEnd": to_unix_end(end_date),
        "interval": "1d",
    }

    headers = {
        "User-Agent": (
            "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
            "AppleWebKit/537.36 (KHTML, like Gecko) "
            "Chrome/120.0 Safari/537.36"
        ),
        "Accept": "application/json, text/plain, */*",
        "Referer": "https://coinmarketcap.com/currencies/ethereum/historical-data/",
        "Origin": "https://coinmarketcap.com",
    }

    last_error = None

    for attempt in range(1, 6):
        try:
            print(f"  {start_date} -> {end_date} | попытка {attempt}/5")

            response = requests.get(
                URL,
                params=params,
                headers=headers,
                timeout=60,
            )

            response.raise_for_status()

            data = response.json()

            status = data.get("status", {})
            error_code = status.get("error_code", 0)

            if error_code not in (0, "0", None):
                raise RuntimeError(status)

            df = parse_cmc_quotes(data)

            if df.empty:
                print("    Пустой кусок")
                return df

            print(
                f"    ✓ строк: {len(df)} | "
                f"{df['date'].min().date()} -> {df['date'].max().date()}"
            )

            return df

        except Exception as e:
            last_error = e
            wait = attempt * 5 + random.uniform(1, 4)
            print(f"    Ошибка: {e}")
            print(f"    Жду {wait:.1f} сек...")
            time.sleep(wait)

    raise RuntimeError(f"Не удалось скачать кусок {start_date} -> {end_date}: {last_error}")


def download_eth_cmc() -> pd.DataFrame:
    """
    Скачивает весь период ETH OHLCV с CoinMarketCap.
    """

    print("Скачиваю ETH OHLCV с CoinMarketCap...")
    print("Период запроса:", START_DATE, "->", END_DATE)
    print("Важно: до августа 2015 нормальных торговых данных ETH быть не должно.\n")

    parts = []

    for start_date, end_date in make_chunks(START_DATE, END_DATE, CHUNK_DAYS):
        part = download_chunk(start_date, end_date)

        if not part.empty:
            parts.append(part)

        time.sleep(random.uniform(1.0, 2.5))

    if not parts:
        raise RuntimeError("CoinMarketCap не вернул данных.")

    df = pd.concat(parts, ignore_index=True)

    df = clean_dataframe(df)

    return df


def clean_dataframe(df: pd.DataFrame) -> pd.DataFrame:
    """
    Чистит и проверяет датасет.
    """

    numeric_cols = [
        "ETH Open",
        "ETH High",
        "ETH Low",
        "ETH Close",
        "Объём торгов ($)",
        "Рыночная капитализация ($)",
    ]

    for col in numeric_cols:
        df[col] = pd.to_numeric(df[col], errors="coerce")

    df = df.dropna(subset=["date", "ETH Close"])

    df = df[df["ETH Close"] > 0].copy()

    df = df.drop_duplicates(subset=["date"], keep="last")

    df = df.sort_values("date")

    df = df[
        (df["date"] >= pd.to_datetime(START_DATE))
        & (df["date"] <= pd.to_datetime(END_DATE))
    ].copy()

    df = df.reset_index(drop=True)

    return df


def validate_dataset(df: pd.DataFrame):
    """
    Печатает диагностику качества.
    """

    print("\nПроверка данных:")

    print("Строк:", len(df))

    if df.empty:
        print("Датасет пустой.")
        return

    first_date = df["date"].min()
    last_date = df["date"].max()

    print("Диапазон:", first_date, "->", last_date)

    print("\nПервые строки:")
    print(df.head(10))

    print("\nПоследние строки:")
    print(df.tail(10))

    full_range = pd.date_range(
        start=first_date,
        end=last_date,
        freq="D"
    )

    missing = full_range.difference(pd.DatetimeIndex(df["date"]))

    print("\nПропущенных дат внутри диапазона:", len(missing))

    if len(missing) > 0:
        print("Первые пропуски:")
        print([d.date() for d in missing[:20]])

    bad_ohlc = df[
        (df["ETH High"] < df["ETH Low"])
        | (df["ETH High"] < df["ETH Open"])
        | (df["ETH High"] < df["ETH Close"])
        | (df["ETH Low"] > df["ETH Open"])
        | (df["ETH Low"] > df["ETH Close"])
    ]

    print("Ошибочных OHLC-строк:", len(bad_ohlc))

    if len(bad_ohlc) > 0:
        print(bad_ohlc.head(20))


def save_excel(df: pd.DataFrame, output_file: str):
    """
    Сохраняет Excel. Если файл открыт, сохраняет копию с timestamp.
    """

    final_file = output_file

    try:
        df.to_excel(final_file, index=False)

    except PermissionError:
        print("\nФайл уже открыт или заблокирован:", output_file)
        stamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        final_file = output_file.replace(".xlsx", f"_{stamp}.xlsx")
        print("Сохраняю под новым именем:", final_file)
        df.to_excel(final_file, index=False)

    from openpyxl import load_workbook
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    wb = load_workbook(final_file)
    ws = wb.active
    ws.title = "ETH OHLCV"

    hdr_fill = PatternFill("solid", fgColor="1F3864")
    hdr_font = Font(color="FFFFFF", bold=True, size=10, name="Calibri")
    body_font = Font(size=10, name="Calibri")

    thin = Side(style="thin", color="DDDDDD")
    brd = Border(left=thin, right=thin, top=thin, bottom=thin)

    for cell in ws[1]:
        cell.fill = hdr_fill
        cell.font = hdr_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = brd

    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.font = body_font
            cell.border = brd

            if cell.column == 1:
                cell.number_format = "yyyy-mm-dd"
            elif isinstance(cell.value, float):
                cell.number_format = "#,##0.000000"

    for col_cells in ws.columns:
        max_len = max(
            len(str(c.value or ""))
            for c in col_cells
        )
        ws.column_dimensions[
            get_column_letter(col_cells[0].column)
        ].width = min(max_len + 3, 35)

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    wb.save(final_file)

    print("\nФайл сохранён:")
    print(Path(final_file).resolve())


def main():
    eth = download_eth_cmc()

    validate_dataset(eth)

    save_excel(eth, OUTPUT_FILE)


if __name__ == "__main__":
    main()