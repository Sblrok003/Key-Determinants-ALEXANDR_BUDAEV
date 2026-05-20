"""
Google Trends — поисковая активность "Chainlink" — 2015 → апрель 2026
=====================================================================

Установка:
    pip install pytrends pandas openpyxl

Запуск:
    python google_trends_link.py

Прогресс сохраняется в google_trends_link_checkpoint.pkl после каждого куска.
При повторном запуске уже скачанные куски пропускаются.

Результат:
    google_trends_link.xlsx

Важно:
    Для Google Trends лучше искать "Chainlink", а не "LINK",
    потому что LINK — обычное английское слово и будет много шума.
"""

import pandas as pd
import numpy as np
import time
import random
import os
import pickle
from datetime import datetime, timedelta
from pytrends.request import TrendReq


SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))

OUT = os.path.join(SCRIPT_DIR, "google_trends_link.xlsx")
CHECKPOINT = os.path.join(SCRIPT_DIR, "google_trends_link_checkpoint.pkl")

KEYWORD = "Chainlink"

START = datetime(2015, 1, 1)
END = datetime(2026, 4, 29)

CHUNK = 180
OVERLAP = 30
SLEEP = 30


def human_sleep(base):
    time.sleep(base + random.uniform(0, base * 0.5))


def load_checkpoint():
    if os.path.exists(CHECKPOINT):
        with open(CHECKPOINT, "rb") as f:
            data = pickle.load(f)

        print(
            f"   Загружен чекпоинт: {len(data['chunks'])} кусков, "
            f"следующий старт: {data['next_start'].date()}"
        )

        return data["chunks"], data["next_start"]

    return [], START


def save_checkpoint(chunks, next_start):
    with open(CHECKPOINT, "wb") as f:
        pickle.dump(
            {
                "chunks": chunks,
                "next_start": next_start,
            },
            f,
        )


def download_chunks():
    pt = TrendReq(
        hl="en-US",
        tz=0,
        timeout=(15, 45),
    )

    chunks, current = load_checkpoint()

    while current < END:
        chunk_end = min(current + timedelta(days=CHUNK), END)

        tf = (
            f"{current.strftime('%Y-%m-%d')} "
            f"{chunk_end.strftime('%Y-%m-%d')}"
        )

        success = False

        for attempt in range(5):
            try:
                pt.build_payload(
                    [KEYWORD],
                    timeframe=tf,
                    geo="",
                )

                df = pt.interest_over_time()

                if df.empty:
                    print(f"   {tf}: пустой ответ")
                    break

                df = df[[KEYWORD]].copy()
                df.index = pd.to_datetime(df.index).tz_localize(None)

                chunks.append(df)

                print(
                    f"   OK {tf}: {len(df)} строк "
                    f"(всего кусков: {len(chunks)})"
                )

                success = True
                break

            except Exception as e:
                wait = 120 * (2 ** attempt) + random.uniform(10, 30)

                print(
                    f"   Ошибка {tf} "
                    f"(попытка {attempt + 1}/5): жду {wait:.0f} сек"
                )
                print(f"      Детали: {e}")

                time.sleep(wait)

        if not success:
            print(f"   Пропущен кусок {tf}")

        if chunk_end >= END:
            save_checkpoint(chunks, END)
            break

        next_start = chunk_end - timedelta(days=OVERLAP)

        save_checkpoint(chunks, next_start)

        current = next_start

        human_sleep(SLEEP)

    return chunks


def normalize_and_stitch(chunks):
    if not chunks:
        return pd.DataFrame()

    result = chunks[0].rename(columns={KEYWORD: "raw"}).copy()

    for chunk in chunks[1:]:
        chunk = chunk.rename(columns={KEYWORD: "raw"}).copy()

        overlap_idx = result.index.intersection(chunk.index)

        if len(overlap_idx) > 0:
            prev_mean = (
                result.loc[overlap_idx, "raw"]
                .replace(0, np.nan)
                .mean()
            )

            curr_mean = (
                chunk.loc[overlap_idx, "raw"]
                .replace(0, np.nan)
                .mean()
            )

            if (
                pd.notna(curr_mean)
                and pd.notna(prev_mean)
                and curr_mean > 0
                and prev_mean > 0
            ):
                scale = prev_mean / curr_mean
                chunk["raw"] = chunk["raw"] * scale

        new_dates = chunk.index.difference(result.index)

        result = pd.concat(
            [
                result,
                chunk.loc[new_dates],
            ]
        )

    result = result.sort_index()

    result.index.name = "Дата"

    result.columns = ["Поисковая активность Chainlink (Google Trends)"]

    col = "Поисковая активность Chainlink (Google Trends)"

    max_val = result[col].max()

    if max_val > 0:
        result[col] = (
            result[col] / max_val * 100
        ).round(2)

    return result


def validate_result(df):
    print("\nПроверка данных:")

    if df.empty:
        print("   Датасет пустой")
        return

    print(f"   Строк: {len(df)}")
    print(f"   Диапазон: {df.index.min().date()} -> {df.index.max().date()}")

    full_range = pd.date_range(
        start=df.index.min(),
        end=df.index.max(),
        freq="D",
    )

    missing = full_range.difference(df.index)

    print(f"   Пропущенных дат внутри диапазона: {len(missing)}")

    if len(missing) > 0:
        print("   Первые пропуски:")
        print([d.date() for d in missing[:20]])

    col = "Поисковая активность Chainlink (Google Trends)"

    print("\n   Описание ряда:")
    print(df[col].describe())

    print("\n   Первые строки:")
    print(df.head(10))

    print("\n   Последние строки:")
    print(df.tail(10))


def save_excel(df, path):
    print(f"\nСохраняю Excel: {os.path.basename(path)}")

    df.to_excel(
        path,
        sheet_name="Google Trends",
    )

    from openpyxl import load_workbook
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from openpyxl.formatting.rule import ColorScaleRule

    wb = load_workbook(path)
    ws = wb["Google Trends"]

    hdr_fill = PatternFill("solid", fgColor="1F3864")
    hdr_font = Font(
        color="FFFFFF",
        bold=True,
        size=10,
        name="Calibri",
    )

    body_font = Font(
        size=10,
        name="Calibri",
    )

    thin = Side(
        style="thin",
        color="DDDDDD",
    )

    brd = Border(
        left=thin,
        right=thin,
        top=thin,
        bottom=thin,
    )

    for cell in ws[1]:
        cell.fill = hdr_fill
        cell.font = hdr_font
        cell.alignment = Alignment(
            horizontal="center",
            vertical="center",
            wrap_text=True,
        )
        cell.border = brd

    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.font = body_font
            cell.border = brd

            if isinstance(cell.value, float):
                cell.number_format = "#,##0.00"

    for col_cells in ws.columns:
        max_len = max(
            (
                len(str(c.value or ""))
                for c in col_cells
            ),
            default=8,
        )

        ws.column_dimensions[
            get_column_letter(col_cells[0].column)
        ].width = min(max_len + 3, 45)

    if ws.max_row >= 2:
        ws.conditional_formatting.add(
            f"B2:B{ws.max_row}",
            ColorScaleRule(
                start_type="min",
                start_color="FFFFFF",
                end_type="max",
                end_color="FF6600",
            ),
        )

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    wb.save(path)

    print(f"   Готово. Строк: {len(df)}")


if __name__ == "__main__":
    print("=" * 70)
    print("  Google Trends — Chainlink / LINK  2015 -> апрель 2026")
    print("=" * 70)

    chunks = download_chunks()

    print(f"\nСклеиваю {len(chunks)} кусков...")

    result = normalize_and_stitch(chunks)

    if result.empty:
        print("Нет данных")
    else:
        result = result.loc[
            START.strftime("%Y-%m-%d"):
            END.strftime("%Y-%m-%d")
        ]

        print(
            f"Итого: {len(result)} строк | "
            f"{result.index.min().date()} -> {result.index.max().date()}"
        )

        validate_result(result)

        save_excel(result, OUT)

        if os.path.exists(CHECKPOINT):
            os.remove(CHECKPOINT)
            print("Чекпоинт удалён.")

        print("\n" + "=" * 70)
        print(f"Готово: {OUT}")
        print("=" * 70)