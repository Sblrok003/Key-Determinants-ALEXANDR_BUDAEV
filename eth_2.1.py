import os
import re
import warnings
from pathlib import Path
from itertools import combinations

import numpy as np
import pandas as pd

import matplotlib
matplotlib.use("Agg")
import matplotlib.pyplot as plt

from sklearn.dummy import DummyClassifier
from sklearn.linear_model import LogisticRegression
from sklearn.ensemble import RandomForestClassifier, ExtraTreesClassifier
from sklearn.pipeline import Pipeline
from sklearn.preprocessing import StandardScaler
from sklearn.neural_network import MLPClassifier
from sklearn.metrics import (
    accuracy_score,
    balanced_accuracy_score,
    f1_score,
    roc_auc_score,
)

warnings.filterwarnings("ignore")


# ============================================================
# 1. SETTINGS ETH
# ============================================================

ASSET_NAME = "ETH"

FILE_PATH = r"C:\Users\mrsas\PycharmProjects\BTC\BTC_pr\БОЛЬШАЯ БАЗА ДАННЫХ ПО ВСЕМ.xlsx"
SHEET_NAME = "ML_ETH"

BASE_DIR = Path(FILE_PATH).parent

OUTPUT_DIR = BASE_DIR / "eth_2_1_joint_model_outputs_fast_selection_final_heavy"
PLOTS_DIR = OUTPUT_DIR / "plots"

STAGE1_DIR = BASE_DIR / "eth_1_1_baseline_feature_test_outputs"

CANDIDATE_POOL_PATH = STAGE1_DIR / "candidate_pool_for_joint_stage.xlsx"

CANDIDATE_POOL_SHEET = "candidate_pool_hz_threshold"

TRAIN_START = "2015-08-07"
TRAIN_END = "2021-12-31"

VAL_START = "2022-01-01"
VAL_END = "2023-12-31"

TEST_START = "2024-01-01"
TEST_END = "2026-03-31"

HORIZONS = [1, 3, 7, 14, 30]
TARGET_THRESHOLDS = [0.00, 0.01, 0.02, 0.03, 0.05]

MIN_TRAIN_ROWS = 100
MIN_VAL_ROWS = 50
MIN_TEST_ROWS = 50

DO_NOT_MIX_THRESHOLDS = True

EXCLUDED_FEATURES = {
    "us_cpi_inflation",
}

GROUP_ORDER = ["market", "behavior", "onchain", "macro"]

LAG_DAYS = [1, 2, 3, 7]

REGIME_ROLLING_WINDOW = 90
REGIME_Q_LOW = 0.25
REGIME_Q_HIGH = 0.75

MAX_INTERACTION_BASE_FEATURES = 10
MAX_INTERACTION_CANDIDATES = 60

CANDIDATE_POOL_MAX_PER_H_THR = 30


# ============================================================
# SPEED LOGIC
# ============================================================

USE_XGBOOST = True
USE_CATBOOST = True
USE_MLP = True
USE_LSTM = True

INCLUDE_LSTM_IN_SELECTION = False

LSTM_SEQUENCE_LENGTHS = [30, 60]
LSTM_MIN_TRAIN_SEQUENCES = 300
LSTM_EPOCHS = 45
LSTM_BATCH_SIZE = 64
LSTM_PATIENCE = 8

SELECTION_MODELS = [
    "LogisticRegression",
    "RandomForest",
    "ExtraTrees",
]

FINAL_MODEL_NAMES = [
    "LogisticRegression",
    "RandomForest",
    "ExtraTrees",
    "XGBoost",
    "CatBoost",
    "MLP_small",
    "MLP_deep",
    "MLP_wide",
    "MLP_regularized",
    "LSTM_deep_seq30",
    "LSTM_deep_seq60",
]

CONFIDENCE_SELECTION_MODELS = FINAL_MODEL_NAMES.copy()

MIN_SCORE_IMPROVEMENT = 0.00001
MAX_ALLOWED_AUC_DROP = 0.00000
MAX_ALLOWED_BA_DROP = 0.00000

CONFIDENCE_GRID = [round(x, 2) for x in np.arange(0.50, 0.91, 0.01)]

PRACTICAL_MIN_ACC = 0.60
PRACTICAL_MIN_COVERAGE = 0.30
PRACTICAL_MAX_COVERAGE = 0.50

RARE_MIN_ACC = 0.65
RARE_MIN_COVERAGE = 0.05
RARE_MAX_COVERAGE = 0.30

BALANCED_MIN_COVERAGE = 0.50


# ============================================================
# 2. COLUMN MAP ETH
# ============================================================

RENAME_MAP = {
    "Дата": "date",

    "Лог-доходность": "log_return",
    "log_return_lag_1": "log_return_lag_1",
    "log_return_lag_2": "log_return_lag_2",
    "log_return_lag_3": "log_return_lag_3",

    "atr_14": "atr_14",
    "rolling_vol_7": "rolling_vol_7",
    "rolling_vol_14": "rolling_vol_14",
    "rolling_vol_30": "rolling_vol_30",

    "ETH_mcap_log_return": "eth_mcap_log_return",
    "eth_mcap_log_return": "eth_mcap_log_return",

    "total_mcap_log_return": "total_mcap_log_return",
    "btc_dominance_change": "btc_dominance_change",

    "RSI (14)": "rsi_14",
    "RSI_14": "rsi_14",
    "rsi_14": "rsi_14",

    "rsi_oversold_dummy": "rsi_oversold_dummy",
    "rsi_overbought_dummy": "rsi_overbought_dummy",
    "sma_ratio": "sma_ratio",

    "Индекс страха и жадности": "fear_greed_index",
    "fear_greed_change": "fear_greed_change",
    "fear_dummy": "fear_dummy",
    "greed_dummy": "greed_dummy",
    "Поисковая активность (Google Trends)": "google_trends",

    "candle_body": "candle_body",
    "ohlc_range": "ohlc_range",
    "volume_growth": "volume_growth",
    "Давление покупателей/продавцов (buy/sell pressure)": "buy_sell_pressure",

    "transactions_log_return": "transactions_log_return",
    "fees_log_return": "fees_log_return",
    "active_addresses_log_return": "active_addresses_log_return",
    "avg_fee_log_return": "avg_fee_log_return",

    "sp500_return": "sp500_return",
    "dxy_return": "dxy_return",
    "Процентные ставки (ФРС)": "fed_rate",
    "Инфляция (CPI) США": "us_cpi_inflation",
    "gold_log_return": "gold_log_return",

    "target_return_1d": "target_return_1d",
    "target_direction_1d": "target_direction_1d",
    "target_return_3d": "target_return_3d",
    "target_direction_3d": "target_direction_3d",
    "target_return_7d": "target_return_7d",
    "target_direction_7d": "target_direction_7d",
    "target_return_14d": "target_return_14d",
    "target_direction_14d": "target_direction_14d",
    "target_return_30d": "target_return_30d",
    "target_direction_30d": "target_direction_30d",
}


# ============================================================
# 3. FEATURE GROUPS ETH
# ============================================================

SET0_BASE = [
    "log_return_lag_1",
    "log_return_lag_2",
    "log_return_lag_3",
    "rolling_vol_7",
    "rolling_vol_14",
    "rolling_vol_30",
]

FEATURE_GROUP_DICT = {
    "SET0": [
        "log_return_lag_1",
        "log_return_lag_2",
        "log_return_lag_3",
        "rolling_vol_7",
        "rolling_vol_14",
        "rolling_vol_30",
    ],

    "market": [
        "rsi_14",
        "rsi_oversold_dummy",
        "rsi_overbought_dummy",
        "atr_14",
        "sma_ratio",
        "candle_body",
        "ohlc_range",
        "volume_growth",
        "buy_sell_pressure",
        "eth_mcap_log_return",
        "total_mcap_log_return",
        "btc_dominance_change",
    ],

    "behavior": [
        "fear_greed_index",
        "fear_greed_change",
        "fear_dummy",
        "greed_dummy",
        "google_trends",
    ],

    "onchain": [
        "transactions_log_return",
        "fees_log_return",
        "active_addresses_log_return",
        "avg_fee_log_return",
    ],

    "macro": [
        "sp500_return",
        "dxy_return",
        "gold_log_return",
        "fed_rate",
        "us_cpi_inflation",
    ],
}

GROUP_RU = {
    "SET0": "базовые признаки цены и волатильности ETH",
    "market": "рыночные признаки ETH",
    "behavior": "поведенческие признаки",
    "onchain": "ончейн-признаки ETH",
    "macro": "макроэкономические признаки",
    "unknown": "неизвестная группа",
}


# ============================================================
# 4. LSTM / PYTORCH
# ============================================================

TORCH_AVAILABLE = False
torch = None
nn = None
DataLoader = None
TensorDataset = None

if USE_LSTM:
    try:
        import torch
        import torch.nn as nn
        from torch.utils.data import DataLoader, TensorDataset
        TORCH_AVAILABLE = True
    except Exception:
        TORCH_AVAILABLE = False


if TORCH_AVAILABLE:
    class TorchLSTMNet(nn.Module):
        def __init__(self, input_size, hidden_size=64, num_layers=3, dropout=0.25):
            super().__init__()

            self.lstm = nn.LSTM(
                input_size=input_size,
                hidden_size=hidden_size,
                num_layers=num_layers,
                dropout=dropout if num_layers > 1 else 0.0,
                batch_first=True,
            )

            self.head = nn.Sequential(
                nn.Linear(hidden_size, 32),
                nn.ReLU(),
                nn.Dropout(dropout),
                nn.Linear(32, 1),
            )

        def forward(self, x):
            out, _ = self.lstm(x)
            last = out[:, -1, :]
            logits = self.head(last).squeeze(-1)
            return logits
else:
    TorchLSTMNet = None


class TorchLSTMClassifier:
    def __init__(
        self,
        seq_len=30,
        hidden_size=64,
        num_layers=3,
        dropout=0.25,
        epochs=45,
        batch_size=64,
        lr=0.001,
        weight_decay=0.001,
        patience=8,
        min_train_sequences=300,
        random_state=42,
    ):
        self.seq_len = int(seq_len)
        self.hidden_size = hidden_size
        self.num_layers = num_layers
        self.dropout = dropout
        self.epochs = epochs
        self.batch_size = batch_size
        self.lr = lr
        self.weight_decay = weight_decay
        self.patience = patience
        self.min_train_sequences = min_train_sequences
        self.random_state = random_state

        self.scaler = StandardScaler()
        self.model = None
        self.classes_ = np.array([0, 1])
        self._is_sequence_model = True

        if TORCH_AVAILABLE:
            self.device = "cuda" if torch.cuda.is_available() else "cpu"
        else:
            self.device = "cpu"

    def _make_sequences(self, X, y=None):
        X = np.asarray(X, dtype=np.float32)
        n = len(X)

        if n < self.seq_len:
            if y is None:
                return np.empty((0, self.seq_len, X.shape[1]), dtype=np.float32)

            return (
                np.empty((0, self.seq_len, X.shape[1]), dtype=np.float32),
                np.empty((0,), dtype=np.float32),
            )

        seqs = []
        ys = []

        for end in range(self.seq_len - 1, n):
            start = end - self.seq_len + 1
            seqs.append(X[start:end + 1])

            if y is not None:
                ys.append(y[end])

        X_seq = np.asarray(seqs, dtype=np.float32)

        if y is None:
            return X_seq

        y_seq = np.asarray(ys, dtype=np.float32)
        return X_seq, y_seq

    def fit(self, X, y):
        if not TORCH_AVAILABLE:
            raise RuntimeError("PyTorch is not installed. LSTM skipped.")

        torch.manual_seed(self.random_state)
        np.random.seed(self.random_state)

        X = np.asarray(X, dtype=np.float32)
        y = np.asarray(y).astype(np.float32)

        X_scaled = self.scaler.fit_transform(X)

        X_seq, y_seq = self._make_sequences(X_scaled, y)

        if len(X_seq) < self.min_train_sequences:
            raise RuntimeError(
                f"Too few effective LSTM train sequences: {len(X_seq)}. "
                f"Need at least {self.min_train_sequences}. "
                f"Rows={len(X)}, seq_len={self.seq_len}."
            )

        if len(np.unique(y_seq)) < 2:
            raise RuntimeError("LSTM y_train sequence contains one class only.")

        n = len(X_seq)
        val_size = max(int(n * 0.15), 1)
        train_size = n - val_size

        X_tr = X_seq[:train_size]
        y_tr = y_seq[:train_size]

        X_va = X_seq[train_size:]
        y_va = y_seq[train_size:]

        train_ds = TensorDataset(
            torch.tensor(X_tr, dtype=torch.float32),
            torch.tensor(y_tr, dtype=torch.float32),
        )

        val_x = torch.tensor(X_va, dtype=torch.float32).to(self.device)
        val_y = torch.tensor(y_va, dtype=torch.float32).to(self.device)

        loader = DataLoader(
            train_ds,
            batch_size=self.batch_size,
            shuffle=False,
            drop_last=False,
        )

        input_size = X_seq.shape[2]

        self.model = TorchLSTMNet(
            input_size=input_size,
            hidden_size=self.hidden_size,
            num_layers=self.num_layers,
            dropout=self.dropout,
        ).to(self.device)

        pos = float((y_tr == 1).sum())
        neg = float((y_tr == 0).sum())

        if pos > 0:
            pos_weight = torch.tensor([neg / pos], dtype=torch.float32).to(self.device)
        else:
            pos_weight = torch.tensor([1.0], dtype=torch.float32).to(self.device)

        criterion = nn.BCEWithLogitsLoss(pos_weight=pos_weight)

        optimizer = torch.optim.AdamW(
            self.model.parameters(),
            lr=self.lr,
            weight_decay=self.weight_decay,
        )

        best_val_loss = np.inf
        best_state = None
        no_improve = 0

        for epoch in range(self.epochs):
            self.model.train()

            for xb, yb in loader:
                xb = xb.to(self.device)
                yb = yb.to(self.device)

                optimizer.zero_grad()
                logits = self.model(xb)
                loss = criterion(logits, yb)
                loss.backward()
                optimizer.step()

            self.model.eval()

            with torch.no_grad():
                logits = self.model(val_x)
                val_loss = criterion(logits, val_y).item()

            if val_loss < best_val_loss - 1e-5:
                best_val_loss = val_loss
                best_state = {
                    k: v.detach().cpu().clone()
                    for k, v in self.model.state_dict().items()
                }
                no_improve = 0
            else:
                no_improve += 1

            if no_improve >= self.patience:
                break

        if best_state is not None:
            self.model.load_state_dict(best_state)

        return self

    def predict_proba(self, X):
        if self.model is None:
            raise RuntimeError("LSTM model is not fitted.")

        X = np.asarray(X, dtype=np.float32)
        X_scaled = self.scaler.transform(X)

        X_seq = self._make_sequences(X_scaled, y=None)

        out = np.full((len(X), 2), np.nan, dtype=np.float32)

        if len(X_seq) == 0:
            return out

        self.model.eval()

        probs = []

        with torch.no_grad():
            for i in range(0, len(X_seq), self.batch_size):
                xb = torch.tensor(
                    X_seq[i:i + self.batch_size],
                    dtype=torch.float32,
                ).to(self.device)

                logits = self.model(xb)
                p = torch.sigmoid(logits).detach().cpu().numpy()
                probs.append(p)

        p = np.concatenate(probs)

        start = self.seq_len - 1
        out[start:, 1] = p
        out[start:, 0] = 1.0 - p

        return out

    def predict(self, X):
        proba = self.predict_proba(X)[:, 1]
        pred = np.where(proba >= 0.5, 1, 0)
        pred[~np.isfinite(proba)] = -1
        return pred


# ============================================================
# 5. UTILS
# ============================================================

def ensure_dirs():
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    PLOTS_DIR.mkdir(parents=True, exist_ok=True)


def clean_col_name(name):
    name = str(name)
    name = name.replace("\n", " ").replace("\r", " ")
    name = name.replace('"', "").replace("'", "")
    name = re.sub(r"\s+", " ", name)
    return name.strip()


def fix_duplicate_target_columns(cols):
    fixed = []
    target_seen = {}

    for col in cols:
        c = clean_col_name(col)
        base = re.sub(r"\.\d+$", "", c)

        m = re.match(r"^target_return_(\d+)d$", base)

        if m:
            h = m.group(1)

            if base not in target_seen:
                target_seen[base] = 0
                fixed.append(f"target_return_{h}d")
            else:
                target_seen[base] += 1
                fixed.append(f"target_direction_{h}d")
        else:
            fixed.append(c)

    return fixed


def collapse_duplicate_columns(df):
    out = {}

    for i, col in enumerate(df.columns):
        s = df.iloc[:, i]

        if col in out:
            out[col] = out[col].combine_first(s)
        else:
            out[col] = s

    return pd.DataFrame(out)


def to_numeric_safe(s):
    if pd.api.types.is_numeric_dtype(s):
        return s

    s = s.astype(str)
    s = s.str.replace("\u00a0", "", regex=False)
    s = s.str.replace(" ", "", regex=False)
    s = s.str.replace("%", "", regex=False)
    s = s.str.replace("$", "", regex=False)
    s = s.str.replace("−", "-", regex=False)
    s = s.str.replace(",", ".", regex=False)
    s = s.replace({"nan": np.nan, "None": np.nan, "": np.nan})

    return pd.to_numeric(s, errors="coerce")


def safe_auc(y_true, proba):
    y_true = np.asarray(y_true)
    proba = np.asarray(proba)

    mask = np.isfinite(proba)
    y_true = y_true[mask]
    proba = proba[mask]

    if len(y_true) < 2:
        return np.nan

    if len(np.unique(y_true)) < 2:
        return np.nan

    try:
        return roc_auc_score(y_true, proba)
    except Exception:
        return np.nan


def calc_metrics(y_true, pred, proba=None):
    y_true = np.asarray(y_true)
    pred = np.asarray(pred)

    mask = (pred == 0) | (pred == 1)

    if proba is not None:
        proba = np.asarray(proba)
        mask = mask & np.isfinite(proba)

    y_eval = y_true[mask]
    pred_eval = pred[mask]

    if proba is not None:
        proba_eval = proba[mask]
    else:
        proba_eval = None

    if len(y_eval) == 0:
        return {
            "accuracy": np.nan,
            "balanced_accuracy": np.nan,
            "f1": np.nan,
            "roc_auc": np.nan,
            "effective_eval_rows": 0,
        }

    result = {
        "accuracy": accuracy_score(y_eval, pred_eval),
        "balanced_accuracy": balanced_accuracy_score(y_eval, pred_eval)
        if len(np.unique(y_eval)) >= 2 else np.nan,
        "f1": f1_score(y_eval, pred_eval, zero_division=0),
        "effective_eval_rows": len(y_eval),
    }

    if proba_eval is not None:
        result["roc_auc"] = safe_auc(y_eval, proba_eval)
    else:
        result["roc_auc"] = np.nan

    return result


def get_positive_proba(model, X):
    if not hasattr(model, "predict_proba"):
        return None

    proba = model.predict_proba(X)

    classes = getattr(model, "classes_", None)

    if classes is None and hasattr(model, "steps"):
        classes = getattr(model.steps[-1][1], "classes_", None)

    if classes is None:
        if proba.shape[1] == 2:
            return proba[:, 1]

        return np.full(len(X), np.nan)

    classes = list(classes)

    if 1 in classes:
        return proba[:, classes.index(1)]

    return np.zeros(len(X))


def safe_filename(name):
    name = str(name)
    name = re.sub(r"[^a-zA-Z0-9а-яА-Я_\-]+", "_", name)
    name = name.strip("_")
    return name[:120]


def feature_group_name(feature):
    for group, features in FEATURE_GROUP_DICT.items():
        if feature in features:
            return group

    return "unknown"


def existing_features(df, features):
    return [f for f in features if f in df.columns]


def unique_keep_order(items):
    seen = set()
    out = []

    for x in items:
        if x not in seen:
            out.append(x)
            seen.add(x)

    return out


def find_file(base_dir, filename):
    direct = base_dir / filename

    if direct.exists():
        return direct

    try:
        matches = list(base_dir.rglob(filename))
        if matches:
            return matches[0]
    except Exception:
        pass

    return None


def threshold_to_pct(thr):
    return int(round(float(thr) * 100))


def target_name(horizon, threshold_pct):
    return f"target_direction_{horizon}d_thr_{int(threshold_pct)}pct"


def write_log(messages):
    with open(OUTPUT_DIR / "2_1_log.txt", "w", encoding="utf-8") as f:
        for msg in messages:
            f.write(str(msg) + "\n")


def normalize_threshold_pct_series(s):
    x = pd.to_numeric(s, errors="coerce")

    finite = x.dropna()

    if not finite.empty and finite.max() <= 1.0:
        x = x * 100

    return x.round().astype("Int64")


# ============================================================
# 6. LOAD ML DATA
# ============================================================

def read_excel_with_fallback():
    df = pd.read_excel(FILE_PATH, sheet_name=SHEET_NAME, header=1)
    cols_clean = [clean_col_name(c) for c in df.columns]

    if "Дата" in cols_clean or "date" in cols_clean:
        return df

    print("WARNING: header=1 did not find date column. Trying header=0.")
    df = pd.read_excel(FILE_PATH, sheet_name=SHEET_NAME, header=0)

    return df


def load_ml_data():
    print("Loading ETH ML data...")

    df = read_excel_with_fallback()
    df.columns = fix_duplicate_target_columns(df.columns)

    keep_cols = []

    for c in df.columns:
        if str(c).lower().startswith("unnamed") and df[c].isna().all():
            continue
        keep_cols.append(c)

    df = df[keep_cols].copy()

    with open(OUTPUT_DIR / "columns_before_rename.txt", "w", encoding="utf-8") as f:
        for c in df.columns:
            f.write(str(c) + "\n")

    rename_actual = {}

    for c in df.columns:
        base = clean_col_name(c)
        if base in RENAME_MAP:
            rename_actual[c] = RENAME_MAP[base]

    df = df.rename(columns=rename_actual)
    df = collapse_duplicate_columns(df)

    with open(OUTPUT_DIR / "columns_after_rename.txt", "w", encoding="utf-8") as f:
        for c in df.columns:
            f.write(str(c) + "\n")

    if "date" not in df.columns:
        raise ValueError("Не найдена колонка date. Проверь колонку 'Дата' в листе ML_ETH.")

    df["date"] = pd.to_datetime(df["date"], errors="coerce", dayfirst=True)
    df = df.dropna(subset=["date"])
    df = df.sort_values("date").reset_index(drop=True)

    for c in df.columns:
        if c != "date":
            df[c] = to_numeric_safe(df[c])

    df = create_targets(df)
    df = df.replace([np.inf, -np.inf], np.nan)

    df.head(300).to_csv(
        OUTPUT_DIR / "prepared_preview.csv",
        index=False,
        encoding="utf-8-sig"
    )

    print(f"Rows: {len(df)}")
    print(f"Date range: {df['date'].min()} — {df['date'].max()}")

    return df


def create_targets(df):
    for h in HORIZONS:
        ret_col = f"target_return_{h}d"

        if ret_col not in df.columns:
            print(f"WARNING: no {ret_col}")
            continue

        for thr in TARGET_THRESHOLDS:
            pct = threshold_to_pct(thr)
            target_col = f"target_direction_{h}d_thr_{pct}pct"

            if thr == 0:
                df[target_col] = np.where(
                    df[ret_col] > 0,
                    1,
                    np.where(df[ret_col] <= 0, 0, np.nan)
                )
            else:
                df[target_col] = np.where(
                    df[ret_col] > thr,
                    1,
                    np.where(df[ret_col] < -thr, 0, np.nan)
                )

    return df


# ============================================================
# 7. LOAD OR BUILD CANDIDATE POOL
# ============================================================

def build_candidate_pool_from_stage1():
    """
    Если candidate_pool_for_joint_stage.xlsx не создан первым этапом,
    эта функция собирает его из 1_1_C_individual_deltas_vs_set0.csv.

    Логика:
    - берём validation;
    - берём только модели отбора: LogisticRegression, RandomForest, ExtraTrees;
    - по каждой паре horizon + threshold + feature считаем средний delta AUC и delta balanced accuracy;
    - ранжируем признаки внутри каждой пары horizon + threshold;
    - сохраняем candidate_pool_for_joint_stage.xlsx.
    """

    deltas_path = STAGE1_DIR / "1_1_C_individual_deltas_vs_set0.csv"
    rec_path = STAGE1_DIR / "1_1_D_feature_recommendations.csv"

    if not deltas_path.exists():
        raise FileNotFoundError(
            "Не найден candidate_pool_for_joint_stage.xlsx и не найден "
            f"{deltas_path}. Сначала запусти этап 1.1 для ETH."
        )

    print("\nCandidate pool xlsx not found.")
    print("Building candidate pool from Stage 1.1 deltas:")
    print(deltas_path)

    deltas = pd.read_csv(deltas_path)

    deltas.columns = [clean_col_name(c) for c in deltas.columns]

    required = [
        "feature",
        "horizon",
        "target_threshold_pct",
        "model",
        "eval_split",
        "delta_roc_auc",
        "delta_balanced_accuracy",
    ]

    missing = [c for c in required if c not in deltas.columns]

    if missing:
        raise ValueError(
            f"В {deltas_path.name} нет нужных колонок: {missing}. "
            f"Есть колонки: {list(deltas.columns)}"
        )

    d = deltas[
        (deltas["eval_split"] == "validation") &
        (deltas["model"].isin(SELECTION_MODELS))
    ].copy()

    d["horizon"] = pd.to_numeric(d["horizon"], errors="coerce").astype("Int64")
    d["target_threshold_pct"] = normalize_threshold_pct_series(d["target_threshold_pct"])

    d["delta_roc_auc"] = pd.to_numeric(d["delta_roc_auc"], errors="coerce")
    d["delta_balanced_accuracy"] = pd.to_numeric(d["delta_balanced_accuracy"], errors="coerce")
    d["delta_accuracy"] = pd.to_numeric(d.get("delta_accuracy", np.nan), errors="coerce")
    d["delta_f1"] = pd.to_numeric(d.get("delta_f1", np.nan), errors="coerce")

    d = d.dropna(subset=["feature", "horizon", "target_threshold_pct"])

    d = d[~d["feature"].isin(EXCLUDED_FEATURES)].copy()

    if d.empty:
        raise RuntimeError("После фильтрации Stage 1.1 deltas пустые. Candidate pool собрать нельзя.")

    agg = (
        d.groupby(["horizon", "target_threshold_pct", "feature"], as_index=False)
        .agg(
            mean_delta_roc_auc=("delta_roc_auc", "mean"),
            mean_delta_balanced_accuracy=("delta_balanced_accuracy", "mean"),
            mean_delta_accuracy=("delta_accuracy", "mean"),
            mean_delta_f1=("delta_f1", "mean"),
            positive_auc_rate=("delta_roc_auc", lambda x: (x > 0).mean()),
            positive_balanced_accuracy_rate=("delta_balanced_accuracy", lambda x: (x > 0).mean()),
            n_validation_model_tests=("model", "count"),
        )
    )

    agg["feature_group"] = agg["feature"].apply(feature_group_name)

    agg["candidate_score"] = (
        0.55 * agg["mean_delta_roc_auc"].fillna(0)
        + 0.45 * agg["mean_delta_balanced_accuracy"].fillna(0)
    )

    if rec_path.exists():
        rec = pd.read_csv(rec_path)
        rec.columns = [clean_col_name(c) for c in rec.columns]

        if "feature" in rec.columns and "recommendation" in rec.columns:
            rec_small = rec[["feature", "recommendation"]].drop_duplicates("feature")
            agg = agg.merge(rec_small, on="feature", how="left")
        else:
            agg["recommendation"] = ""
    else:
        agg["recommendation"] = ""

    # Не душим слишком жёстко: оставляем всё, что не явно EXCLUDE,
    # а если всё EXCLUDE, то всё равно берём топ по score.
    non_exclude = agg[agg["recommendation"].fillna("") != "EXCLUDE"].copy()

    if not non_exclude.empty:
        agg_use = non_exclude.copy()
    else:
        agg_use = agg.copy()

    agg_use = agg_use.sort_values(
        ["horizon", "target_threshold_pct", "candidate_score", "mean_delta_roc_auc"],
        ascending=[True, True, False, False],
    )

    selected_parts = []

    for (h, pct), g in agg_use.groupby(["horizon", "target_threshold_pct"]):
        g = g.sort_values(
            ["candidate_score", "mean_delta_roc_auc", "mean_delta_balanced_accuracy"],
            ascending=[False, False, False],
        ).head(CANDIDATE_POOL_MAX_PER_H_THR).copy()

        g["rank"] = np.arange(1, len(g) + 1)
        selected_parts.append(g)

    pool = pd.concat(selected_parts, ignore_index=True)

    pool = pool[
        [
            "horizon",
            "target_threshold_pct",
            "rank",
            "feature",
            "feature_group",
            "candidate_score",
            "mean_delta_roc_auc",
            "mean_delta_balanced_accuracy",
            "mean_delta_accuracy",
            "mean_delta_f1",
            "positive_auc_rate",
            "positive_balanced_accuracy_rate",
            "n_validation_model_tests",
            "recommendation",
        ]
    ].copy()

    with pd.ExcelWriter(CANDIDATE_POOL_PATH, engine="openpyxl") as writer:
        readme = pd.DataFrame({
            "section": [
                "Описание",
                "Источник",
                "Логика",
                "Важно",
            ],
            "text": [
                "Автоматически собранный candidate pool для ETH Stage 2.1.",
                str(deltas_path),
                "Кандидаты ранжируются внутри каждой пары horizon + threshold по validation delta ROC-AUC и delta balanced accuracy.",
                "Test не использовался для отбора кандидатов.",
            ],
        })

        readme.to_excel(writer, sheet_name="README", index=False)
        pool.to_excel(writer, sheet_name=CANDIDATE_POOL_SHEET, index=False)

    print(f"Candidate pool created: {CANDIDATE_POOL_PATH}")
    print(pool.head(50))

    return CANDIDATE_POOL_PATH


def load_candidate_pool():
    path = CANDIDATE_POOL_PATH

    if not path.exists():
        found = find_file(STAGE1_DIR, "candidate_pool_for_joint_stage.xlsx")
        if found is not None:
            path = found

    if not path.exists():
        path = build_candidate_pool_from_stage1()

    print(f"Loading candidate pool: {path}")

    xls = pd.ExcelFile(path)

    print("Available sheets in candidate pool:")
    for s in xls.sheet_names:
        print(" -", s)

    sheet_to_use = None

    if CANDIDATE_POOL_SHEET in xls.sheet_names:
        sheet_to_use = CANDIDATE_POOL_SHEET
    else:
        preferred = [
            "final_nested_h_thr",
            "final_candidates_nested_by_horizon_threshold",
            "final_candidates_horizon_threshold",
            "final_top_candidates_by_horizon_threshold",
            "next_stage_plan",
            "candidate_pool_h_thr",
        ]

        for s in preferred:
            if s in xls.sheet_names:
                sheet_to_use = s
                break

        if sheet_to_use is None:
            for s in xls.sheet_names:
                sl = s.lower()
                if "final" in sl and ("h_thr" in sl or "horizon" in sl):
                    sheet_to_use = s
                    break

        if sheet_to_use is None:
            for s in xls.sheet_names:
                sl = s.lower()
                if "candidate" in sl and ("h_thr" in sl or "horizon" in sl):
                    sheet_to_use = s
                    break

    if sheet_to_use is None:
        raise ValueError(
            "В candidate_pool_for_joint_stage.xlsx не найден лист с кандидатами. "
            "Нужен лист final_nested_h_thr."
        )

    print(f"Using candidate pool sheet: {sheet_to_use}")

    cand = pd.read_excel(path, sheet_name=sheet_to_use)
    cand.columns = [clean_col_name(c) for c in cand.columns]

    if "target_threshold" in cand.columns and "target_threshold_pct" not in cand.columns:
        cand["target_threshold_pct"] = cand["target_threshold"]

    if "threshold" in cand.columns and "target_threshold_pct" not in cand.columns:
        cand["target_threshold_pct"] = cand["threshold"]

    required = [
        "horizon",
        "target_threshold_pct",
        "feature",
    ]

    missing = [c for c in required if c not in cand.columns]

    if missing:
        raise ValueError(
            f"В листе {sheet_to_use} нет нужных колонок: {missing}. "
            f"Есть колонки: {list(cand.columns)}"
        )

    cand["horizon"] = pd.to_numeric(cand["horizon"], errors="coerce").astype("Int64")
    cand["target_threshold_pct"] = normalize_threshold_pct_series(cand["target_threshold_pct"])

    if "rank" not in cand.columns:
        cand["rank"] = cand.groupby(["horizon", "target_threshold_pct"]).cumcount() + 1

    cand["rank"] = pd.to_numeric(cand["rank"], errors="coerce").astype("Int64")

    cand = cand.dropna(subset=["horizon", "target_threshold_pct", "rank", "feature"]).copy()

    cand["horizon"] = cand["horizon"].astype(int)
    cand["target_threshold_pct"] = cand["target_threshold_pct"].astype(int)
    cand["rank"] = cand["rank"].astype(int)

    cand["feature"] = cand["feature"].astype(str).str.strip()

    cand["feature_group"] = cand["feature"].apply(feature_group_name)

    before = len(cand)
    cand = cand[~cand["feature"].isin(EXCLUDED_FEATURES)].copy()
    after = len(cand)

    if before != after:
        print(f"Excluded features: {EXCLUDED_FEATURES}. Removed rows: {before - after}")

    cand = cand.sort_values(
        ["horizon", "target_threshold_pct", "rank"],
        ascending=True
    ).reset_index(drop=True)

    cand.to_csv(
        OUTPUT_DIR / "2_1_used_candidate_pool_without_excluded_features.csv",
        index=False,
        encoding="utf-8-sig"
    )

    print("Candidate pool loaded.")
    print(
        cand[
            ["horizon", "target_threshold_pct", "rank", "feature", "feature_group"]
        ].head(50)
    )

    return cand


# ============================================================
# 8. MODELS
# ============================================================

def get_models(y_train, model_scope="selection"):
    models = {}

    models["Baseline_MajorityClass"] = DummyClassifier(strategy="most_frequent")

    models["LogisticRegression"] = Pipeline([
        ("scaler", StandardScaler()),
        ("model", LogisticRegression(
            max_iter=5000,
            solver="lbfgs",
            class_weight="balanced",
            random_state=42,
        ))
    ])

    models["RandomForest"] = RandomForestClassifier(
        n_estimators=400,
        max_depth=4,
        min_samples_leaf=20,
        class_weight="balanced",
        random_state=42,
        n_jobs=-1,
    )

    models["ExtraTrees"] = ExtraTreesClassifier(
        n_estimators=400,
        max_depth=4,
        min_samples_leaf=20,
        class_weight="balanced",
        random_state=42,
        n_jobs=-1,
    )

    if model_scope == "selection":
        return models

    if USE_MLP:
        models["MLP_small"] = Pipeline([
            ("scaler", StandardScaler()),
            ("model", MLPClassifier(
                hidden_layer_sizes=(64, 32),
                activation="relu",
                solver="adam",
                alpha=0.001,
                learning_rate_init=0.001,
                max_iter=700,
                early_stopping=True,
                validation_fraction=0.15,
                n_iter_no_change=25,
                random_state=42,
            ))
        ])

        models["MLP_deep"] = Pipeline([
            ("scaler", StandardScaler()),
            ("model", MLPClassifier(
                hidden_layer_sizes=(128, 64, 32, 16),
                activation="relu",
                solver="adam",
                alpha=0.001,
                learning_rate_init=0.0007,
                max_iter=1000,
                early_stopping=True,
                validation_fraction=0.15,
                n_iter_no_change=35,
                random_state=42,
            ))
        ])

        models["MLP_wide"] = Pipeline([
            ("scaler", StandardScaler()),
            ("model", MLPClassifier(
                hidden_layer_sizes=(256, 128, 64),
                activation="relu",
                solver="adam",
                alpha=0.0015,
                learning_rate_init=0.0007,
                max_iter=1000,
                early_stopping=True,
                validation_fraction=0.15,
                n_iter_no_change=35,
                random_state=42,
            ))
        ])

        models["MLP_regularized"] = Pipeline([
            ("scaler", StandardScaler()),
            ("model", MLPClassifier(
                hidden_layer_sizes=(128, 64, 32),
                activation="relu",
                solver="adam",
                alpha=0.01,
                learning_rate_init=0.0005,
                max_iter=1000,
                early_stopping=True,
                validation_fraction=0.15,
                n_iter_no_change=35,
                random_state=42,
            ))
        ])

    if USE_LSTM:
        if TORCH_AVAILABLE:
            for seq_len in LSTM_SEQUENCE_LENGTHS:
                models[f"LSTM_deep_seq{seq_len}"] = TorchLSTMClassifier(
                    seq_len=seq_len,
                    hidden_size=64,
                    num_layers=3,
                    dropout=0.25,
                    epochs=LSTM_EPOCHS,
                    batch_size=LSTM_BATCH_SIZE,
                    lr=0.001,
                    weight_decay=0.001,
                    patience=LSTM_PATIENCE,
                    min_train_sequences=LSTM_MIN_TRAIN_SEQUENCES,
                    random_state=42,
                )
        else:
            print("PyTorch is not installed. LSTM models skipped.")

    if USE_XGBOOST:
        try:
            from xgboost import XGBClassifier

            pos = int((y_train == 1).sum())
            neg = int((y_train == 0).sum())
            scale_pos_weight = neg / pos if pos > 0 else 1.0

            models["XGBoost"] = XGBClassifier(
                n_estimators=500,
                max_depth=3,
                learning_rate=0.03,
                subsample=0.8,
                colsample_bytree=0.8,
                min_child_weight=5,
                reg_lambda=5.0,
                reg_alpha=1.0,
                eval_metric="logloss",
                random_state=42,
                n_jobs=-1,
                scale_pos_weight=scale_pos_weight,
            )
        except Exception as e:
            print(f"XGBoost skipped: {e}")

    if USE_CATBOOST:
        try:
            from catboost import CatBoostClassifier

            models["CatBoost"] = CatBoostClassifier(
                iterations=500,
                depth=4,
                learning_rate=0.03,
                loss_function="Logloss",
                eval_metric="AUC",
                random_seed=42,
                verbose=False,
                auto_class_weights="Balanced",
                l2_leaf_reg=8.0,
            )
        except Exception as e:
            print(f"CatBoost skipped: {e}")

    return models


# ============================================================
# 9. SPLIT AND EVALUATION
# ============================================================

def make_model_data(df, features, target):
    needed = ["date"] + features + [target]
    missing = [c for c in needed if c not in df.columns]

    if missing:
        return pd.DataFrame()

    data = df[needed].copy()
    data = data.replace([np.inf, -np.inf], np.nan)
    data = data.dropna(subset=features + [target])

    data = data[
        (data["date"] >= pd.Timestamp(TRAIN_START)) &
        (data["date"] <= pd.Timestamp(TEST_END))
    ].copy()

    return data


def split_data(data):
    train = data[
        (data["date"] >= pd.Timestamp(TRAIN_START)) &
        (data["date"] <= pd.Timestamp(TRAIN_END))
    ].copy()

    val = data[
        (data["date"] >= pd.Timestamp(VAL_START)) &
        (data["date"] <= pd.Timestamp(VAL_END))
    ].copy()

    test = data[
        (data["date"] >= pd.Timestamp(TEST_START)) &
        (data["date"] <= pd.Timestamp(TEST_END))
    ].copy()

    return train, val, test


def data_ok(train, val, test, target):
    if len(train) < MIN_TRAIN_ROWS:
        return False

    if len(val) < MIN_VAL_ROWS:
        return False

    if len(test) < MIN_TEST_ROWS:
        return False

    if train[target].nunique() < 2:
        return False

    if val[target].nunique() < 2:
        return False

    if test[target].nunique() < 2:
        return False

    return True


def split_info(train, val, test, target):
    return {
        "train_rows": len(train),
        "validation_rows": len(val),
        "test_rows": len(test),

        "train_start": train["date"].min() if len(train) else pd.NaT,
        "train_end": train["date"].max() if len(train) else pd.NaT,

        "validation_start": val["date"].min() if len(val) else pd.NaT,
        "validation_end": val["date"].max() if len(val) else pd.NaT,

        "test_start": test["date"].min() if len(test) else pd.NaT,
        "test_end": test["date"].max() if len(test) else pd.NaT,

        "train_class_1_share": train[target].mean() if len(train) else np.nan,
        "validation_class_1_share": val[target].mean() if len(val) else np.nan,
        "test_class_1_share": test[target].mean() if len(test) else np.nan,
    }


def effective_sequence_rows(model_name, n_rows):
    if not model_name.startswith("LSTM_deep_seq"):
        return np.nan

    seq_len = int(model_name.replace("LSTM_deep_seq", ""))
    return max(0, n_rows - seq_len + 1)


def fit_eval_feature_set(
    df,
    features,
    target,
    stage,
    feature_set_name,
    horizon,
    threshold_pct,
    extra=None,
    model_scope="selection",
):
    if extra is None:
        extra = {}

    features = unique_keep_order([f for f in features if f in df.columns])

    data = make_model_data(df, features, target)

    if data.empty:
        return []

    train, val, test = split_data(data)

    if not data_ok(train, val, test, target):
        return []

    X_train = train[features] if len(features) > 0 else np.zeros((len(train), 1))
    y_train = train[target].astype(int).values

    X_val = val[features] if len(features) > 0 else np.zeros((len(val), 1))
    y_val = val[target].astype(int).values

    X_test = test[features] if len(features) > 0 else np.zeros((len(test), 1))
    y_test = test[target].astype(int).values

    info = split_info(train, val, test, target)

    rows = []

    models = get_models(y_train, model_scope=model_scope)

    for model_name, model in models.items():
        if model_name != "Baseline_MajorityClass" and len(features) == 0:
            continue

        try:
            model.fit(X_train, y_train)
        except Exception as e:
            rows.append({
                "asset": ASSET_NAME,
                "stage": stage,
                "feature_set_name": feature_set_name,
                "model": model_name,
                "model_scope": model_scope,
                "eval_split": "fit_error",
                "horizon": horizon,
                "target_threshold_pct": threshold_pct,
                "target": target,
                "n_features": len(features),
                "features": ",".join(features),
                "error": str(e),
                "is_neural_model": model_name.startswith("MLP") or model_name.startswith("LSTM"),
                "is_lstm_model": model_name.startswith("LSTM"),
                "is_xgboost_model": model_name == "XGBoost",
                "is_catboost_model": model_name == "CatBoost",
                "used_for_feature_selection": model_name in SELECTION_MODELS,
                "lstm_effective_train_sequences": effective_sequence_rows(model_name, len(train)),
                "lstm_effective_val_sequences": effective_sequence_rows(model_name, len(val)),
                "lstm_effective_test_sequences": effective_sequence_rows(model_name, len(test)),
                **info,
                **extra,
            })
            continue

        for split_name, X_eval, y_eval in [
            ("validation", X_val, y_val),
            ("test", X_test, y_test),
        ]:
            try:
                pred = model.predict(X_eval)
                proba = get_positive_proba(model, X_eval)

                metrics = calc_metrics(y_eval, pred, proba)

                rows.append({
                    "asset": ASSET_NAME,
                    "stage": stage,
                    "feature_set_name": feature_set_name,
                    "model": model_name,
                    "model_scope": model_scope,
                    "eval_split": split_name,
                    "horizon": horizon,
                    "target_threshold_pct": threshold_pct,
                    "target": target,
                    "n_features": len(features),
                    "features": ",".join(features),
                    "is_neural_model": model_name.startswith("MLP") or model_name.startswith("LSTM"),
                    "is_lstm_model": model_name.startswith("LSTM"),
                    "is_xgboost_model": model_name == "XGBoost",
                    "is_catboost_model": model_name == "CatBoost",
                    "used_for_feature_selection": model_name in SELECTION_MODELS,
                    "lstm_effective_train_sequences": effective_sequence_rows(model_name, len(train)),
                    "lstm_effective_val_sequences": effective_sequence_rows(model_name, len(val)),
                    "lstm_effective_test_sequences": effective_sequence_rows(model_name, len(test)),
                    **info,
                    **metrics,
                    **extra,
                })

            except Exception as e:
                rows.append({
                    "asset": ASSET_NAME,
                    "stage": stage,
                    "feature_set_name": feature_set_name,
                    "model": model_name,
                    "model_scope": model_scope,
                    "eval_split": split_name + "_error",
                    "horizon": horizon,
                    "target_threshold_pct": threshold_pct,
                    "target": target,
                    "n_features": len(features),
                    "features": ",".join(features),
                    "error": str(e),
                    "is_neural_model": model_name.startswith("MLP") or model_name.startswith("LSTM"),
                    "is_lstm_model": model_name.startswith("LSTM"),
                    "is_xgboost_model": model_name == "XGBoost",
                    "is_catboost_model": model_name == "CatBoost",
                    "used_for_feature_selection": model_name in SELECTION_MODELS,
                    "lstm_effective_train_sequences": effective_sequence_rows(model_name, len(train)),
                    "lstm_effective_val_sequences": effective_sequence_rows(model_name, len(val)),
                    "lstm_effective_test_sequences": effective_sequence_rows(model_name, len(test)),
                    **info,
                    **extra,
                })

    return rows


def aggregate_validation(rows, model_names=None):
    if not rows:
        return None

    if model_names is None:
        model_names = SELECTION_MODELS

    df = pd.DataFrame(rows)

    val = df[
        (df["eval_split"] == "validation") &
        (df["model"].isin(model_names))
    ].copy()

    if val.empty:
        return None

    out = {
        "mean_accuracy": val["accuracy"].mean(),
        "mean_balanced_accuracy": val["balanced_accuracy"].mean(),
        "mean_f1": val["f1"].mean(),
        "mean_roc_auc": val["roc_auc"].mean(),
        "n_models": val["model"].nunique(),
        "models_used": ",".join(sorted(val["model"].dropna().unique())),
    }

    out["validation_selection_score"] = (
        0.55 * out["mean_roc_auc"] +
        0.45 * out["mean_balanced_accuracy"]
    )

    return out


def is_improved(new_agg, current_agg):
    if new_agg is None:
        return False

    if current_agg is None:
        return True

    new_score = new_agg.get("validation_selection_score", -np.inf)
    cur_score = current_agg.get("validation_selection_score", -np.inf)

    new_auc = new_agg.get("mean_roc_auc", -np.inf)
    cur_auc = current_agg.get("mean_roc_auc", -np.inf)

    new_ba = new_agg.get("mean_balanced_accuracy", -np.inf)
    cur_ba = current_agg.get("mean_balanced_accuracy", -np.inf)

    score_ok = new_score > cur_score + MIN_SCORE_IMPROVEMENT
    auc_ok = new_auc >= cur_auc - MAX_ALLOWED_AUC_DROP
    ba_ok = new_ba >= cur_ba - MAX_ALLOWED_BA_DROP

    return bool(score_ok and auc_ok and ba_ok)


def agg_to_prefixed_dict(agg, prefix):
    if agg is None:
        return {
            f"{prefix}_mean_accuracy": np.nan,
            f"{prefix}_mean_balanced_accuracy": np.nan,
            f"{prefix}_mean_f1": np.nan,
            f"{prefix}_mean_roc_auc": np.nan,
            f"{prefix}_validation_selection_score": np.nan,
            f"{prefix}_n_models": np.nan,
            f"{prefix}_models_used": "",
        }

    return {
        f"{prefix}_mean_accuracy": agg.get("mean_accuracy", np.nan),
        f"{prefix}_mean_balanced_accuracy": agg.get("mean_balanced_accuracy", np.nan),
        f"{prefix}_mean_f1": agg.get("mean_f1", np.nan),
        f"{prefix}_mean_roc_auc": agg.get("mean_roc_auc", np.nan),
        f"{prefix}_validation_selection_score": agg.get("validation_selection_score", np.nan),
        f"{prefix}_n_models": agg.get("n_models", np.nan),
        f"{prefix}_models_used": agg.get("models_used", ""),
    }


# ============================================================
# 10. FEATURE ENGINEERING
# ============================================================

def add_lag_features(df, base_features, lag_days):
    new_cols = []

    for f in base_features:
        if f not in df.columns:
            continue

        if f in SET0_BASE:
            continue

        if f in EXCLUDED_FEATURES:
            continue

        for lag in lag_days:
            col = f"{f}_lag_{lag}"

            if col not in df.columns:
                df[col] = df[f].shift(lag)

            new_cols.append(col)

    return df, new_cols


def add_regime_features(df, base_features):
    new_cols = []

    for f in base_features:
        if f not in df.columns:
            continue

        if f in EXCLUDED_FEATURES:
            continue

        if f.endswith("_dummy"):
            continue

        if df[f].nunique(dropna=True) <= 3:
            continue

        low_col = f"{f}_regime_low_q25_w{REGIME_ROLLING_WINDOW}"
        high_col = f"{f}_regime_high_q75_w{REGIME_ROLLING_WINDOW}"

        if low_col not in df.columns:
            rolling_low = (
                df[f]
                .rolling(REGIME_ROLLING_WINDOW, min_periods=30)
                .quantile(REGIME_Q_LOW)
                .shift(1)
            )

            df[low_col] = np.where(df[f] < rolling_low, 1, 0)
            df.loc[rolling_low.isna(), low_col] = np.nan

        if high_col not in df.columns:
            rolling_high = (
                df[f]
                .rolling(REGIME_ROLLING_WINDOW, min_periods=30)
                .quantile(REGIME_Q_HIGH)
                .shift(1)
            )

            df[high_col] = np.where(df[f] > rolling_high, 1, 0)
            df.loc[rolling_high.isna(), high_col] = np.nan

        new_cols.extend([low_col, high_col])

    return df, new_cols


def add_interaction_features(df, base_features):
    base_features = [
        f for f in base_features
        if f in df.columns and f not in EXCLUDED_FEATURES
    ]

    base_features = unique_keep_order(base_features)
    base_features = base_features[:MAX_INTERACTION_BASE_FEATURES]

    new_cols = []

    for a, b in combinations(base_features, 2):
        col = f"interact__{safe_filename(a)}__x__{safe_filename(b)}"

        if len(new_cols) >= MAX_INTERACTION_CANDIDATES:
            break

        if col not in df.columns:
            df[col] = df[a] * df[b]

        new_cols.append(col)

    return df, new_cols


# ============================================================
# 11. SELECTION STAGES
# ============================================================

def get_regime_candidates(candidate_pool, horizon, threshold_pct, df):
    sub = candidate_pool[
        (candidate_pool["horizon"] == horizon) &
        (candidate_pool["target_threshold_pct"] == threshold_pct)
    ].copy()

    if sub.empty:
        return pd.DataFrame(columns=["feature", "feature_group", "rank"])

    sub = sub[~sub["feature"].isin(EXCLUDED_FEATURES)].copy()
    sub = sub[sub["feature"].isin(df.columns)].copy()

    if "rank" not in sub.columns:
        sub["rank"] = np.arange(1, len(sub) + 1)

    sub = sub.sort_values("rank").reset_index(drop=True)

    return sub


def run_groupwise_stage(df, base_features, candidates, target, horizon, threshold_pct):
    stage_rows = []
    summary_rows = []

    current_features = base_features.copy()

    base_eval_rows = fit_eval_feature_set(
        df,
        current_features,
        target,
        stage="2_1_A_groupwise",
        feature_set_name="SET0",
        horizon=horizon,
        threshold_pct=threshold_pct,
        extra={
            "step": "base",
            "group_added": "",
            "accepted": True,
            "heavy_models_used_here": False,
        },
        model_scope="selection",
    )

    stage_rows.extend(base_eval_rows)
    current_agg = aggregate_validation(base_eval_rows)

    selected_groups = []

    for group in GROUP_ORDER:
        group_feats = candidates[candidates["feature_group"] == group]["feature"].tolist()

        group_feats = [
            f for f in group_feats
            if f not in current_features and f in df.columns and f not in EXCLUDED_FEATURES
        ]

        group_feats = unique_keep_order(group_feats)

        if not group_feats:
            continue

        print(
            f"[GROUP] h={horizon}, thr={threshold_pct}% | "
            f"trying group={group} | n_features_add={len(group_feats)}"
        )

        proposed_features = unique_keep_order(current_features + group_feats)

        rows = fit_eval_feature_set(
            df,
            proposed_features,
            target,
            stage="2_1_A_groupwise",
            feature_set_name=f"ADD_GROUP_{group}",
            horizon=horizon,
            threshold_pct=threshold_pct,
            extra={
                "step": f"try_group_{group}",
                "group_added": group,
                "candidate_group_features": ",".join(group_feats),
                "heavy_models_used_here": False,
            },
            model_scope="selection",
        )

        stage_rows.extend(rows)
        new_agg = aggregate_validation(rows)

        accepted = is_improved(new_agg, current_agg)

        summary_rows.append({
            "horizon": horizon,
            "target_threshold_pct": threshold_pct,
            "stage": "groupwise",
            "group": group,
            "added_features": ",".join(group_feats),
            "accepted": accepted,
            **agg_to_prefixed_dict(current_agg, "before"),
            **agg_to_prefixed_dict(new_agg, "after"),
        })

        print(f"[GROUP] group={group} | accepted={accepted}")

        if accepted:
            current_features = proposed_features
            current_agg = new_agg
            selected_groups.append(group)

    return current_features, selected_groups, current_agg, stage_rows, summary_rows


def run_forward_selection_stage(
    df,
    start_features,
    candidate_features,
    target,
    horizon,
    threshold_pct,
    stage_name,
    feature_prefix,
):
    stage_rows = []
    summary_rows = []

    current_features = unique_keep_order([f for f in start_features if f in df.columns])

    base_rows = fit_eval_feature_set(
        df,
        current_features,
        target,
        stage=stage_name,
        feature_set_name=f"{feature_prefix}_START",
        horizon=horizon,
        threshold_pct=threshold_pct,
        extra={
            "step": "start",
            "feature_tried": "",
            "accepted": True,
            "heavy_models_used_here": False,
        },
        model_scope="selection",
    )

    stage_rows.extend(base_rows)
    current_agg = aggregate_validation(base_rows)

    selected_new_features = []

    for f in candidate_features:
        if f in current_features:
            continue

        if f not in df.columns:
            continue

        if f in EXCLUDED_FEATURES:
            continue

        print(
            f"[{stage_name}] h={horizon}, thr={threshold_pct}% | "
            f"trying feature={f} | current_n_features={len(current_features)}"
        )

        proposed_features = unique_keep_order(current_features + [f])

        rows = fit_eval_feature_set(
            df,
            proposed_features,
            target,
            stage=stage_name,
            feature_set_name=f"{feature_prefix}_TRY_{safe_filename(f)}",
            horizon=horizon,
            threshold_pct=threshold_pct,
            extra={
                "step": "try_feature",
                "feature_tried": f,
                "heavy_models_used_here": False,
            },
            model_scope="selection",
        )

        stage_rows.extend(rows)
        new_agg = aggregate_validation(rows)

        accepted = is_improved(new_agg, current_agg)

        summary_rows.append({
            "horizon": horizon,
            "target_threshold_pct": threshold_pct,
            "stage": stage_name,
            "feature_tried": f,
            "accepted": accepted,
            **agg_to_prefixed_dict(current_agg, "before"),
            **agg_to_prefixed_dict(new_agg, "after"),
        })

        print(f"[{stage_name}] feature={f} | accepted={accepted}")

        if accepted:
            current_features = proposed_features
            current_agg = new_agg
            selected_new_features.append(f)

    return current_features, selected_new_features, current_agg, stage_rows, summary_rows


def run_final_all_models_stage(df, final_features, target, horizon, threshold_pct):
    print(
        f"[FINAL ALL MODELS] h={horizon}, thr={threshold_pct}% | "
        f"n_features={len(final_features)} | heavy_models=ON"
    )

    rows = fit_eval_feature_set(
        df=df,
        features=final_features,
        target=target,
        stage="2_1_G_final_all_models",
        feature_set_name="FINAL_SELECTED_FEATURES_ALL_MODELS",
        horizon=horizon,
        threshold_pct=threshold_pct,
        extra={
            "step": "final_all_models_after_feature_selection",
            "feature_selection_finished": True,
            "test_used_for_selection": False,
            "heavy_models_used_here": True,
        },
        model_scope="final",
    )

    return rows


# ============================================================
# 12. CONFIDENCE / ASSURANCE
# ============================================================

def confidence_metrics(y_true, proba, confidence_threshold):
    proba = np.asarray(proba)
    y_true = np.asarray(y_true).astype(int)

    mask_finite = np.isfinite(proba)

    proba = proba[mask_finite]
    y_true = y_true[mask_finite]

    long_mask = proba >= confidence_threshold
    short_mask = proba <= (1.0 - confidence_threshold)
    signal_mask = long_mask | short_mask

    coverage = signal_mask.mean() if len(signal_mask) else np.nan
    n_signals = int(signal_mask.sum())

    if n_signals == 0:
        return {
            "confidence_threshold": confidence_threshold,
            "n_obs": len(y_true),
            "n_signals": 0,
            "coverage": 0.0,
            "signal_accuracy": np.nan,
            "signal_balanced_accuracy": np.nan,
            "signal_f1": np.nan,
            "long_signals": 0,
            "short_signals": 0,
            "long_share": np.nan,
            "short_share": np.nan,
        }

    pred = np.where(long_mask[signal_mask], 1, 0)
    y_sig = y_true[signal_mask]

    acc = accuracy_score(y_sig, pred)

    if len(np.unique(y_sig)) >= 2:
        ba = balanced_accuracy_score(y_sig, pred)
    else:
        ba = np.nan

    f1 = f1_score(y_sig, pred, zero_division=0)

    long_signals = int(long_mask.sum())
    short_signals = int(short_mask.sum())

    return {
        "confidence_threshold": confidence_threshold,
        "n_obs": len(y_true),
        "n_signals": n_signals,
        "coverage": coverage,
        "signal_accuracy": acc,
        "signal_balanced_accuracy": ba,
        "signal_f1": f1,
        "long_signals": long_signals,
        "short_signals": short_signals,
        "long_share": long_signals / n_signals if n_signals else np.nan,
        "short_share": short_signals / n_signals if n_signals else np.nan,
    }


def run_confidence_stage(df, final_features, target, horizon, threshold_pct):
    rows = []

    data = make_model_data(df, final_features, target)

    if data.empty:
        return pd.DataFrame(), pd.DataFrame()

    train, val, test = split_data(data)

    if not data_ok(train, val, test, target):
        return pd.DataFrame(), pd.DataFrame()

    X_train = train[final_features]
    y_train = train[target].astype(int).values

    X_val = val[final_features]
    y_val = val[target].astype(int).values

    X_test = test[final_features]
    y_test = test[target].astype(int).values

    print(
        f"[CONFIDENCE] h={horizon}, thr={threshold_pct}% | "
        f"n_features={len(final_features)} | heavy_models=ON"
    )

    models = get_models(y_train, model_scope="final")

    for model_name, model in models.items():
        if model_name == "Baseline_MajorityClass":
            continue

        try:
            model.fit(X_train, y_train)
        except Exception as e:
            rows.append({
                "asset": ASSET_NAME,
                "horizon": horizon,
                "target_threshold_pct": threshold_pct,
                "target": target,
                "model": model_name,
                "model_scope": "final",
                "eval_split": "fit_error",
                "error": str(e),
                "n_features": len(final_features),
                "features": ",".join(final_features),
                "is_neural_model": model_name.startswith("MLP") or model_name.startswith("LSTM"),
                "is_lstm_model": model_name.startswith("LSTM"),
                "is_xgboost_model": model_name == "XGBoost",
                "is_catboost_model": model_name == "CatBoost",
                "used_for_confidence_selection": model_name in CONFIDENCE_SELECTION_MODELS,
                "heavy_models_used_here": True,
            })
            continue

        for split_name, X_eval, y_eval in [
            ("validation", X_val, y_val),
            ("test", X_test, y_test),
        ]:
            try:
                proba = get_positive_proba(model, X_eval)

                if proba is None:
                    continue

                for c in CONFIDENCE_GRID:
                    m = confidence_metrics(y_eval, proba, c)

                    rows.append({
                        "asset": ASSET_NAME,
                        "horizon": horizon,
                        "target_threshold_pct": threshold_pct,
                        "target": target,
                        "model": model_name,
                        "model_scope": "final",
                        "eval_split": split_name,
                        "n_features": len(final_features),
                        "features": ",".join(final_features),
                        "is_neural_model": model_name.startswith("MLP") or model_name.startswith("LSTM"),
                        "is_lstm_model": model_name.startswith("LSTM"),
                        "is_xgboost_model": model_name == "XGBoost",
                        "is_catboost_model": model_name == "CatBoost",
                        "used_for_confidence_selection": model_name in CONFIDENCE_SELECTION_MODELS,
                        "heavy_models_used_here": True,
                        **m,
                    })

            except Exception as e:
                rows.append({
                    "asset": ASSET_NAME,
                    "horizon": horizon,
                    "target_threshold_pct": threshold_pct,
                    "target": target,
                    "model": model_name,
                    "model_scope": "final",
                    "eval_split": split_name + "_error",
                    "error": str(e),
                    "n_features": len(final_features),
                    "features": ",".join(final_features),
                    "is_neural_model": model_name.startswith("MLP") or model_name.startswith("LSTM"),
                    "is_lstm_model": model_name.startswith("LSTM"),
                    "is_xgboost_model": model_name == "XGBoost",
                    "is_catboost_model": model_name == "CatBoost",
                    "used_for_confidence_selection": model_name in CONFIDENCE_SELECTION_MODELS,
                    "heavy_models_used_here": True,
                })

    conf = pd.DataFrame(rows)

    if conf.empty:
        return conf, pd.DataFrame()

    selected = select_confidence_variants(conf, horizon, threshold_pct)

    return conf, selected


def select_confidence_variants(conf, horizon, threshold_pct):
    val = conf[
        (conf["eval_split"] == "validation") &
        (conf["model"].isin(CONFIDENCE_SELECTION_MODELS))
    ].copy()

    if val.empty:
        return pd.DataFrame()

    val["assurance_score"] = (
        0.55 * val["signal_accuracy"].fillna(0) +
        0.25 * val["signal_balanced_accuracy"].fillna(0) +
        0.20 * val["coverage"].fillna(0)
    )

    selected_rows = []

    def pick_variant(name, subset, fallback_subset=None):
        if subset is not None and not subset.empty:
            row = subset.sort_values(
                ["assurance_score", "signal_accuracy", "coverage"],
                ascending=[False, False, False]
            ).head(1).copy()

            row["assurance_variant"] = name
            row["variant_selection_note"] = "matched_rule_on_validation"
            return row

        if fallback_subset is not None and not fallback_subset.empty:
            row = fallback_subset.sort_values(
                ["assurance_score", "signal_accuracy", "coverage"],
                ascending=[False, False, False]
            ).head(1).copy()

            row["assurance_variant"] = name
            row["variant_selection_note"] = "fallback_best_available_on_validation"
            return row

        return pd.DataFrame()

    balanced_subset = val[
        val["coverage"] >= BALANCED_MIN_COVERAGE
    ].copy()

    practical_subset = val[
        (val["signal_accuracy"] >= PRACTICAL_MIN_ACC) &
        (val["coverage"] >= PRACTICAL_MIN_COVERAGE) &
        (val["coverage"] <= PRACTICAL_MAX_COVERAGE)
    ].copy()

    rare_subset = val[
        (val["signal_accuracy"] >= RARE_MIN_ACC) &
        (val["coverage"] >= RARE_MIN_COVERAGE) &
        (val["coverage"] < RARE_MAX_COVERAGE)
    ].copy()

    any_reasonable = val[val["coverage"] >= 0.05].copy()

    for name, subset in [
        ("balanced_coverage", balanced_subset),
        ("accuracy_60_coverage_30_50", practical_subset),
        ("rare_strong_signal", rare_subset),
    ]:
        picked = pick_variant(name, subset, any_reasonable)

        if not picked.empty:
            selected_rows.append(picked)

    if not selected_rows:
        return pd.DataFrame()

    selected_val = pd.concat(selected_rows, ignore_index=True)

    test = conf[conf["eval_split"] == "test"].copy()

    out_rows = []

    for _, row in selected_val.iterrows():
        model_name = row["model"]
        c = row["confidence_threshold"]

        test_match = test[
            (test["model"] == model_name) &
            (test["confidence_threshold"] == c)
        ]

        test_row = test_match.iloc[0].to_dict() if not test_match.empty else {}

        combined = row.to_dict()

        for k, v in test_row.items():
            if k in [
                "asset",
                "horizon",
                "target_threshold_pct",
                "target",
                "model",
                "features",
            ]:
                continue

            combined[f"test_{k}"] = v

        out_rows.append(combined)

    out = pd.DataFrame(out_rows)

    out["test_is_used_for_selection"] = False
    out["selection_basis"] = "validation_only"
    out["feature_selection_models"] = ",".join(SELECTION_MODELS)
    out["confidence_selection_models"] = ",".join(CONFIDENCE_SELECTION_MODELS)
    out["heavy_models_used_for_confidence"] = True

    return out


# ============================================================
# 13. MAIN PIPELINE
# ============================================================

def run_stage2_pipeline(df, candidate_pool):
    base_features = existing_features(df, SET0_BASE)

    all_eval_rows = []
    groupwise_summary = []
    forward_summary = []
    lag_summary = []
    regime_summary = []
    interaction_summary = []
    combined_summary = []
    final_all_models_rows = []
    confidence_rows = []
    confidence_selected_rows = []
    final_plan_rows = []

    for h in HORIZONS:
        for thr in TARGET_THRESHOLDS:
            pct = threshold_to_pct(thr)
            target = target_name(h, pct)

            print("\n" + "=" * 80)
            print(f"Running ETH stage 2.1 | horizon={h} | threshold={pct}%")
            print("=" * 80)
            print("Feature selection models:", SELECTION_MODELS)
            print("Heavy models are OFF during selection and ON only at final stages.")

            if target not in df.columns:
                print(f"SKIP: target not found: {target}")
                continue

            candidates = get_regime_candidates(candidate_pool, h, pct, df)

            if candidates.empty:
                print("No candidates for this horizon + threshold. Running SET0 only.")

            group_features, selected_groups, group_agg, rows, summary = run_groupwise_stage(
                df=df,
                base_features=base_features,
                candidates=candidates,
                target=target,
                horizon=h,
                threshold_pct=pct,
            )

            all_eval_rows.extend(rows)
            groupwise_summary.extend(summary)

            print(f"Selected groups: {selected_groups}")

            if selected_groups:
                forward_candidates = candidates[
                    candidates["feature_group"].isin(selected_groups)
                ]["feature"].tolist()
            else:
                forward_candidates = []

            forward_candidates = unique_keep_order([
                f for f in forward_candidates
                if f not in EXCLUDED_FEATURES and f in df.columns
            ])

            forward_features, selected_plain_features, forward_agg, rows, summary = run_forward_selection_stage(
                df=df,
                start_features=base_features,
                candidate_features=forward_candidates,
                target=target,
                horizon=h,
                threshold_pct=pct,
                stage_name="2_1_B_forward_selection",
                feature_prefix="FORWARD",
            )

            all_eval_rows.extend(rows)
            forward_summary.extend(summary)

            print(f"Selected plain features: {selected_plain_features}")

            df, lag_candidates = add_lag_features(df, selected_plain_features, LAG_DAYS)

            lag_features, selected_lag_features, lag_agg, rows, summary = run_forward_selection_stage(
                df=df,
                start_features=forward_features,
                candidate_features=lag_candidates,
                target=target,
                horizon=h,
                threshold_pct=pct,
                stage_name="2_1_C_lagged_determinants",
                feature_prefix="LAGS",
            )

            all_eval_rows.extend(rows)
            lag_summary.extend(summary)

            print(f"Selected lag features: {selected_lag_features}")

            regime_base = unique_keep_order(selected_plain_features + selected_lag_features)
            df, regime_candidates = add_regime_features(df, regime_base)

            regime_features, selected_regime_features, regime_agg, rows, summary = run_forward_selection_stage(
                df=df,
                start_features=lag_features,
                candidate_features=regime_candidates,
                target=target,
                horizon=h,
                threshold_pct=pct,
                stage_name="2_1_D_regime_features",
                feature_prefix="REGIME",
            )

            all_eval_rows.extend(rows)
            regime_summary.extend(summary)

            print(f"Selected regime features: {selected_regime_features}")

            interaction_base = unique_keep_order(
                selected_plain_features +
                selected_lag_features +
                selected_regime_features
            )

            df, interaction_candidates = add_interaction_features(df, interaction_base)

            interaction_features, selected_interaction_features, interaction_agg, rows, summary = run_forward_selection_stage(
                df=df,
                start_features=regime_features,
                candidate_features=interaction_candidates,
                target=target,
                horizon=h,
                threshold_pct=pct,
                stage_name="2_1_E_interaction_features",
                feature_prefix="INTERACTIONS",
            )

            all_eval_rows.extend(rows)
            interaction_summary.extend(summary)

            print(f"Selected interaction features: {selected_interaction_features}")

            combined_candidates = unique_keep_order(
                selected_plain_features +
                selected_lag_features +
                selected_regime_features +
                selected_interaction_features
            )

            combined_features, selected_combined_features, combined_agg, rows, summary = run_forward_selection_stage(
                df=df,
                start_features=base_features,
                candidate_features=combined_candidates,
                target=target,
                horizon=h,
                threshold_pct=pct,
                stage_name="2_1_F_combined_all_selected_blocks",
                feature_prefix="COMBINED",
            )

            all_eval_rows.extend(rows)
            combined_summary.extend(summary)

            print(f"Final combined features: {selected_combined_features}")

            final_features = combined_features
            final_non_set0 = [f for f in final_features if f not in base_features]

            final_rows = run_final_all_models_stage(
                df=df,
                final_features=final_features,
                target=target,
                horizon=h,
                threshold_pct=pct,
            )

            all_eval_rows.extend(final_rows)
            final_all_models_rows.extend(final_rows)
            final_all_agg = aggregate_validation(final_rows, model_names=FINAL_MODEL_NAMES)

            conf, selected_conf = run_confidence_stage(
                df=df,
                final_features=final_features,
                target=target,
                horizon=h,
                threshold_pct=pct,
            )

            if not conf.empty:
                confidence_rows.append(conf)

            if not selected_conf.empty:
                confidence_selected_rows.append(selected_conf)

            final_plan_rows.append({
                "asset": ASSET_NAME,
                "horizon": h,
                "target_threshold_pct": pct,
                "target": target,

                "base_features": ",".join(base_features),
                "selected_groups": ",".join(selected_groups),

                "selected_plain_features": ",".join(selected_plain_features),
                "selected_lag_features": ",".join(selected_lag_features),
                "selected_regime_features": ",".join(selected_regime_features),
                "selected_interaction_features": ",".join(selected_interaction_features),

                "final_features": ",".join(final_features),
                "final_non_set0_features": ",".join(final_non_set0),
                "n_final_features": len(final_features),
                "n_final_non_set0_features": len(final_non_set0),

                **agg_to_prefixed_dict(group_agg, "after_groupwise"),
                **agg_to_prefixed_dict(forward_agg, "after_forward"),
                **agg_to_prefixed_dict(lag_agg, "after_lags"),
                **agg_to_prefixed_dict(regime_agg, "after_regimes"),
                **agg_to_prefixed_dict(interaction_agg, "after_interactions"),
                **agg_to_prefixed_dict(combined_agg, "after_combined"),
                **agg_to_prefixed_dict(final_all_agg, "final_all_models"),

                "selection_basis": "validation_only",
                "test_used_for_selection": False,
                "excluded_features": ",".join(sorted(EXCLUDED_FEATURES)),
                "selection_models": ",".join(SELECTION_MODELS),
                "final_models": ",".join(FINAL_MODEL_NAMES),
                "heavy_models_used_during_feature_selection": False,
                "heavy_models_used_only_at_final": True,
                "xgboost_counted_final_only": USE_XGBOOST,
                "catboost_counted_final_only": USE_CATBOOST,
                "mlp_models_enabled_final_only": USE_MLP,
                "lstm_models_enabled_final_only": USE_LSTM,
                "torch_available": TORCH_AVAILABLE,
                "include_lstm_in_selection": INCLUDE_LSTM_IN_SELECTION,
            })

    eval_df = pd.DataFrame(all_eval_rows)
    groupwise_df = pd.DataFrame(groupwise_summary)
    forward_df = pd.DataFrame(forward_summary)
    lag_df = pd.DataFrame(lag_summary)
    regime_df = pd.DataFrame(regime_summary)
    interaction_df = pd.DataFrame(interaction_summary)
    combined_df = pd.DataFrame(combined_summary)
    final_all_models_df = pd.DataFrame(final_all_models_rows)
    final_plan_df = pd.DataFrame(final_plan_rows)

    confidence_df = pd.concat(confidence_rows, ignore_index=True) if confidence_rows else pd.DataFrame()
    confidence_selected_df = pd.concat(confidence_selected_rows, ignore_index=True) if confidence_selected_rows else pd.DataFrame()

    return {
        "eval": eval_df,
        "groupwise": groupwise_df,
        "forward": forward_df,
        "lags": lag_df,
        "regimes": regime_df,
        "interactions": interaction_df,
        "combined": combined_df,
        "final_all_models": final_all_models_df,
        "confidence": confidence_df,
        "confidence_selected": confidence_selected_df,
        "final_plan": final_plan_df,
    }


# ============================================================
# 14. SAVE REPORTS
# ============================================================

def save_outputs(outputs):
    print("\nSaving CSV outputs...")

    file_map = {
        "eval": "2_1_ALL_eval_rows.csv",
        "groupwise": "2_1_A_groupwise_results.csv",
        "forward": "2_1_B_forward_selection_results.csv",
        "lags": "2_1_C_lagged_features_results.csv",
        "regimes": "2_1_D_regime_features_results.csv",
        "interactions": "2_1_E_interaction_features_results.csv",
        "combined": "2_1_F_combined_all_blocks_results.csv",
        "final_all_models": "2_1_G_final_all_models_results.csv",
        "confidence": "2_1_H_confidence_signals_all.csv",
        "confidence_selected": "2_1_I_confidence_selected_variants.csv",
        "final_plan": "2_1_J_final_plan_by_horizon_threshold.csv",
    }

    for key, filename in file_map.items():
        df = outputs.get(key, pd.DataFrame())

        if df is not None and not df.empty:
            df.to_csv(OUTPUT_DIR / filename, index=False, encoding="utf-8-sig")
        else:
            pd.DataFrame().to_csv(OUTPUT_DIR / filename, index=False, encoding="utf-8-sig")

    write_excel_report(outputs)
    write_text_summary(outputs)


def write_excel_report(outputs):
    xlsx_path = OUTPUT_DIR / "2_1_ETH_joint_model_report_fast_selection_final_heavy.xlsx"

    with pd.ExcelWriter(xlsx_path, engine="openpyxl") as writer:
        readme = pd.DataFrame({
            "section": [
                "Описание",
                "Главное правило",
                "Asset",
                "Input sheet",
                "Thresholds",
                "Порядок этапов",
                "Candidate pool",
                "CPI",
                "Selection models",
                "Heavy models",
                "XGBoost / CatBoost",
                "MLP",
                "LSTM",
                "Confidence",
            ],
            "text": [
                "Это второй этап ETH: joint-модели по кандидатам из первого этапа.",
                "Все решения принимаются только по validation. Test не используется для выбора.",
                ASSET_NAME,
                SHEET_NAME,
                "Каждый horizon + threshold обрабатывается отдельно. Пороги не усредняются.",
                "SET0 -> groups -> forward -> lags -> regimes -> interactions -> combined -> final all models -> confidence -> test check.",
                f"Файл: {CANDIDATE_POOL_PATH}; лист: {CANDIDATE_POOL_SHEET}",
                f"Исключённые признаки: {', '.join(sorted(EXCLUDED_FEATURES))}",
                f"Для отбора признаков используются только: {', '.join(SELECTION_MODELS)}",
                "XGBoost, CatBoost, MLP и LSTM НЕ считаются на каждом шаге отбора, а включаются только на финальном наборе признаков.",
                "XGBoost и CatBoost считаются в финальном отчёте и confidence stage.",
                "MLP_small, MLP_deep, MLP_wide, MLP_regularized считаются только в финале.",
                f"LSTM_deep_seq30 и LSTM_deep_seq60 считаются только в финале. PyTorch available={TORCH_AVAILABLE}.",
                "Confidence thresholds выбираются по validation; test только проверяет перенос.",
            ]
        })

        readme.to_excel(writer, sheet_name="README", index=False)

        sheet_map = {
            "final_plan": "final_plan",
            "confidence_selected": "confidence_selected",
            "confidence": "confidence_all",
            "final_all_models": "final_all_models",
            "combined": "combined_stage",
            "interactions": "interactions",
            "regimes": "regimes",
            "lags": "lags",
            "forward": "forward_selection",
            "groupwise": "groupwise",
            "eval": "all_eval_rows",
        }

        for key, sheet in sheet_map.items():
            df = outputs.get(key, pd.DataFrame())

            if df is not None and not df.empty:
                df.to_excel(writer, sheet_name=sheet[:31], index=False)

    try:
        from openpyxl import load_workbook
        from openpyxl.styles import Font

        wb = load_workbook(xlsx_path)

        for ws in wb.worksheets:
            ws.freeze_panes = "A2"
            ws.auto_filter.ref = ws.dimensions

            for cell in ws[1]:
                cell.font = Font(bold=True)

            for col in ws.columns:
                max_len = 0
                col_letter = col[0].column_letter

                for cell in col[:200]:
                    try:
                        max_len = max(max_len, len(str(cell.value)))
                    except Exception:
                        pass

                ws.column_dimensions[col_letter].width = min(max_len + 2, 45)

        wb.save(xlsx_path)

    except Exception as e:
        print(f"Excel formatting skipped: {e}")

    print(f"Excel saved: {xlsx_path}")


def write_text_summary(outputs):
    path = OUTPUT_DIR / "2_1_summary.txt"

    final_plan = outputs.get("final_plan", pd.DataFrame())
    confidence_selected = outputs.get("confidence_selected", pd.DataFrame())
    eval_df = outputs.get("eval", pd.DataFrame())
    final_all_models = outputs.get("final_all_models", pd.DataFrame())

    with open(path, "w", encoding="utf-8") as f:
        f.write("ETH — STAGE 2.1 JOINT MODEL PIPELINE\n")
        f.write("FAST FEATURE SELECTION + HEAVY MODELS ONLY AT FINAL\n")
        f.write("=" * 80 + "\n\n")

        f.write("Это НЕ подбор по test.\n")
        f.write("Все решения по признакам принимаются только по validation.\n")
        f.write("TEST IS NOT USED FOR SELECTION.\n\n")

        f.write("Пороговые задачи НЕ смешивались.\n")
        f.write("Каждая пара horizon + threshold обрабатывалась отдельно.\n\n")

        f.write("Input:\n")
        f.write(f"- file: {FILE_PATH}\n")
        f.write(f"- sheet: {SHEET_NAME}\n\n")

        f.write("Candidate pool:\n")
        f.write(f"- file: {CANDIDATE_POOL_PATH}\n")
        f.write(f"- sheet: {CANDIDATE_POOL_SHEET}\n\n")

        f.write("Порядок этапов:\n")
        f.write("1. SET0\n")
        f.write("2. Groups: market -> behavior -> onchain -> macro\n")
        f.write("3. Forward selection внутри выбранных групп\n")
        f.write("4. Лаги выбранных детерминантов\n")
        f.write("5. Regime-признаки\n")
        f.write("6. Interaction-признаки\n")
        f.write("7. Объединение всех удачных проверок в одну модель\n")
        f.write("8. Финальный прогон всех моделей: XGBoost/CatBoost/MLP/LSTM включаются здесь\n")
        f.write("9. Confidence / assurance signals на финальном наборе\n")
        f.write("10. Test только после финального выбора по validation\n\n")

        f.write(f"Исключённые признаки: {', '.join(sorted(EXCLUDED_FEATURES))}\n\n")

        f.write("Модели, которые участвуют в отборе признаков:\n")
        for m in SELECTION_MODELS:
            f.write(f"- {m}\n")

        f.write("\nТяжёлые модели, которые считаются только в финале:\n")
        for m in FINAL_MODEL_NAMES:
            if m not in SELECTION_MODELS:
                f.write(f"- {m}\n")

        f.write("\nНейронные сети:\n")
        f.write("- MLP_small: (64, 32), final only\n")
        f.write("- MLP_deep: (128, 64, 32, 16), final only\n")
        f.write("- MLP_wide: (256, 128, 64), final only\n")
        f.write("- MLP_regularized: (128, 64, 32), stronger regularization, final only\n")
        f.write("- LSTM_deep_seq30: deep LSTM with 30-day sequences, final only\n")
        f.write("- LSTM_deep_seq60: deep LSTM with 60-day sequences, final only\n")
        f.write(f"PyTorch available: {TORCH_AVAILABLE}\n\n")

        if eval_df is not None and not eval_df.empty:
            if "model_scope" in eval_df.columns:
                f.write("Model scope counts:\n")
                counts = eval_df["model_scope"].value_counts(dropna=False)
                for scope, cnt in counts.items():
                    f.write(f"- {scope}: {cnt}\n")
                f.write("\n")

            xgb_rows = eval_df[eval_df["model"].astype(str) == "XGBoost"].copy()
            cat_rows = eval_df[eval_df["model"].astype(str) == "CatBoost"].copy()
            mlp_rows = eval_df[eval_df["model"].astype(str).str.startswith("MLP")].copy()
            lstm_rows = eval_df[eval_df["model"].astype(str).str.startswith("LSTM")].copy()

            f.write("Heavy model diagnostics:\n")
            f.write(f"- XGBoost rows: {len(xgb_rows)}\n")
            f.write(f"- CatBoost rows: {len(cat_rows)}\n")
            f.write(f"- MLP rows: {len(mlp_rows)}\n")
            f.write(f"- LSTM rows: {len(lstm_rows)}\n")
            f.write("- Heavy models are counted only at final_all_models and confidence stages.\n\n")

        if final_all_models is not None and not final_all_models.empty:
            f.write("FINAL ALL MODELS summary:\n")
            val = final_all_models[final_all_models["eval_split"] == "validation"].copy()
            if not val.empty:
                by_model = (
                    val.groupby("model")[["accuracy", "balanced_accuracy", "f1", "roc_auc"]]
                    .mean()
                    .sort_values("roc_auc", ascending=False)
                )
                for model, row in by_model.iterrows():
                    f.write(
                        f"- {model}: acc={row['accuracy']:.4f}, "
                        f"BA={row['balanced_accuracy']:.4f}, "
                        f"F1={row['f1']:.4f}, AUC={row['roc_auc']:.4f}\n"
                    )
                f.write("\n")

        if final_plan is not None and not final_plan.empty:
            f.write("ФИНАЛЬНЫЕ НАБОРЫ ПО horizon + threshold\n")
            f.write("-" * 80 + "\n")

            for _, row in final_plan.iterrows():
                f.write(
                    f"h={row['horizon']}, threshold={row['target_threshold_pct']}%:\n"
                    f"  selected_groups: {row.get('selected_groups', '')}\n"
                    f"  final_non_SET0: {row.get('final_non_set0_features', '')}\n"
                    f"  n_final_features: {row.get('n_final_features', '')}\n"
                    f"  validation score after combined: {row.get('after_combined_validation_selection_score', np.nan)}\n"
                    f"  validation score final all models: {row.get('final_all_models_validation_selection_score', np.nan)}\n\n"
                )

        if confidence_selected is not None and not confidence_selected.empty:
            f.write("\nВЫБРАННЫЕ CONFIDENCE / ASSURANCE РЕЖИМЫ\n")
            f.write("-" * 80 + "\n")

            for _, row in confidence_selected.iterrows():
                f.write(
                    f"h={row.get('horizon')}, thr={row.get('target_threshold_pct')}%, "
                    f"variant={row.get('assurance_variant')}, model={row.get('model')}, "
                    f"conf={row.get('confidence_threshold')}, "
                    f"val_acc={row.get('signal_accuracy')}, val_cov={row.get('coverage')}, "
                    f"test_acc={row.get('test_signal_accuracy')}, test_cov={row.get('test_coverage')}\n"
                )

        f.write("\nСозданные файлы:\n")
        f.write("- 2_1_ALL_eval_rows.csv\n")
        f.write("- 2_1_A_groupwise_results.csv\n")
        f.write("- 2_1_B_forward_selection_results.csv\n")
        f.write("- 2_1_C_lagged_features_results.csv\n")
        f.write("- 2_1_D_regime_features_results.csv\n")
        f.write("- 2_1_E_interaction_features_results.csv\n")
        f.write("- 2_1_F_combined_all_blocks_results.csv\n")
        f.write("- 2_1_G_final_all_models_results.csv\n")
        f.write("- 2_1_H_confidence_signals_all.csv\n")
        f.write("- 2_1_I_confidence_selected_variants.csv\n")
        f.write("- 2_1_J_final_plan_by_horizon_threshold.csv\n")
        f.write("- 2_1_ETH_joint_model_report_fast_selection_final_heavy.xlsx\n")

    print(f"Summary saved: {path}")


# ============================================================
# 15. PLOTS
# ============================================================

def make_plots(outputs):
    final_plan = outputs.get("final_plan", pd.DataFrame())
    confidence_selected = outputs.get("confidence_selected", pd.DataFrame())

    if final_plan is not None and not final_plan.empty:
        plot_df = final_plan.copy()

        plot_df["regime"] = (
            "h=" + plot_df["horizon"].astype(str) +
            ",thr=" + plot_df["target_threshold_pct"].astype(str) + "%"
        )

        plot_df["score"] = plot_df["after_combined_validation_selection_score"]

        plot_df = plot_df.sort_values("score", ascending=False).head(25)

        plt.figure(figsize=(13, 8))
        plt.barh(plot_df["regime"][::-1], plot_df["score"][::-1])
        plt.title("ETH stage 2.1: top regimes by validation score after combined stage")
        plt.xlabel("Validation selection score")
        plt.tight_layout()
        plt.savefig(PLOTS_DIR / "top_regimes_after_combined_validation_score.png", dpi=180)
        plt.close()

    if confidence_selected is not None and not confidence_selected.empty:
        conf = confidence_selected.copy()

        plt.figure(figsize=(13, 8))
        plt.scatter(conf["coverage"], conf["signal_accuracy"])

        for _, row in conf.iterrows():
            plt.text(
                row["coverage"],
                row["signal_accuracy"],
                str(row["horizon"]) + "/" + str(row["target_threshold_pct"]),
                fontsize=7
            )

        plt.axhline(0.60, linewidth=1)
        plt.axvline(0.30, linewidth=1)
        plt.axvline(0.50, linewidth=1)
        plt.title("ETH confidence regimes: validation accuracy vs coverage")
        plt.xlabel("Coverage")
        plt.ylabel("Signal accuracy")
        plt.tight_layout()
        plt.savefig(PLOTS_DIR / "confidence_validation_accuracy_vs_coverage.png", dpi=180)
        plt.close()


# ============================================================
# 16. MAIN
# ============================================================

def main():
    ensure_dirs()

    print("ETH Stage 2.1 joint pipeline")
    print("Fast feature selection + heavy models only at final")
    print(f"Output folder: {OUTPUT_DIR}")

    logs = []
    logs.append(f"ASSET_NAME={ASSET_NAME}")
    logs.append(f"FILE_PATH={FILE_PATH}")
    logs.append(f"SHEET_NAME={SHEET_NAME}")
    logs.append(f"OUTPUT_DIR={OUTPUT_DIR}")
    logs.append(f"STAGE1_DIR={STAGE1_DIR}")
    logs.append(f"CANDIDATE_POOL_PATH={CANDIDATE_POOL_PATH}")
    logs.append(f"CANDIDATE_POOL_SHEET={CANDIDATE_POOL_SHEET}")
    logs.append(f"EXCLUDED_FEATURES={sorted(EXCLUDED_FEATURES)}")
    logs.append(f"SELECTION_MODELS={SELECTION_MODELS}")
    logs.append(f"FINAL_MODEL_NAMES={FINAL_MODEL_NAMES}")
    logs.append("Heavy models are used only at final_all_models and confidence stages.")
    logs.append(f"USE_XGBOOST={USE_XGBOOST}")
    logs.append(f"USE_CATBOOST={USE_CATBOOST}")
    logs.append(f"USE_MLP={USE_MLP}")
    logs.append(f"USE_LSTM={USE_LSTM}")
    logs.append(f"TORCH_AVAILABLE={TORCH_AVAILABLE}")
    logs.append(f"INCLUDE_LSTM_IN_SELECTION={INCLUDE_LSTM_IN_SELECTION}")
    logs.append("TEST IS NOT USED FOR SELECTION.")
    logs.append("Thresholds are processed separately.")

    df = load_ml_data()

    print("\nSET0 features found:")
    print(existing_features(df, SET0_BASE))

    print("\nETH feature groups found/missing:")
    for group, feats in FEATURE_GROUP_DICT.items():
        if group == "SET0":
            continue
        print(f"\n{group}:")
        print("found:", existing_features(df, feats))
        print("missing:", [f for f in feats if f not in df.columns])

    print("\nLoading candidate pool...")
    candidate_pool = load_candidate_pool()

    print("\nCandidate pool preview:")
    print(candidate_pool[["horizon", "target_threshold_pct", "rank", "feature", "feature_group"]].head(20))

    print("\nRunning stage 2.1...")
    outputs = run_stage2_pipeline(df, candidate_pool)

    print("\nSaving outputs...")
    save_outputs(outputs)

    print("\nSaving plots...")
    make_plots(outputs)

    write_log(logs)

    print("\nDONE.")
    print(f"Results saved to: {OUTPUT_DIR}")
    print("\nMain files:")
    print("- 2_1_ETH_joint_model_report_fast_selection_final_heavy.xlsx")
    print("- 2_1_summary.txt")
    print("- 2_1_J_final_plan_by_horizon_threshold.csv")
    print("- 2_1_I_confidence_selected_variants.csv")
    print("- 2_1_H_confidence_signals_all.csv")
    print("- 2_1_G_final_all_models_results.csv")
    print("- 2_1_ALL_eval_rows.csv")


if __name__ == "__main__":
    main()