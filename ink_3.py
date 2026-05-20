"""
LINK Stage 3.0 — Final Walk-Forward Stability Check
====================================================

Что делает:
1. Берёт финальные режимы из LINK Stage 2.1:
   - horizon
   - target_threshold_pct
   - final_features

2. НЕ подбирает признаки заново.
   Это важно: Stage 3 — не selection, а честная проверка устойчивости.

3. Делает walk-forward проверку:
   train до конца прошлого года -> test следующий год.

4. Проверяет:
   - обычные модели на полном покрытии;
   - confidence-сигналы из Stage 2.1;
   - переносимость результатов во времени.

5. Создаёт:
   - CSV-файлы;
   - Excel-отчёт;
   - PNG-графики;
   - summary txt.

Важно для LINK:
    Данные LINK начинаются примерно с 2017-09-20.
    Поэтому walk-forward train_start начинается с 2017-09-20, а не с 2015.

Установка:
    pip install pandas numpy openpyxl matplotlib scikit-learn

Опционально:
    pip install xgboost catboost

Запуск:
    python link_3_0_walk_forward_final_check.py
"""

import re
import warnings
from pathlib import Path

import numpy as np
import pandas as pd

import matplotlib
matplotlib.use("Agg")
import matplotlib.pyplot as plt

from sklearn.dummy import DummyClassifier
from sklearn.linear_model import LogisticRegression
from sklearn.ensemble import RandomForestClassifier, ExtraTreesClassifier
from sklearn.pipeline import Pipeline
from sklearn.preprocessing import StandardScaler
from sklearn.metrics import (
    accuracy_score,
    balanced_accuracy_score,
    f1_score,
    roc_auc_score,
)

warnings.filterwarnings("ignore")


# ============================================================
# 1. SETTINGS LINK
# ============================================================

ASSET_NAME = "LINK"

FILE_PATH = r"C:\Users\mrsas\PycharmProjects\BTC\BTC_pr\БОЛЬШАЯ БАЗА ДАННЫХ ПО ВСЕМ.xlsx"
SHEET_NAME = "ML_LINK"

BASE_DIR = Path(FILE_PATH).parent

# Основная папка второй стадии LINK.
STAGE2_DIR = BASE_DIR / "link_2_1_joint_model_outputs_fast_selection_final_heavy"

FINAL_PLAN_PATH = STAGE2_DIR / "2_1_J_final_plan_by_horizon_threshold.csv"
CONFIDENCE_SELECTED_PATH = STAGE2_DIR / "2_1_I_confidence_selected_variants.csv"
FINAL_ALL_MODELS_PATH = STAGE2_DIR / "2_1_G_final_all_models_results.csv"

OUTPUT_DIR = BASE_DIR / "link_3_0_walk_forward_final_check_outputs"
PLOTS_DIR = OUTPUT_DIR / "plots_for_diploma_stage_3_0"

OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
PLOTS_DIR.mkdir(parents=True, exist_ok=True)

# Если None — проверяем все режимы из Stage 2.1.
# Если поставить, например, 12 — возьмём top-12 по after_combined score.
TOP_N_REGIMES = None

# LSTM специально не включаем в Stage 3 по умолчанию:
# она дорогая, уменьшает эффективную выборку и уже была нестабильна.
RUN_MODELS = [
    "LogisticRegression",
    "RandomForest",
    "ExtraTrees",
    "XGBoost",
    "CatBoost",
]

# Можно добавить MLP_small, но будет дольше:
# RUN_MODELS = ["LogisticRegression", "RandomForest", "ExtraTrees", "XGBoost", "CatBoost", "MLP_small"]

RANDOM_STATE = 42

HORIZONS = [1, 3, 7, 14, 30]
TARGET_THRESHOLDS = [0.00, 0.01, 0.02, 0.03, 0.05]

MIN_TRAIN_ROWS = 300
MIN_TEST_ROWS = 50

# Для LINK train_start должен начинаться с реального старта данных.
LINK_SAMPLE_START = "2017-09-20"

# Walk-forward:
# train расширяется, test — следующий год.
WALK_FORWARD_FOLDS = [
    {
        "fold": "WF_2022",
        "train_start": LINK_SAMPLE_START,
        "train_end": "2021-12-31",
        "test_start": "2022-01-01",
        "test_end": "2022-12-31",
    },
    {
        "fold": "WF_2023",
        "train_start": LINK_SAMPLE_START,
        "train_end": "2022-12-31",
        "test_start": "2023-01-01",
        "test_end": "2023-12-31",
    },
    {
        "fold": "WF_2024",
        "train_start": LINK_SAMPLE_START,
        "train_end": "2023-12-31",
        "test_start": "2024-01-01",
        "test_end": "2024-12-31",
    },
    {
        "fold": "WF_2025",
        "train_start": LINK_SAMPLE_START,
        "train_end": "2024-12-31",
        "test_start": "2025-01-01",
        "test_end": "2025-12-31",
    },
    {
        "fold": "WF_2026_Q1",
        "train_start": LINK_SAMPLE_START,
        "train_end": "2025-12-31",
        "test_start": "2026-01-01",
        "test_end": "2026-03-31",
    },
]

SET0_BASE = [
    "log_return_lag_1",
    "log_return_lag_2",
    "log_return_lag_3",
    "rolling_vol_7",
    "rolling_vol_14",
    "rolling_vol_30",
]

REGIME_ROLLING_WINDOW = 90
REGIME_Q_LOW = 0.25
REGIME_Q_HIGH = 0.75


# ============================================================
# 2. COLUMN MAP LINK
# ============================================================

RENAME_MAP = {
    "Дата": "date",
    "date": "date",

    "Лог-доходность": "log_return",
    "log_return": "log_return",

    "log_return_lag_1": "log_return_lag_1",
    "log_return_lag_2": "log_return_lag_2",
    "log_return_lag_3": "log_return_lag_3",

    "atr_14": "atr_14",
    "ATR_14": "atr_14",
    "rolling_vol_7": "rolling_vol_7",
    "rolling_vol_14": "rolling_vol_14",
    "rolling_vol_30": "rolling_vol_30",

    # Правильный asset-specific market cap признак для LINK.
    "LINK_mcap_log_return": "link_mcap_log_return",
    "link_mcap_log_return": "link_mcap_log_return",

    # Fallback: если в Excel осталась старая ETH-колонка,
    # но внутри уже лежит LINK-значение, код не упадёт.
    "ETH_mcap_log_return": "link_mcap_log_return",
    "eth_mcap_log_return": "link_mcap_log_return",

    "total_mcap_log_return": "total_mcap_log_return",
    "btc_dominance_change": "btc_dominance_change",

    "RSI (14)": "rsi_14",
    "RSI_14": "rsi_14",
    "rsi_14": "rsi_14",

    "rsi_oversold_dummy": "rsi_oversold_dummy",
    "rsi_overbought_dummy": "rsi_overbought_dummy",
    "sma_ratio": "sma_ratio",

    "Индекс страха и жадности": "fear_greed_index",
    "fear_greed_index": "fear_greed_index",
    "fear_greed_change": "fear_greed_change",
    "fear_dummy": "fear_dummy",
    "greed_dummy": "greed_dummy",

    "Поисковая активность (Google Trends)": "google_trends",
    "Поисковая активность Chainlink (Google Trends)": "google_trends",
    "google_trends": "google_trends",

    "candle_body": "candle_body",
    "ohlc_range": "ohlc_range",
    "volume_growth": "volume_growth",
    "Давление покупателей/продавцов (buy/sell pressure)": "buy_sell_pressure",
    "buy_sell_pressure": "buy_sell_pressure",

    "transactions_log_return": "transactions_log_return",
    "fees_log_return": "fees_log_return",
    "active_addresses_log_return": "active_addresses_log_return",
    "avg_fee_log_return": "avg_fee_log_return",

    "sp500_return": "sp500_return",
    "dxy_return": "dxy_return",
    "Процентные ставки (ФРС)": "fed_rate",
    "fed_rate": "fed_rate",
    "Инфляция (CPI) США": "us_cpi_inflation",
    "us_cpi_inflation": "us_cpi_inflation",
    "gold_log_return": "gold_log_return",

    "target_return_1d": "target_return_1d",
    "target_direction_1d": "target_direction_1d",
    "target_return_3d": "target_return_3d",
    "target_direction_3d": "target_direction_3d",
    "target_return_7d": "target_return_7d",
    "target_direction_7d": "target_direction_7d",
    "target_return_14d": "target_return_14d",
    "target_direction_14d": "target_direction_14d",
    "target_return_30d": "target_return_30d",
    "target_direction_30d": "target_direction_30d",
}


# ============================================================
# 3. UTILS
# ============================================================

def clean_col_name(name):
    name = str(name)
    name = name.replace("\n", " ").replace("\r", " ")
    name = name.replace('"', "").replace("'", "")
    name = re.sub(r"\s+", " ", name)
    return name.strip()


def fix_duplicate_target_columns(cols):
    """
    Иногда в Excel бывают дубли:
        target_return_1d, target_return_1d
    Второй дубль считаем target_direction_1d.
    """
    fixed = []
    target_seen = {}

    for col in cols:
        c = clean_col_name(col)
        base = re.sub(r"\.\d+$", "", c)

        m = re.match(r"^target_return_(\d+)d$", base)

        if m:
            h = m.group(1)

            if base not in target_seen:
                target_seen[base] = 0
                fixed.append(f"target_return_{h}d")
            else:
                target_seen[base] += 1
                fixed.append(f"target_direction_{h}d")
        else:
            fixed.append(c)

    return fixed


def collapse_duplicate_columns(df):
    out = {}

    for i, col in enumerate(df.columns):
        s = df.iloc[:, i]

        if col in out:
            out[col] = out[col].combine_first(s)
        else:
            out[col] = s

    return pd.DataFrame(out)


def to_numeric_safe(s):
    if pd.api.types.is_numeric_dtype(s):
        return s

    s = s.astype(str)
    s = s.str.replace("\u00a0", "", regex=False)
    s = s.str.replace(" ", "", regex=False)
    s = s.str.replace("%", "", regex=False)
    s = s.str.replace("$", "", regex=False)
    s = s.str.replace("−", "-", regex=False)
    s = s.str.replace(",", ".", regex=False)
    s = s.replace({"nan": np.nan, "None": np.nan, "": np.nan})

    return pd.to_numeric(s, errors="coerce")


def threshold_to_pct(thr):
    return int(round(float(thr) * 100))


def target_name(horizon, threshold_pct):
    return f"target_direction_{int(horizon)}d_thr_{int(threshold_pct)}pct"


def parse_feature_list(x):
    if pd.isna(x):
        return []

    text = str(x).strip()

    if not text:
        return []

    return [f.strip() for f in text.split(",") if f.strip()]


def safe_auc(y_true, proba):
    y_true = np.asarray(y_true)
    proba = np.asarray(proba)

    mask = np.isfinite(proba)

    y_true = y_true[mask]
    proba = proba[mask]

    if len(y_true) < 2:
        return np.nan

    if len(np.unique(y_true)) < 2:
        return np.nan

    try:
        return roc_auc_score(y_true, proba)
    except Exception:
        return np.nan


def calc_metrics(y_true, pred, proba):
    y_true = np.asarray(y_true)
    pred = np.asarray(pred)
    proba = np.asarray(proba)

    mask = np.isfinite(proba)
    y_true = y_true[mask]
    pred = pred[mask]
    proba = proba[mask]

    if len(y_true) == 0:
        return {
            "accuracy": np.nan,
            "balanced_accuracy": np.nan,
            "f1": np.nan,
            "roc_auc": np.nan,
            "effective_eval_rows": 0,
        }

    return {
        "accuracy": accuracy_score(y_true, pred),
        "balanced_accuracy": balanced_accuracy_score(y_true, pred)
        if len(np.unique(y_true)) >= 2 else np.nan,
        "f1": f1_score(y_true, pred, zero_division=0),
        "roc_auc": safe_auc(y_true, proba),
        "effective_eval_rows": len(y_true),
    }


def get_positive_proba(model, X):
    if not hasattr(model, "predict_proba"):
        return np.full(len(X), np.nan)

    proba = model.predict_proba(X)

    classes = getattr(model, "classes_", None)

    if classes is None and hasattr(model, "steps"):
        classes = getattr(model.steps[-1][1], "classes_", None)

    if classes is None:
        return proba[:, 1] if proba.shape[1] == 2 else np.full(len(X), np.nan)

    classes = list(classes)

    if 1 in classes:
        return proba[:, classes.index(1)]

    return np.zeros(len(X))


def make_safe_sheet_name(name):
    name = str(name)
    name = re.sub(r"[\[\]\:\*\?\/\\]", "_", name)
    return name[:31]


def resolve_stage2_paths():
    """
    Проверяет, что папка второй стадии LINK найдена.
    Если стандартный путь не найден, пробует найти папку по файлу final_plan.
    """

    global STAGE2_DIR
    global FINAL_PLAN_PATH
    global CONFIDENCE_SELECTED_PATH
    global FINAL_ALL_MODELS_PATH

    if FINAL_PLAN_PATH.exists():
        return

    print("WARNING: стандартная папка LINK Stage 2 не найдена:")
    print(STAGE2_DIR)
    print("Пробую найти 2_1_J_final_plan_by_horizon_threshold.csv внутри BASE_DIR...")

    candidates = list(BASE_DIR.rglob("2_1_J_final_plan_by_horizon_threshold.csv"))

    if not candidates:
        raise FileNotFoundError(
            "Не найден файл 2_1_J_final_plan_by_horizon_threshold.csv. "
            "Проверь, что Stage 2.1 для LINK уже был запущен."
        )

    # Предпочитаем путь, где есть link в названии папки.
    link_candidates = [
        p for p in candidates
        if "link" in str(p.parent).lower()
    ]

    chosen = link_candidates[0] if link_candidates else candidates[0]

    STAGE2_DIR = chosen.parent
    FINAL_PLAN_PATH = STAGE2_DIR / "2_1_J_final_plan_by_horizon_threshold.csv"
    CONFIDENCE_SELECTED_PATH = STAGE2_DIR / "2_1_I_confidence_selected_variants.csv"
    FINAL_ALL_MODELS_PATH = STAGE2_DIR / "2_1_G_final_all_models_results.csv"

    print("Найдена папка Stage 2:")
    print(STAGE2_DIR)


# ============================================================
# 4. LOAD DATA
# ============================================================

def read_excel_with_fallback():
    df = pd.read_excel(FILE_PATH, sheet_name=SHEET_NAME, header=1)
    cols_clean = [clean_col_name(c) for c in df.columns]

    if "Дата" in cols_clean or "date" in cols_clean:
        return df

    print("WARNING: header=1 did not find date column. Trying header=0.")
    return pd.read_excel(FILE_PATH, sheet_name=SHEET_NAME, header=0)


def create_targets(df):
    for h in HORIZONS:
        ret_col = f"target_return_{h}d"

        if ret_col not in df.columns:
            print(f"WARNING: no {ret_col}")
            continue

        for thr in TARGET_THRESHOLDS:
            pct = threshold_to_pct(thr)
            target_col = target_name(h, pct)

            if thr == 0:
                df[target_col] = np.where(
                    df[ret_col] > 0,
                    1,
                    np.where(df[ret_col] <= 0, 0, np.nan)
                )
            else:
                df[target_col] = np.where(
                    df[ret_col] > thr,
                    1,
                    np.where(df[ret_col] < -thr, 0, np.nan)
                )

    return df


def load_ml_data():
    print("Loading LINK ML data...")

    df = read_excel_with_fallback()
    df.columns = fix_duplicate_target_columns(df.columns)

    keep_cols = []

    for c in df.columns:
        if str(c).lower().startswith("unnamed") and df[c].isna().all():
            continue
        keep_cols.append(c)

    df = df[keep_cols].copy()

    rename_actual = {}

    for c in df.columns:
        base = clean_col_name(c)

        if base in RENAME_MAP:
            rename_actual[c] = RENAME_MAP[base]

    df = df.rename(columns=rename_actual)
    df = collapse_duplicate_columns(df)

    if "date" not in df.columns:
        raise ValueError("Не найдена колонка date. Проверь колонку 'Дата' в листе ML_LINK.")

    df["date"] = pd.to_datetime(df["date"], errors="coerce", dayfirst=True)
    df = df.dropna(subset=["date"])
    df = df.sort_values("date").reset_index(drop=True)

    # Для LINK не используем доисторический период до появления данных.
    df = df[df["date"] >= pd.Timestamp(LINK_SAMPLE_START)].copy()

    for c in df.columns:
        if c != "date":
            df[c] = to_numeric_safe(df[c])

    df = create_targets(df)
    df = df.replace([np.inf, -np.inf], np.nan)

    print(f"Rows: {len(df)}")
    print(f"Date range: {df['date'].min()} — {df['date'].max()}")

    return df


def load_stage2_files():
    resolve_stage2_paths()

    if not FINAL_PLAN_PATH.exists():
        raise FileNotFoundError(f"Не найден final plan: {FINAL_PLAN_PATH}")

    final_plan = pd.read_csv(FINAL_PLAN_PATH)

    if CONFIDENCE_SELECTED_PATH.exists():
        confidence_selected = pd.read_csv(CONFIDENCE_SELECTED_PATH)
    else:
        confidence_selected = pd.DataFrame()
        print(f"WARNING: confidence selected file not found: {CONFIDENCE_SELECTED_PATH}")

    if FINAL_ALL_MODELS_PATH.exists():
        final_all_models = pd.read_csv(FINAL_ALL_MODELS_PATH)
    else:
        final_all_models = pd.DataFrame()
        print(f"WARNING: final all models file not found: {FINAL_ALL_MODELS_PATH}")

    return final_plan, confidence_selected, final_all_models


# ============================================================
# 5. RECREATE ENGINEERED FEATURES FROM STAGE 2.1
# ============================================================

def ensure_feature_exists(df, feature, warnings_list):
    """
    В Stage 2.1 признаки могли быть созданы внутри кода:
    - lag
    - regime
    - interaction

    Здесь мы их пересоздаём, чтобы проверить финальные наборы.
    """

    if feature in df.columns:
        return True

    # fallback для старого ETH-названия, если оно вдруг осталось в final_features
    if feature in {"eth_mcap_log_return", "ETH_mcap_log_return"}:
        if "link_mcap_log_return" in df.columns:
            df[feature] = df["link_mcap_log_return"]
            warnings_list.append(
                f"Feature {feature} was mapped to link_mcap_log_return for LINK."
            )
            return True

    if feature == "link_mcap_log_return":
        for old_col in ["eth_mcap_log_return", "ETH_mcap_log_return"]:
            if old_col in df.columns:
                df[feature] = df[old_col]
                warnings_list.append(
                    f"link_mcap_log_return was recreated from old column {old_col}."
                )
                return True

    # 1) lag feature: atr_14_lag_2
    m_lag = re.match(r"^(.*)_lag_(\d+)$", feature)

    if m_lag:
        base = m_lag.group(1)
        lag = int(m_lag.group(2))

        if ensure_feature_exists(df, base, warnings_list):
            df[feature] = df[base].shift(lag)
            return True

        warnings_list.append(f"Cannot create lag feature {feature}: base {base} not found.")
        return False

    # 2) regime low/high:
    # atr_14_regime_low_q25_w90
    # atr_14_regime_high_q75_w90
    m_reg = re.match(r"^(.*)_regime_(low|high)_q(25|75)_w(\d+)$", feature)

    if m_reg:
        base = m_reg.group(1)
        side = m_reg.group(2)
        q = int(m_reg.group(3)) / 100.0
        window = int(m_reg.group(4))

        if ensure_feature_exists(df, base, warnings_list):
            rolling_q = (
                df[base]
                .rolling(window, min_periods=30)
                .quantile(q)
                .shift(1)
            )

            if side == "low":
                df[feature] = np.where(df[base] < rolling_q, 1, 0)
            else:
                df[feature] = np.where(df[base] > rolling_q, 1, 0)

            df.loc[rolling_q.isna(), feature] = np.nan
            return True

        warnings_list.append(f"Cannot create regime feature {feature}: base {base} not found.")
        return False

    # 3) interaction:
    # interact__atr_14__x__atr_14_regime_low_q25_w90
    if feature.startswith("interact__") and "__x__" in feature:
        raw = feature.replace("interact__", "", 1)
        parts = raw.split("__x__")

        if len(parts) == 2:
            a, b = parts[0], parts[1]

            ok_a = ensure_feature_exists(df, a, warnings_list)
            ok_b = ensure_feature_exists(df, b, warnings_list)

            if ok_a and ok_b:
                df[feature] = df[a] * df[b]
                return True

            warnings_list.append(f"Cannot create interaction {feature}: {a} or {b} not found.")
            return False

    warnings_list.append(f"Feature not found and cannot be recreated: {feature}")
    return False


def ensure_all_features(df, features, warnings_list):
    ok_features = []

    for f in features:
        if ensure_feature_exists(df, f, warnings_list):
            ok_features.append(f)

    return ok_features


# ============================================================
# 6. MODELS
# ============================================================

def get_models(y_train):
    models = {}

    models["Baseline_MajorityClass"] = DummyClassifier(strategy="most_frequent")

    models["LogisticRegression"] = Pipeline([
        ("scaler", StandardScaler()),
        ("model", LogisticRegression(
            max_iter=5000,
            solver="lbfgs",
            class_weight="balanced",
            random_state=RANDOM_STATE,
        ))
    ])

    models["RandomForest"] = RandomForestClassifier(
        n_estimators=400,
        max_depth=4,
        min_samples_leaf=20,
        class_weight="balanced",
        random_state=RANDOM_STATE,
        n_jobs=-1,
    )

    models["ExtraTrees"] = ExtraTreesClassifier(
        n_estimators=400,
        max_depth=4,
        min_samples_leaf=20,
        class_weight="balanced",
        random_state=RANDOM_STATE,
        n_jobs=-1,
    )

    if "MLP_small" in RUN_MODELS:
        try:
            from sklearn.neural_network import MLPClassifier

            models["MLP_small"] = Pipeline([
                ("scaler", StandardScaler()),
                ("model", MLPClassifier(
                    hidden_layer_sizes=(64, 32),
                    activation="relu",
                    solver="adam",
                    alpha=0.001,
                    learning_rate_init=0.001,
                    max_iter=700,
                    early_stopping=True,
                    validation_fraction=0.15,
                    n_iter_no_change=25,
                    random_state=RANDOM_STATE,
                ))
            ])
        except Exception as e:
            print(f"MLP_small skipped: {e}")

    if "XGBoost" in RUN_MODELS:
        try:
            from xgboost import XGBClassifier

            pos = int((y_train == 1).sum())
            neg = int((y_train == 0).sum())
            scale_pos_weight = neg / pos if pos > 0 else 1.0

            models["XGBoost"] = XGBClassifier(
                n_estimators=500,
                max_depth=3,
                learning_rate=0.03,
                subsample=0.8,
                colsample_bytree=0.8,
                min_child_weight=5,
                reg_lambda=5.0,
                reg_alpha=1.0,
                eval_metric="logloss",
                random_state=RANDOM_STATE,
                n_jobs=-1,
                scale_pos_weight=scale_pos_weight,
            )
        except Exception as e:
            print(f"XGBoost skipped: {e}")

    if "CatBoost" in RUN_MODELS:
        try:
            from catboost import CatBoostClassifier

            models["CatBoost"] = CatBoostClassifier(
                iterations=500,
                depth=4,
                learning_rate=0.03,
                loss_function="Logloss",
                eval_metric="AUC",
                random_seed=RANDOM_STATE,
                verbose=False,
                auto_class_weights="Balanced",
                l2_leaf_reg=8.0,
            )
        except Exception as e:
            print(f"CatBoost skipped: {e}")

    return {
        k: v
        for k, v in models.items()
        if k in RUN_MODELS or k == "Baseline_MajorityClass"
    }


# ============================================================
# 7. DATA SPLITS
# ============================================================

def make_model_data(df, features, target):
    needed = ["date"] + features + [target]
    missing = [c for c in needed if c not in df.columns]

    if missing:
        return pd.DataFrame(), missing

    data = df[needed].copy()
    data = data.replace([np.inf, -np.inf], np.nan)
    data = data.dropna(subset=features + [target])
    data = data.sort_values("date").reset_index(drop=True)

    return data, []


def split_fold(data, fold):
    train = data[
        (data["date"] >= pd.Timestamp(fold["train_start"])) &
        (data["date"] <= pd.Timestamp(fold["train_end"]))
    ].copy()

    test = data[
        (data["date"] >= pd.Timestamp(fold["test_start"])) &
        (data["date"] <= pd.Timestamp(fold["test_end"]))
    ].copy()

    return train, test


def is_data_ok(train, test, target):
    if len(train) < MIN_TRAIN_ROWS:
        return False, "too_few_train_rows"

    if len(test) < MIN_TEST_ROWS:
        return False, "too_few_test_rows"

    if train[target].nunique() < 2:
        return False, "one_class_train"

    if test[target].nunique() < 2:
        return False, "one_class_test"

    return True, ""


# ============================================================
# 8. WALK-FORWARD FULL-COVERAGE CHECK
# ============================================================

def prepare_final_plan(final_plan):
    fp = final_plan.copy()

    if "after_combined_validation_selection_score" in fp.columns:
        fp = fp.sort_values("after_combined_validation_selection_score", ascending=False)

    if TOP_N_REGIMES is not None:
        fp = fp.head(TOP_N_REGIMES).copy()

    return fp.reset_index(drop=True)


def run_walk_forward_full_models(df, final_plan):
    rows = []
    warnings_list = []

    fp = prepare_final_plan(final_plan)

    for _, regime in fp.iterrows():
        h = int(regime["horizon"])
        pct = int(regime["target_threshold_pct"])
        target = target_name(h, pct)

        if target not in df.columns:
            warnings_list.append(f"Target not found: {target}")
            continue

        if "final_features" in regime.index:
            features = parse_feature_list(regime["final_features"])
        else:
            features = SET0_BASE + parse_feature_list(regime.get("final_non_set0_features", ""))

        features = ensure_all_features(df, features, warnings_list)

        if not features:
            warnings_list.append(f"No features for h={h}, thr={pct}")
            continue

        data, missing = make_model_data(df, features, target)

        if missing:
            warnings_list.append(f"Missing columns h={h}, thr={pct}: {missing}")
            continue

        for fold in WALK_FORWARD_FOLDS:
            train, test = split_fold(data, fold)

            ok, reason = is_data_ok(train, test, target)

            if not ok:
                rows.append({
                    "asset": ASSET_NAME,
                    "horizon": h,
                    "target_threshold_pct": pct,
                    "target": target,
                    "fold": fold["fold"],
                    "model": "",
                    "status": "skipped",
                    "skip_reason": reason,
                    "n_features": len(features),
                    "features": ",".join(features),
                    "train_rows": len(train),
                    "test_rows": len(test),
                    "train_class_1_share": train[target].mean() if len(train) else np.nan,
                    "test_class_1_share": test[target].mean() if len(test) else np.nan,
                })
                continue

            X_train = train[features]
            y_train = train[target].astype(int).values

            X_test = test[features]
            y_test = test[target].astype(int).values

            models = get_models(y_train)

            for model_name, model in models.items():
                try:
                    model.fit(X_train, y_train)

                    pred = model.predict(X_test)
                    proba = get_positive_proba(model, X_test)

                    m = calc_metrics(y_test, pred, proba)

                    rows.append({
                        "asset": ASSET_NAME,
                        "horizon": h,
                        "target_threshold_pct": pct,
                        "target": target,
                        "fold": fold["fold"],
                        "model": model_name,
                        "status": "ok",
                        "skip_reason": "",
                        "n_features": len(features),
                        "features": ",".join(features),
                        "train_start": fold["train_start"],
                        "train_end": fold["train_end"],
                        "test_start": fold["test_start"],
                        "test_end": fold["test_end"],
                        "train_rows": len(train),
                        "test_rows": len(test),
                        "train_class_1_share": train[target].mean(),
                        "test_class_1_share": test[target].mean(),
                        **m,
                    })

                except Exception as e:
                    rows.append({
                        "asset": ASSET_NAME,
                        "horizon": h,
                        "target_threshold_pct": pct,
                        "target": target,
                        "fold": fold["fold"],
                        "model": model_name,
                        "status": "error",
                        "skip_reason": str(e),
                        "n_features": len(features),
                        "features": ",".join(features),
                        "train_rows": len(train),
                        "test_rows": len(test),
                    })

        print(f"Done WF full models: h={h}, thr={pct}%, features={len(features)}")

    return pd.DataFrame(rows), warnings_list


def summarize_walk_forward(wf):
    ok = wf[wf["status"] == "ok"].copy()

    if ok.empty:
        return pd.DataFrame()

    gcols = ["horizon", "target_threshold_pct", "model"]

    summary = (
        ok.groupby(gcols)
        .agg(
            n_folds=("fold", "nunique"),
            mean_accuracy=("accuracy", "mean"),
            median_accuracy=("accuracy", "median"),
            min_accuracy=("accuracy", "min"),
            mean_balanced_accuracy=("balanced_accuracy", "mean"),
            median_balanced_accuracy=("balanced_accuracy", "median"),
            min_balanced_accuracy=("balanced_accuracy", "min"),
            mean_f1=("f1", "mean"),
            median_f1=("f1", "median"),
            mean_roc_auc=("roc_auc", "mean"),
            median_roc_auc=("roc_auc", "median"),
            min_roc_auc=("roc_auc", "min"),
            std_roc_auc=("roc_auc", "std"),
            mean_test_rows=("test_rows", "mean"),
            min_test_rows=("test_rows", "min"),
        )
        .reset_index()
    )

    ok["roc_auc_above_0_5"] = ok["roc_auc"] > 0.5
    ok["ba_above_0_5"] = ok["balanced_accuracy"] > 0.5
    ok["acc_above_0_5"] = ok["accuracy"] > 0.5

    rates = (
        ok.groupby(gcols)
        .agg(
            roc_auc_positive_rate=("roc_auc_above_0_5", "mean"),
            ba_positive_rate=("ba_above_0_5", "mean"),
            acc_positive_rate=("acc_above_0_5", "mean"),
        )
        .reset_index()
    )

    summary = summary.merge(rates, on=gcols, how="left")

    summary["wf_selection_score"] = (
        0.55 * summary["mean_roc_auc"].fillna(0) +
        0.45 * summary["mean_balanced_accuracy"].fillna(0)
    )

    summary["stability_flag"] = np.where(
        (
            (summary["n_folds"] >= 3) &
            (summary["mean_roc_auc"] > 0.52) &
            (summary["mean_balanced_accuracy"] > 0.52) &
            (summary["roc_auc_positive_rate"] >= 0.60) &
            (summary["ba_positive_rate"] >= 0.60)
        ),
        "PASS_STABLE",
        np.where(
            (
                (summary["n_folds"] >= 3) &
                (
                    (summary["mean_roc_auc"] > 0.52) |
                    (summary["mean_balanced_accuracy"] > 0.52)
                )
            ),
            "WEAK_REVIEW",
            "FAIL_UNSTABLE",
        )
    )

    summary = summary.sort_values(
        ["stability_flag", "wf_selection_score", "mean_roc_auc"],
        ascending=[True, False, False],
    )

    return summary


# ============================================================
# 9. CONFIDENCE WALK-FORWARD CHECK
# ============================================================

def confidence_metrics(y_true, proba, confidence_threshold):
    y_true = np.asarray(y_true).astype(int)
    proba = np.asarray(proba)

    mask = np.isfinite(proba)
    y_true = y_true[mask]
    proba = proba[mask]

    if len(y_true) == 0:
        return {
            "n_obs": 0,
            "n_signals": 0,
            "coverage": np.nan,
            "signal_accuracy": np.nan,
            "signal_balanced_accuracy": np.nan,
            "signal_f1": np.nan,
            "long_signals": 0,
            "short_signals": 0,
        }

    long_mask = proba >= confidence_threshold
    short_mask = proba <= (1.0 - confidence_threshold)
    signal_mask = long_mask | short_mask

    n_signals = int(signal_mask.sum())
    coverage = n_signals / len(y_true)

    if n_signals == 0:
        return {
            "n_obs": len(y_true),
            "n_signals": 0,
            "coverage": coverage,
            "signal_accuracy": np.nan,
            "signal_balanced_accuracy": np.nan,
            "signal_f1": np.nan,
            "long_signals": 0,
            "short_signals": 0,
        }

    pred_sig = np.where(long_mask[signal_mask], 1, 0)
    y_sig = y_true[signal_mask]

    return {
        "n_obs": len(y_true),
        "n_signals": n_signals,
        "coverage": coverage,
        "signal_accuracy": accuracy_score(y_sig, pred_sig),
        "signal_balanced_accuracy": balanced_accuracy_score(y_sig, pred_sig)
        if len(np.unique(y_sig)) >= 2 else np.nan,
        "signal_f1": f1_score(y_sig, pred_sig, zero_division=0),
        "long_signals": int(long_mask.sum()),
        "short_signals": int(short_mask.sum()),
    }


def run_walk_forward_confidence(df, confidence_selected, final_plan):
    rows = []
    warnings_list = []

    if confidence_selected is None or confidence_selected.empty:
        return pd.DataFrame(), ["No confidence_selected rows found."]

    final_plan_index = {}

    for _, r in final_plan.iterrows():
        key = (int(r["horizon"]), int(r["target_threshold_pct"]))

        if "final_features" in r.index:
            final_plan_index[key] = parse_feature_list(r["final_features"])
        else:
            final_plan_index[key] = SET0_BASE + parse_feature_list(r.get("final_non_set0_features", ""))

    conf = confidence_selected.copy()

    for _, sig in conf.iterrows():
        h = int(sig["horizon"])
        pct = int(sig["target_threshold_pct"])
        target = target_name(h, pct)

        model_name = str(sig["model"])
        conf_thr = float(sig["confidence_threshold"])
        variant = str(sig.get("assurance_variant", ""))

        if model_name not in RUN_MODELS:
            print(f"Skipping confidence model {model_name}, not in RUN_MODELS.")
            continue

        if target not in df.columns:
            warnings_list.append(f"Target not found for confidence: {target}")
            continue

        key = (h, pct)

        if key not in final_plan_index:
            warnings_list.append(f"No final features for confidence key: {key}")
            continue

        features = ensure_all_features(df, final_plan_index[key], warnings_list)

        data, missing = make_model_data(df, features, target)

        if missing:
            warnings_list.append(f"Missing confidence columns h={h}, thr={pct}: {missing}")
            continue

        for fold in WALK_FORWARD_FOLDS:
            train, test = split_fold(data, fold)

            ok, reason = is_data_ok(train, test, target)

            if not ok:
                rows.append({
                    "asset": ASSET_NAME,
                    "horizon": h,
                    "target_threshold_pct": pct,
                    "target": target,
                    "fold": fold["fold"],
                    "model": model_name,
                    "assurance_variant": variant,
                    "confidence_threshold": conf_thr,
                    "status": "skipped",
                    "skip_reason": reason,
                    "train_rows": len(train),
                    "test_rows": len(test),
                    "n_features": len(features),
                    "features": ",".join(features),
                })
                continue

            X_train = train[features]
            y_train = train[target].astype(int).values

            X_test = test[features]
            y_test = test[target].astype(int).values

            models = get_models(y_train)

            if model_name not in models:
                rows.append({
                    "asset": ASSET_NAME,
                    "horizon": h,
                    "target_threshold_pct": pct,
                    "target": target,
                    "fold": fold["fold"],
                    "model": model_name,
                    "assurance_variant": variant,
                    "confidence_threshold": conf_thr,
                    "status": "skipped",
                    "skip_reason": "model_not_available",
                    "train_rows": len(train),
                    "test_rows": len(test),
                    "n_features": len(features),
                    "features": ",".join(features),
                })
                continue

            model = models[model_name]

            try:
                model.fit(X_train, y_train)
                proba = get_positive_proba(model, X_test)

                m = confidence_metrics(y_test, proba, conf_thr)

                rows.append({
                    "asset": ASSET_NAME,
                    "horizon": h,
                    "target_threshold_pct": pct,
                    "target": target,
                    "fold": fold["fold"],
                    "model": model_name,
                    "assurance_variant": variant,
                    "confidence_threshold": conf_thr,
                    "status": "ok",
                    "skip_reason": "",
                    "train_start": fold["train_start"],
                    "train_end": fold["train_end"],
                    "test_start": fold["test_start"],
                    "test_end": fold["test_end"],
                    "train_rows": len(train),
                    "test_rows": len(test),
                    "train_class_1_share": train[target].mean(),
                    "test_class_1_share": test[target].mean(),
                    "n_features": len(features),
                    "features": ",".join(features),
                    **m,
                })

            except Exception as e:
                rows.append({
                    "asset": ASSET_NAME,
                    "horizon": h,
                    "target_threshold_pct": pct,
                    "target": target,
                    "fold": fold["fold"],
                    "model": model_name,
                    "assurance_variant": variant,
                    "confidence_threshold": conf_thr,
                    "status": "error",
                    "skip_reason": str(e),
                    "train_rows": len(train),
                    "test_rows": len(test),
                    "n_features": len(features),
                    "features": ",".join(features),
                })

        print(f"Done WF confidence: h={h}, thr={pct}%, model={model_name}, conf={conf_thr}")

    return pd.DataFrame(rows), warnings_list


def summarize_confidence_wf(conf_wf):
    ok = conf_wf[conf_wf["status"] == "ok"].copy()

    if ok.empty:
        return pd.DataFrame()

    gcols = [
        "horizon",
        "target_threshold_pct",
        "model",
        "assurance_variant",
        "confidence_threshold",
    ]

    summary = (
        ok.groupby(gcols)
        .agg(
            n_folds=("fold", "nunique"),
            mean_signal_accuracy=("signal_accuracy", "mean"),
            median_signal_accuracy=("signal_accuracy", "median"),
            min_signal_accuracy=("signal_accuracy", "min"),
            mean_signal_balanced_accuracy=("signal_balanced_accuracy", "mean"),
            median_signal_balanced_accuracy=("signal_balanced_accuracy", "median"),
            mean_coverage=("coverage", "mean"),
            median_coverage=("coverage", "median"),
            min_coverage=("coverage", "min"),
            total_signals=("n_signals", "sum"),
            mean_signals_per_fold=("n_signals", "mean"),
            min_signals_per_fold=("n_signals", "min"),
            mean_long_signals=("long_signals", "mean"),
            mean_short_signals=("short_signals", "mean"),
        )
        .reset_index()
    )

    ok["signal_acc_above_0_5"] = ok["signal_accuracy"] > 0.5
    ok["coverage_above_0_05"] = ok["coverage"] >= 0.05

    rates = (
        ok.groupby(gcols)
        .agg(
            signal_acc_positive_rate=("signal_acc_above_0_5", "mean"),
            coverage_005_rate=("coverage_above_0_05", "mean"),
        )
        .reset_index()
    )

    summary = summary.merge(rates, on=gcols, how="left")

    summary["confidence_wf_score"] = (
        0.55 * summary["mean_signal_accuracy"].fillna(0) +
        0.25 * summary["mean_signal_balanced_accuracy"].fillna(0) +
        0.20 * summary["mean_coverage"].fillna(0)
    )

    summary["confidence_stability_flag"] = np.where(
        (
            (summary["n_folds"] >= 3) &
            (summary["mean_signal_accuracy"] > 0.55) &
            (summary["signal_acc_positive_rate"] >= 0.60) &
            (summary["mean_coverage"] >= 0.05) &
            (summary["total_signals"] >= 30)
        ),
        "PASS_STABLE_SIGNAL",
        np.where(
            (
                (summary["n_folds"] >= 3) &
                (summary["mean_signal_accuracy"] > 0.53) &
                (summary["total_signals"] >= 20)
            ),
            "WEAK_REVIEW_SIGNAL",
            "FAIL_UNSTABLE_SIGNAL",
        )
    )

    summary = summary.sort_values(
        ["confidence_stability_flag", "confidence_wf_score"],
        ascending=[True, False],
    )

    return summary


# ============================================================
# 10. COMPARISON WITH STAGE 2 VALIDATION / TEST
# ============================================================

def make_stage2_vs_wf_comparison(final_all_models, wf_summary):
    if final_all_models is None or final_all_models.empty:
        return pd.DataFrame()

    if wf_summary is None or wf_summary.empty:
        return pd.DataFrame()

    s2 = final_all_models.copy()

    needed_cols = [
        "horizon",
        "target_threshold_pct",
        "model",
        "eval_split",
        "accuracy",
        "balanced_accuracy",
        "f1",
        "roc_auc",
    ]

    missing = [c for c in needed_cols if c not in s2.columns]

    if missing:
        print("WARNING: Stage 2 final_all_models missing columns:", missing)
        return pd.DataFrame()

    s2 = s2[
        (s2["eval_split"].isin(["validation", "test"])) &
        (s2["model"].isin(RUN_MODELS))
    ].copy()

    if s2.empty:
        return pd.DataFrame()

    pivot = (
        s2.groupby(["horizon", "target_threshold_pct", "model", "eval_split"])
        [["accuracy", "balanced_accuracy", "f1", "roc_auc"]]
        .mean()
        .reset_index()
    )

    pivot = pivot.pivot_table(
        index=["horizon", "target_threshold_pct", "model"],
        columns="eval_split",
        values=["accuracy", "balanced_accuracy", "f1", "roc_auc"],
        aggfunc="mean"
    )

    pivot.columns = [f"stage2_{metric}_{split}" for metric, split in pivot.columns]
    pivot = pivot.reset_index()

    out = wf_summary.merge(
        pivot,
        on=["horizon", "target_threshold_pct", "model"],
        how="left"
    )

    for col in [
        "stage2_roc_auc_validation",
        "stage2_roc_auc_test",
        "stage2_balanced_accuracy_validation",
        "stage2_balanced_accuracy_test",
    ]:
        if col not in out.columns:
            out[col] = np.nan

    out["gap_stage2_validation_auc_minus_wf_auc"] = (
        out["stage2_roc_auc_validation"] - out["mean_roc_auc"]
    )

    out["gap_stage2_test_auc_minus_wf_auc"] = (
        out["stage2_roc_auc_test"] - out["mean_roc_auc"]
    )

    out["gap_stage2_validation_ba_minus_wf_ba"] = (
        out["stage2_balanced_accuracy_validation"] - out["mean_balanced_accuracy"]
    )

    out["gap_stage2_test_ba_minus_wf_ba"] = (
        out["stage2_balanced_accuracy_test"] - out["mean_balanced_accuracy"]
    )

    return out


# ============================================================
# 11. PLOTS
# ============================================================

def save_bar_top(df, value_col, label_col, title, filename, top_n=20):
    if df is None or df.empty or value_col not in df.columns:
        return

    d = df.dropna(subset=[value_col]).copy()
    d = d.sort_values(value_col, ascending=False).head(top_n)

    if d.empty:
        return

    plt.figure(figsize=(14, 8))
    plt.barh(d[label_col][::-1], d[value_col][::-1])
    plt.title(title)
    plt.xlabel(value_col)
    plt.tight_layout()
    plt.savefig(PLOTS_DIR / filename, dpi=250)
    plt.close()


def make_plots(wf_summary, conf_summary, comparison):
    created = []

    if wf_summary is not None and not wf_summary.empty:
        d = wf_summary.copy()
        d["regime_model"] = (
            "h=" + d["horizon"].astype(str) +
            ",thr=" + d["target_threshold_pct"].astype(str) +
            "," + d["model"].astype(str)
        )

        save_bar_top(
            d,
            "mean_roc_auc",
            "regime_model",
            "LINK Stage 3.0 walk-forward: top regimes by mean ROC-AUC",
            "wf_top_mean_roc_auc.png",
        )
        created.append("wf_top_mean_roc_auc.png")

        save_bar_top(
            d,
            "mean_balanced_accuracy",
            "regime_model",
            "LINK Stage 3.0 walk-forward: top regimes by mean balanced accuracy",
            "wf_top_mean_balanced_accuracy.png",
        )
        created.append("wf_top_mean_balanced_accuracy.png")

        heat_data = (
            d.groupby(["horizon", "target_threshold_pct"])["mean_roc_auc"]
            .max()
            .reset_index()
        )

        if not heat_data.empty:
            heat = heat_data.pivot(
                index="horizon",
                columns="target_threshold_pct",
                values="mean_roc_auc"
            )

            plt.figure(figsize=(10, 6))
            plt.imshow(heat.values, aspect="auto")
            plt.xticks(range(len(heat.columns)), heat.columns)
            plt.yticks(range(len(heat.index)), heat.index)
            plt.colorbar(label="max mean ROC-AUC")
            plt.title("LINK Stage 3.0 WF heatmap: horizon x threshold")
            plt.xlabel("target_threshold_pct")
            plt.ylabel("horizon")
            plt.tight_layout()
            plt.savefig(PLOTS_DIR / "wf_heatmap_horizon_threshold_auc.png", dpi=250)
            plt.close()
            created.append("wf_heatmap_horizon_threshold_auc.png")

        flag_counts = d["stability_flag"].value_counts().reset_index()
        flag_counts.columns = ["stability_flag", "count"]

        plt.figure(figsize=(9, 5))
        plt.bar(flag_counts["stability_flag"], flag_counts["count"])
        plt.title("LINK Stage 3.0 stability flags")
        plt.xlabel("stability flag")
        plt.ylabel("count")
        plt.tight_layout()
        plt.savefig(PLOTS_DIR / "wf_stability_flags.png", dpi=250)
        plt.close()
        created.append("wf_stability_flags.png")

    if conf_summary is not None and not conf_summary.empty:
        c = conf_summary.copy()
        c["regime_model"] = (
            "h=" + c["horizon"].astype(str) +
            ",thr=" + c["target_threshold_pct"].astype(str) +
            "," + c["model"].astype(str) +
            ",conf=" + c["confidence_threshold"].astype(str)
        )

        save_bar_top(
            c,
            "mean_signal_accuracy",
            "regime_model",
            "LINK Stage 3.0 confidence: top regimes by signal accuracy",
            "confidence_top_mean_signal_accuracy.png",
        )
        created.append("confidence_top_mean_signal_accuracy.png")

        plt.figure(figsize=(10, 7))
        plt.scatter(c["mean_coverage"], c["mean_signal_accuracy"])

        for _, row in c.iterrows():
            label = f"{int(row['horizon'])}/{int(row['target_threshold_pct'])}"
            plt.text(row["mean_coverage"], row["mean_signal_accuracy"], label, fontsize=8)

        plt.axhline(0.55, linewidth=1)
        plt.axvline(0.05, linewidth=1)
        plt.title("LINK Stage 3.0 confidence: mean signal accuracy vs coverage")
        plt.xlabel("mean coverage")
        plt.ylabel("mean signal accuracy")
        plt.tight_layout()
        plt.savefig(PLOTS_DIR / "confidence_accuracy_vs_coverage.png", dpi=250)
        plt.close()
        created.append("confidence_accuracy_vs_coverage.png")

    if comparison is not None and not comparison.empty:
        comp = comparison.dropna(subset=["stage2_roc_auc_validation", "mean_roc_auc"]).copy()

        if not comp.empty:
            plt.figure(figsize=(9, 7))
            plt.scatter(comp["stage2_roc_auc_validation"], comp["mean_roc_auc"])
            plt.axline((0.5, 0.5), slope=1, linewidth=1)
            plt.title("LINK Stage 2 validation ROC-AUC vs Stage 3 walk-forward ROC-AUC")
            plt.xlabel("Stage 2 validation ROC-AUC")
            plt.ylabel("Stage 3 WF mean ROC-AUC")
            plt.tight_layout()
            plt.savefig(PLOTS_DIR / "stage2_validation_vs_wf_auc.png", dpi=250)
            plt.close()
            created.append("stage2_validation_vs_wf_auc.png")

    return created


# ============================================================
# 12. SAVE OUTPUTS
# ============================================================

def save_excel(
    final_plan,
    confidence_selected,
    final_all_models,
    wf,
    wf_summary,
    conf_wf,
    conf_summary,
    comparison,
    recommended,
    warnings_list,
):
    xlsx_path = OUTPUT_DIR / "LINK_STAGE_3_0_WALK_FORWARD_FINAL_CHECK.xlsx"

    readme = pd.DataFrame({
        "section": [
            "Описание",
            "Главное правило",
            "Актив",
            "Старт данных",
            "Что проверяется",
            "Walk-forward",
            "Признаки",
            "Модели",
            "Confidence",
            "Ограничение",
        ],
        "text": [
            "Stage 3.0 — заключительная проверка устойчивости финальных режимов Stage 2.1 для LINK.",
            "Признаки НЕ подбираются заново. Используются final_features из 2_1_J_final_plan_by_horizon_threshold.csv.",
            ASSET_NAME,
            f"LINK sample starts around {LINK_SAMPLE_START}.",
            "Проверяется переносимость режимов во времени: train расширяется, следующий год используется как test.",
            "Folds: 2022, 2023, 2024, 2025, 2026_Q1.",
            "Lag/regime/interaction-признаки пересоздаются по тем же правилам, что в Stage 2.1.",
            f"Проверяемые модели: {', '.join(RUN_MODELS)}.",
            "Confidence-сигналы проверяются по тем же confidence_threshold, которые были выбраны на validation в Stage 2.1.",
            "Это не доказывает торговую прибыльность. Это проверка качества ML-сигнала до торгового backtest. Для LINK fees_log_return может быть ETH gas proxy.",
        ]
    })

    folds_df = pd.DataFrame(WALK_FORWARD_FOLDS)
    warnings_df = pd.DataFrame({"warning": warnings_list})

    with pd.ExcelWriter(xlsx_path, engine="openpyxl") as writer:
        readme.to_excel(writer, sheet_name="README", index=False)
        folds_df.to_excel(writer, sheet_name="folds", index=False)
        final_plan.to_excel(writer, sheet_name="stage2_final_plan_used", index=False)

        if confidence_selected is not None and not confidence_selected.empty:
            confidence_selected.to_excel(writer, sheet_name="stage2_conf_selected", index=False)

        if final_all_models is not None and not final_all_models.empty:
            final_all_models.to_excel(writer, sheet_name="stage2_final_models", index=False)

        wf.to_excel(writer, sheet_name="wf_all_rows", index=False)
        wf_summary.to_excel(writer, sheet_name="wf_summary", index=False)

        if conf_wf is not None and not conf_wf.empty:
            conf_wf.to_excel(writer, sheet_name="confidence_wf_rows", index=False)

        if conf_summary is not None and not conf_summary.empty:
            conf_summary.to_excel(writer, sheet_name="confidence_wf_summary", index=False)

        if comparison is not None and not comparison.empty:
            comparison.to_excel(writer, sheet_name="stage2_vs_wf", index=False)

        if recommended is not None and not recommended.empty:
            recommended.to_excel(writer, sheet_name="recommended_final", index=False)

        warnings_df.to_excel(writer, sheet_name="warnings", index=False)

    try:
        from openpyxl import load_workbook
        from openpyxl.styles import Font

        wb = load_workbook(xlsx_path)

        for ws in wb.worksheets:
            ws.freeze_panes = "A2"
            ws.auto_filter.ref = ws.dimensions

            for cell in ws[1]:
                cell.font = Font(bold=True)

            for col in ws.columns:
                max_len = 0
                col_letter = col[0].column_letter

                for cell in col[:300]:
                    try:
                        max_len = max(max_len, len(str(cell.value)))
                    except Exception:
                        pass

                ws.column_dimensions[col_letter].width = min(max_len + 2, 45)

        wb.save(xlsx_path)

    except Exception as e:
        print(f"Excel formatting skipped: {e}")

    return xlsx_path


def make_recommended_final(wf_summary, conf_summary):
    parts = []

    if wf_summary is not None and not wf_summary.empty:
        full = wf_summary.copy()
        full = full[full["stability_flag"].isin(["PASS_STABLE", "WEAK_REVIEW"])].copy()
        full["recommendation_type"] = "full_coverage_model"
        full["main_score"] = full["wf_selection_score"]

        parts.append(full)

    if conf_summary is not None and not conf_summary.empty:
        sig = conf_summary.copy()
        sig = sig[sig["confidence_stability_flag"].isin(["PASS_STABLE_SIGNAL", "WEAK_REVIEW_SIGNAL"])].copy()
        sig["recommendation_type"] = "confidence_signal"
        sig["main_score"] = sig["confidence_wf_score"]

        parts.append(sig)

    if not parts:
        return pd.DataFrame()

    out = pd.concat(parts, ignore_index=True, sort=False)
    out = out.sort_values("main_score", ascending=False)

    return out


def write_summary_txt(wf_summary, conf_summary, recommended, warnings_list, plot_files, excel_path):
    path = OUTPUT_DIR / "LINK_STAGE_3_0_SUMMARY.txt"

    with open(path, "w", encoding="utf-8") as f:
        f.write("LINK STAGE 3.0 — WALK-FORWARD FINAL CHECK\n")
        f.write("=" * 80 + "\n\n")

        f.write("Главное:\n")
        f.write("- Признаки НЕ подбирались заново.\n")
        f.write("- Использованы final_features из LINK Stage 2.1.\n")
        f.write("- Проверка сделана через расширяющееся walk-forward окно.\n")
        f.write("- Test внутри каждого fold — следующий год.\n")
        f.write(f"- LINK sample starts around {LINK_SAMPLE_START}.\n")
        f.write("- Для LINK fees_log_return может быть ETH gas proxy, если так было собрано в базе.\n\n")

        f.write("Input:\n")
        f.write(f"- file: {FILE_PATH}\n")
        f.write(f"- sheet: {SHEET_NAME}\n")
        f.write(f"- stage2_dir: {STAGE2_DIR}\n\n")

        f.write("Проверенные модели:\n")
        for m in RUN_MODELS:
            f.write(f"- {m}\n")

        f.write("\nTop full-coverage regimes by WF score:\n")
        if wf_summary is not None and not wf_summary.empty:
            top = wf_summary.sort_values("wf_selection_score", ascending=False).head(15)

            for _, r in top.iterrows():
                f.write(
                    f"h={int(r['horizon'])}, thr={int(r['target_threshold_pct'])}%, "
                    f"model={r['model']}, flag={r['stability_flag']}, "
                    f"mean_auc={r['mean_roc_auc']:.4f}, "
                    f"mean_ba={r['mean_balanced_accuracy']:.4f}, "
                    f"positive_auc_rate={r['roc_auc_positive_rate']:.2f}\n"
                )
        else:
            f.write("No WF summary.\n")

        f.write("\nTop confidence regimes by WF score:\n")
        if conf_summary is not None and not conf_summary.empty:
            topc = conf_summary.sort_values("confidence_wf_score", ascending=False).head(15)

            for _, r in topc.iterrows():
                f.write(
                    f"h={int(r['horizon'])}, thr={int(r['target_threshold_pct'])}%, "
                    f"model={r['model']}, variant={r['assurance_variant']}, "
                    f"conf={r['confidence_threshold']}, "
                    f"flag={r['confidence_stability_flag']}, "
                    f"mean_signal_acc={r['mean_signal_accuracy']:.4f}, "
                    f"mean_cov={r['mean_coverage']:.4f}, "
                    f"total_signals={int(r['total_signals'])}\n"
                )
        else:
            f.write("No confidence WF summary.\n")

        f.write("\nRecommended final candidates:\n")
        if recommended is not None and not recommended.empty:
            top_rec = recommended.head(20)

            for _, r in top_rec.iterrows():
                rec_type = r.get("recommendation_type", "")

                if rec_type == "full_coverage_model":
                    f.write(
                        f"[FULL] h={int(r['horizon'])}, thr={int(r['target_threshold_pct'])}%, "
                        f"model={r['model']}, flag={r['stability_flag']}, "
                        f"score={r['main_score']:.4f}\n"
                    )
                else:
                    f.write(
                        f"[SIGNAL] h={int(r['horizon'])}, thr={int(r['target_threshold_pct'])}%, "
                        f"model={r['model']}, variant={r.get('assurance_variant', '')}, "
                        f"flag={r.get('confidence_stability_flag', '')}, "
                        f"score={r['main_score']:.4f}\n"
                    )
        else:
            f.write("No recommended candidates passed stability filters.\n")

        f.write("\nCreated files:\n")
        f.write(f"- {excel_path}\n")
        f.write("- 3_0_walk_forward_all_rows.csv\n")
        f.write("- 3_0_walk_forward_summary.csv\n")
        f.write("- 3_0_confidence_walk_forward_rows.csv\n")
        f.write("- 3_0_confidence_walk_forward_summary.csv\n")
        f.write("- 3_0_stage2_vs_walk_forward_comparison.csv\n")
        f.write("- 3_0_recommended_final_candidates.csv\n")

        f.write("\nPlots:\n")
        for p in plot_files:
            f.write(f"- {p}\n")

        f.write("\nWarnings:\n")
        for w in warnings_list:
            f.write(f"- {w}\n")

    return path


# ============================================================
# 13. MAIN
# ============================================================

def main():
    print("=" * 80)
    print("LINK Stage 3.0 — Walk-Forward Final Check")
    print("=" * 80)

    print(f"Input Excel: {FILE_PATH}")
    print(f"Sheet: {SHEET_NAME}")
    print(f"Stage 2 dir: {STAGE2_DIR}")
    print(f"Output dir: {OUTPUT_DIR}")
    print(f"LINK sample start: {LINK_SAMPLE_START}")

    df = load_ml_data()

    final_plan, confidence_selected, final_all_models = load_stage2_files()

    print("\nLoaded Stage 2 files:")
    print(f"final_plan rows: {len(final_plan)}")
    print(f"confidence_selected rows: {len(confidence_selected)}")
    print(f"final_all_models rows: {len(final_all_models)}")

    final_plan_used = prepare_final_plan(final_plan)

    print("\nRunning walk-forward full models...")
    wf, warnings_full = run_walk_forward_full_models(df, final_plan_used)

    print("\nSummarizing walk-forward full models...")
    wf_summary = summarize_walk_forward(wf)

    print("\nRunning walk-forward confidence signals...")
    conf_wf, warnings_conf = run_walk_forward_confidence(df, confidence_selected, final_plan_used)

    print("\nSummarizing confidence walk-forward...")
    conf_summary = summarize_confidence_wf(conf_wf)

    print("\nComparing Stage 2 vs Stage 3...")
    comparison = make_stage2_vs_wf_comparison(final_all_models, wf_summary)

    print("\nCreating final recommendations...")
    recommended = make_recommended_final(wf_summary, conf_summary)

    warnings_list = warnings_full + warnings_conf

    print("\nSaving CSV outputs...")
    wf.to_csv(OUTPUT_DIR / "3_0_walk_forward_all_rows.csv", index=False, encoding="utf-8-sig")
    wf_summary.to_csv(OUTPUT_DIR / "3_0_walk_forward_summary.csv", index=False, encoding="utf-8-sig")
    conf_wf.to_csv(OUTPUT_DIR / "3_0_confidence_walk_forward_rows.csv", index=False, encoding="utf-8-sig")
    conf_summary.to_csv(OUTPUT_DIR / "3_0_confidence_walk_forward_summary.csv", index=False, encoding="utf-8-sig")
    comparison.to_csv(OUTPUT_DIR / "3_0_stage2_vs_walk_forward_comparison.csv", index=False, encoding="utf-8-sig")
    recommended.to_csv(OUTPUT_DIR / "3_0_recommended_final_candidates.csv", index=False, encoding="utf-8-sig")

    print("\nMaking plots...")
    plot_files = make_plots(wf_summary, conf_summary, comparison)

    print("\nSaving Excel report...")
    excel_path = save_excel(
        final_plan_used,
        confidence_selected,
        final_all_models,
        wf,
        wf_summary,
        conf_wf,
        conf_summary,
        comparison,
        recommended,
        warnings_list,
    )

    print("\nWriting text summary...")
    summary_path = write_summary_txt(
        wf_summary,
        conf_summary,
        recommended,
        warnings_list,
        plot_files,
        excel_path,
    )

    print("\nDONE.")
    print(f"Excel report: {excel_path}")
    print(f"Summary: {summary_path}")
    print(f"Plots dir: {PLOTS_DIR}")

    print("\nMain outputs:")
    print("- 3_0_walk_forward_all_rows.csv")
    print("- 3_0_walk_forward_summary.csv")
    print("- 3_0_confidence_walk_forward_rows.csv")
    print("- 3_0_confidence_walk_forward_summary.csv")
    print("- 3_0_stage2_vs_walk_forward_comparison.csv")
    print("- 3_0_recommended_final_candidates.csv")
    print("- LINK_STAGE_3_0_WALK_FORWARD_FINAL_CHECK.xlsx")
    print("- LINK_STAGE_3_0_SUMMARY.txt")

    if recommended is not None and not recommended.empty:
        print("\nTop recommended candidates:")
        cols = [
            "recommendation_type",
            "horizon",
            "target_threshold_pct",
            "model",
            "main_score",
        ]

        existing = [c for c in cols if c in recommended.columns]
        print(recommended[existing].head(20))
    else:
        print("\nNo candidates passed stability filters.")

    if warnings_list:
        print("\nWarnings:")
        for w in warnings_list[:30]:
            print("-", w)

        if len(warnings_list) > 30:
            print(f"... and {len(warnings_list) - 30} more warnings.")


if __name__ == "__main__":
    main()
